import streamlit as st
import pandas as pd
import json, os, io, base64, traceback, threading
from datetime import datetime, timedelta

# ── Módulo RPA (importação segura — falha silenciosa se libs ausentes) ────────
try:
    import robot_captura as _robot
    ROBOT_DISPONIVEL = True
except Exception as _robot_err:
    ROBOT_DISPONIVEL = False
    _robot_err_msg = str(_robot_err)

try:
    import rateio_sunne_bot as _rateio_bot
    RATEIO_BOT_OK = True
except Exception as _rateio_bot_err:
    RATEIO_BOT_OK = False
    _rateio_bot_err_msg = str(_rateio_bot_err)

# ── PAGE CONFIG ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sunne · Hub Operacional v12",
    page_icon="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>☀️</text></svg>",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS PREMIUM SUNNE ─────────────────────────────────────────────────────────
SUNNE_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Geist:wght@300;400;500;600;700&family=Geist+Mono:wght@400;500&family=Syne:wght@600;700;800&display=swap');

/* ════════════════════════════════════════════════════════════════
   SUNNE HUB v12 — Design System
   Aesthetic: Pro Dark · Linear/Vercel-grade · Precision data UI
   ════════════════════════════════════════════════════════════════ */

:root {
  /* ── Core surfaces ── */
  --bg:          #FDF8F5;
  --bg-2:        #ffffff;
  --bg-3:    #32001a;
  --surface:     #13131C;
  --surface-2:   #18182A;
  --surface-3:   #1E1E30;
  --overlay:     rgba(255,255,255,.03);

  /* ── Brand ── */
  --orange:      #F36E21;
  --orange-d:    #D45E14;
  --orange-glow: rgba(243,110,33,.18);
  --orange-dim:  rgba(243,110,33,.10);
  --orange-xs:   rgba(243,110,33,.06);

  /* ── Accent ── */
  --accent:      #6366F1;
  --accent-d:    #4F52D9;
  --accent-dim:  rgba(99,102,241,.12);

  /* ── Text ── */
  --text-1:      #F59E0B;
  --text-2:      #ffffff;
  --text-3:      #606080;
  --text-4:      #FDF8F5;

  /* ── Borders ── */
  --border:      rgba(255,255,255,.06);
  --border-m:    rgba(255,255,255,.10);
  --border-h:    rgba(255,255,255,.16);

  /* ── Status ── */
  --red:         #F43F5E;
  --red-dim:     rgba(244,63,94,.12);
  --amber:       #F59E0B;
  --amber-dim:   rgba(245,158,11,.12);
  --green:       #10B981;
  --green-dim:   rgba(16,185,129,.12);
  --blue:        #3B82F6;
  --blue-dim:    rgba(59,130,246,.12);

  /* ── Shadows ── */
  --shadow-s:    0 1px 3px rgba(0,0,0,.4);
  --shadow:      0 4px 16px rgba(0,0,0,.5);
  --shadow-m:    0 8px 32px rgba(0,0,0,.6);
  --shadow-l:    0 20px 60px rgba(0,0,0,.7);
  --glow-o:      0 0 24px rgba(243,110,33,.22);
  --glow-a:      0 0 24px rgba(99,102,241,.18);

  /* ── Radii ── */
  --r-xs: 4px;
  --r-s:  6px;
  --r:    10px;
  --r-l:  14px;
  --r-xl: 20px;
}

/* ── Reset & base ───────────────────────────────────────────────── */
html, body, [class*="css"] {
  font-family: 'Geist', 'SF Pro Display', -apple-system, sans-serif !important;
  color: var(--text-1) !important;
  -webkit-font-smoothing: antialiased !important;
  text-rendering: optimizeLegibility !important;
}
[data-testid="stAppViewContainer"] { background: var(--bg) !important; }
[data-testid="stMain"]             { background: var(--bg) !important; }
#MainMenu, footer, header          { visibility: hidden; }
.block-container {
  padding: 2rem 2.5rem 4rem !important;
  max-width: 1440px !important;
}

/* ── Scrollbar ──────────────────────────────────────────────────── */
::-webkit-scrollbar              { width: 4px; height: 4px; }
::-webkit-scrollbar-track        { background: #32001a; }
::-webkit-scrollbar-thumb        { background: var(--border-m); border-radius: 4px; }
::-webkit-scrollbar-thumb:hover  { background: var(--border-h); }

/* ════════════════════════════════════════════════════════════════
   SIDEBAR
   ════════════════════════════════════════════════════════════════ */
[data-testid="stSidebar"] {
  background: var(--bg-3) !important;
  border-right: 1px solid var(--border) !important;
  box-shadow: none !important;
}
[data-testid="stSidebar"] > div { padding: 0 !important; }

[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
  background:  #32001a !important;
  border: none !important;
  padding: 0 !important;
  margin-bottom: 1px !important;
  width: 100% !important;
  box-shadow: none !important;
  border-radius: var(--r-s) !important;
  transition: background .12s ease !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"]:hover {
  background: var(--overlay) !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] p {
  color: var(--text-2) !important;
  font-size: 13px !important;
  font-weight: 400 !important;
  letter-spacing: -.01em !important;
  padding: 5px 4px !important;
  transition: color .12s !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"]:hover p {
  color: var(--text-1) !important;
}

/* ── Sidebar components ─────────────────────────────────────────── */
.sb-brand {
  padding: 1.5rem 1.2rem .9rem;
  border-bottom: 1px solid var(--border);
  margin-bottom: .4rem;
  display: flex; align-items: center; gap: 10px;
}
.sb-brand-icon {
  width: 32px; height: 32px;
  background: linear-gradient(135deg, var(--orange), var(--orange-d));
  border-radius: var(--r-s);
  display: flex; align-items: center; justify-content: center;
  font-size: 16px;
  box-shadow: var(--glow-o);
  flex-shrink: 0;
}
.sb-brand-wrap { display: flex; flex-direction: column; }
.sb-brand-name {
  font-family: 'Syne', sans-serif;
  font-size: 1rem; font-weight: 700;
  color: var(--text-1);
  letter-spacing: -.02em;
  line-height: 1;
}
.sb-brand-sub {
  font-size: 9px; color: var(--text-3);
  letter-spacing: .12em; text-transform: uppercase; margin-top: 2px;
}
.sb-user {
  margin: 0 .7rem .6rem;
  background: var(--overlay);
  border-radius: var(--r);
  padding: .6rem .8rem;
  border: 1px solid var(--border);
}
.sb-user-name  { font-size: 12.5px; color: var(--text-1); font-weight: 500; }
.sb-user-role  { font-size: 9.5px; color: var(--text-3); text-transform: uppercase; letter-spacing: .08em; margin-top: 1px; }
.sb-google-badge {
  margin: 0 .7rem .4rem;
  background: rgba(16,185,129,.08);
  border: 1px solid rgba(16,185,129,.18);
  border-radius: var(--r-s); padding: .35rem .7rem;
  font-size: 11px; color: var(--green);
}
.sec-label {
  font-size: 9px; font-weight: 700; letter-spacing: .16em;
  text-transform: uppercase; color: var(--text-4);
  margin-bottom: 2px; margin-top: 14px; padding-left: .9rem;
}
.sb-divider { height: 1px; background: var(--border); margin: .4rem .7rem; }

/* ════════════════════════════════════════════════════════════════
   TYPOGRAPHY
   ════════════════════════════════════════════════════════════════ */
h1 {
  font-family: 'Syne', sans-serif !important;
  font-size: 1.65rem !important; font-weight: 700 !important;
  color: var(--text-1) !important;
  letter-spacing: -.03em !important;
  margin-bottom: .1rem !important; line-height: 1.1 !important;
}
h2 {
  font-size: 1rem !important; font-weight: 600 !important;
  color: var(--text-1) !important; letter-spacing: -.02em !important;
}
h3 {
  font-size: .875rem !important; font-weight: 600 !important;
  color: var(--text-2) !important;
}

/* ════════════════════════════════════════════════════════════════
   BUTTONS
   ════════════════════════════════════════════════════════════════ */
.stButton > button {
  background: var(--orange) !important;
  color: #fff !important; border: none !important;
  border-radius: var(--r-s) !important;
  font-weight: 600 !important; font-size: 13px !important;
  padding: .5rem 1.2rem !important;
  letter-spacing: -.01em !important;
  box-shadow: 0 1px 0 rgba(0,0,0,.3), var(--glow-o) !important;
  transition: all .14s ease !important;
  position: relative !important; overflow: hidden !important;
}
.stButton > button::after {
  content: ''; position: absolute; inset: 0;
  background: linear-gradient(180deg, rgba(255,255,255,.1) 0%, transparent 100%);
  pointer-events: none;
}
.stButton > button:hover {
  background: var(--orange-d) !important;
  transform: translateY(-1px) !important;
  box-shadow: 0 4px 16px rgba(243,110,33,.4) !important;
}
.stButton > button:active { transform: translateY(0) !important; }

.stDownloadButton > button {
  background: var(--surface-2) !important;
  color: var(--text-2) !important;
  border: 1px solid var(--border-m) !important;
  border-radius: var(--r-s) !important;
  font-size: 12.5px !important; font-weight: 500 !important;
  padding: .42rem 1rem !important;
  transition: all .14s !important;
}
.stDownloadButton > button:hover {
  background: var(--surface-3) !important;
  color: var(--text-1) !important;
  border-color: var(--border-h) !important;
}

/* ════════════════════════════════════════════════════════════════
   INPUTS
   ════════════════════════════════════════════════════════════════ */
[data-testid="stTextInput"] input,
[data-testid="stSelectbox"] > div,
[data-testid="stNumberInput"] input,
[data-testid="stTextArea"] textarea {
  background: var(--surface-2) !important;
  border: 1px solid var(--border-m) !important;
  border-radius: var(--r-s) !important;
  color: var(--text-1) !important;
  font-size: 13px !important;
  transition: border-color .14s, box-shadow .14s !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stTextArea"] textarea:focus {
  border-color: var(--orange) !important;
  box-shadow: 0 0 0 3px var(--orange-dim) !important;
  outline: none !important;
}
[data-testid="stSelectbox"] svg { color: var(--text-3) !important; }

/* Labels */
[data-testid="stTextInput"] label,
[data-testid="stSelectbox"] label,
[data-testid="stNumberInput"] label,
[data-testid="stTextArea"] label,
[data-testid="stFileUploader"] label {
  font-size: 11.5px !important; font-weight: 500 !important;
  color: var(--text-3) !important;
  letter-spacing: .04em !important; text-transform: uppercase !important;
}

/* File uploader */
[data-testid="stFileUploader"] > div {
  background: var(--surface-2) !important;
  border: 1px dashed var(--border-m) !important;
  border-radius: var(--r) !important;
  transition: border-color .14s !important;
}
[data-testid="stFileUploader"] > div:hover {
  border-color: var(--orange) !important;
}

/* ════════════════════════════════════════════════════════════════
   TABS
   ════════════════════════════════════════════════════════════════ */
[data-testid="stTabs"] [role="tablist"] {
  background: var(--surface) !important;
  border: 1px solid var(--border) !important;
  border-radius: var(--r) !important;
  padding: 3px !important; gap: 2px !important;
  margin-bottom: 1.2rem !important;
}
[data-testid="stTabs"] button[role="tab"] {
  font-size: 12.5px !important; font-weight: 500 !important;
  color: var(--text-3) !important;
  border-radius: var(--r-s) !important;
  padding: .45rem 1rem !important;
  border: none !important;
  transition: all .14s !important; letter-spacing: -.01em !important;
}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
  background: var(--surface-3) !important;
  color: var(--text-1) !important; font-weight: 600 !important;
  box-shadow: var(--shadow-s) !important;
}
[data-testid="stTabs"] button[role="tab"]:hover { color: var(--text-2) !important; }

/* ════════════════════════════════════════════════════════════════
   DATAFRAMES
   ════════════════════════════════════════════════════════════════ */
[data-testid="stDataFrame"] {
  border: 1px solid var(--border) !important;
  border-radius: var(--r) !important;
  overflow: hidden !important;
  background: var(--surface) !important;
}
[data-testid="stDataFrame"] thead tr th {
  background: var(--surface-2) !important;
  color: var(--text-3) !important;
  font-size: 10.5px !important; font-weight: 600 !important;
  text-transform: uppercase !important; letter-spacing: .08em !important;
  border-bottom: 1px solid var(--border-m) !important;
}
[data-testid="stDataFrame"] tbody tr td {
  background: var(--surface) !important;
  color: var(--text-2) !important;
  font-size: 12.5px !important;
  border-bottom: 1px solid var(--border) !important;
  font-family: 'Geist Mono', monospace !important;
}
[data-testid="stDataFrame"] tbody tr:hover td {
  background: var(--surface-2) !important;
  color: var(--text-1) !important;
}

/* ════════════════════════════════════════════════════════════════
   EXPANDERS
   ════════════════════════════════════════════════════════════════ */
[data-testid="stExpander"] {
  background: var(--surface) !important;
  border: 1px solid var(--border) !important;
  border-radius: var(--r) !important;
  box-shadow: none !important;
  transition: border-color .14s !important;
}
[data-testid="stExpander"]:hover { border-color: var(--border-m) !important; }
[data-testid="stExpander"] summary {
  font-size: 13px !important; font-weight: 500 !important;
  color: var(--text-2) !important; padding: .7rem 1rem !important;
}
[data-testid="stExpander"] summary:hover { color: var(--text-1) !important; }

/* ════════════════════════════════════════════════════════════════
   PROGRESS / SPINNER
   ════════════════════════════════════════════════════════════════ */
[data-testid="stProgress"] > div {
  background: var(--surface-3) !important;
  border-radius: 4px !important; overflow: hidden !important;
}
[data-testid="stProgress"] > div > div {
  background: linear-gradient(90deg, var(--orange), var(--orange-d)) !important;
  border-radius: 4px !important;
  transition: width .3s ease !important;
}

/* ════════════════════════════════════════════════════════════════
   METRIC
   ════════════════════════════════════════════════════════════════ */
[data-testid="stMetricValue"] {
  font-family: 'Geist Mono', monospace !important;
  color: var(--text-1) !important; font-size: 1.6rem !important;
}
[data-testid="stMetricLabel"] { color: var(--text-3) !important; font-size: 11px !important; }

/* ════════════════════════════════════════════════════════════════
   TOAST
   ════════════════════════════════════════════════════════════════ */
[data-testid="stToast"] {
  background: var(--surface-3) !important;
  border: 1px solid var(--border-m) !important;
  border-radius: var(--r) !important;
  color: var(--text-1) !important;
  box-shadow: var(--shadow-m) !important;
}

/* ════════════════════════════════════════════════════════════════
   CUSTOM COMPONENTS
   ════════════════════════════════════════════════════════════════ */

/* ── KPI Card ─────────────────────────────────────────────────── */
.kpi-box {
  background: #ffffff;
  border: 1px solid var(--border);
  border-radius: var(--r-l);
  padding: 1.1rem 1.3rem .9rem;
  position: relative; overflow: hidden;
  transition: border-color .18s, box-shadow .18s;
}
.kpi-box:hover {
  border-color: var(--border-m);
  box-shadow: var(--shadow);
}
.kpi-box::before {
  content: '';
  position: absolute; top: 0; left: 0; right: 0; height: 1px;
  background: linear-gradient(90deg, transparent, var(--orange), transparent);
  opacity: .6;
}
.kpi-value {
  font-family: 'Geist Mono', monospace;
  font-size: 1.9rem; font-weight: 500;
  color: var(--text-1); line-height: 1; margin-bottom: .35rem;
  letter-spacing: -.04em;
}
.kpi-label {
  font-size: 10px; font-weight: 600;
  color: var(--text-3);
  text-transform: uppercase; letter-spacing: .12em;
}
.kpi-delta { font-size: 11px; margin-top: 6px; font-family: 'Geist Mono', monospace; }
.kpi-up   { color: var(--green); }
.kpi-down { color: var(--red);   }

/* ── Divider ─────────────────────────────────────────────────── */
.sdiv {
  height: 1px;
  background: var(--border);
  margin: 1.4rem 0;
}

/* ── Section header ──────────────────────────────────────────── */
.sec-head {
  display: flex; align-items: center; gap: .6rem;
  margin-bottom: .9rem;
}
.sec-head-title {
  font-family: 'Syne', sans-serif;
  font-size: .9rem; font-weight: 700;
  color: var(--text-2);
  letter-spacing: -.01em;
  white-space: nowrap;
}
.sec-head-line {
  flex: 1; height: 1px; background: var(--border);
}

/* ── Alerts ──────────────────────────────────────────────────── */
.alert {
  border-radius: var(--r);
  padding: .7rem .9rem;
  margin-bottom: .5rem;
  font-size: 12.5px; line-height: 1.55;
  display: flex; align-items: flex-start; gap: 8px;
}
.alert-r {
  background: var(--red-dim);
  border: 1px solid rgba(244,63,94,.25);
  color: #FDA4AF;
}
.alert-y {
  background: var(--amber-dim);
  border: 1px solid rgba(245,158,11,.25);
  color: #FCD34D;
}
.alert-g {
  background: var(--green-dim);
  border: 1px solid rgba(16,185,129,.25);
  color: #6EE7B7;
}
.alert-b {
  background: var(--blue-dim);
  border: 1px solid rgba(59,130,246,.25);
  color: #93C5FD;
}

/* ── Notification banner ─────────────────────────────────────── */
.notif-banner {
  background: linear-gradient(135deg, rgba(245,158,11,.08), rgba(243,110,33,.06));
  border: 1px solid rgba(245,158,11,.22);
  border-radius: var(--r); padding: .75rem 1rem;
  margin-bottom: 1rem; font-size: 12.5px; color: #FCD34D;
  line-height: 1.6;
}

/* ── Kanban ──────────────────────────────────────────────────── */
.kb-col-head {
  font-size: 10px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .12em;
  padding: .4rem .8rem; border-radius: var(--r-s);
  margin-bottom: .6rem;
  display: flex; align-items: center; justify-content: space-between;
}
.kb-aberto    { background: rgba(243,110,33,.1);  color: #FBA672; }
.kb-andamento { background: rgba(59,130,246,.1);  color: #93C5FD; }
.kb-travado   { background: rgba(244,63,94,.1);   color: #FDA4AF; }
.kb-concluido { background: rgba(16,185,129,.1);  color: #6EE7B7; }
.kb-cancelado { background: var(--overlay);       color: var(--text-3); }

.kb-card {
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--r);
  padding: .8rem .9rem;
  margin-bottom: .4rem;
  cursor: pointer;
  transition: border-color .14s, box-shadow .14s, transform .14s;
}
.kb-card:hover {
  border-color: var(--border-m);
  box-shadow: var(--shadow);
  transform: translateY(-1px);
}
.kb-card-title {
  font-weight: 600; font-size: 12.5px;
  color: var(--text-1); margin-bottom: 4px;
  letter-spacing: -.01em;
}
.kb-card-meta  { font-size: 11px; color: var(--text-3); line-height: 1.7; }
.kb-sla-ok     { color: var(--green);  font-size: 11px; font-weight: 600; margin-top: 4px; font-family: 'Geist Mono', monospace; }
.kb-sla-med    { color: var(--amber);  font-size: 11px; font-weight: 600; margin-top: 4px; font-family: 'Geist Mono', monospace; }
.kb-sla-bad    { color: var(--red);    font-size: 11px; font-weight: 600; margin-top: 4px; font-family: 'Geist Mono', monospace; }
.kb-motivo     { font-size: 11px; color: var(--red); margin-top: 4px; font-style: italic; }
.kb-obs-badge  {
  display: inline-block; background: var(--orange-dim);
  color: var(--orange); border-radius: var(--r-xs);
  padding: 1px 6px; font-size: 10px; font-weight: 700; margin-top: 3px;
}
.kb-wrap  { min-height: 60px; }
.kb-empty { font-size: 12px; color: var(--text-4); text-align: center; padding: 1.5rem 0; }
.kb-metric {
  font-size: 11px; color: var(--text-3); text-align: center;
  margin-top: .4rem; background: var(--overlay);
  border-radius: var(--r-s); padding: .25rem;
  font-family: 'Geist Mono', monospace;
}
.kb-tag-atrasada {
  display: inline-block;
  background: var(--red-dim); color: var(--red);
  border: 1px solid rgba(244,63,94,.25);
  border-radius: var(--r-xs); padding: 1px 6px;
  font-size: 9.5px; font-weight: 700; margin-top: 3px;
  letter-spacing: .04em; text-transform: uppercase;
}
.kb-tag-hoje {
  display: inline-block;
  background: var(--amber-dim); color: var(--amber);
  border: 1px solid rgba(245,158,11,.25);
  border-radius: var(--r-xs); padding: 1px 6px;
  font-size: 9.5px; font-weight: 700; margin-top: 3px;
}
.kb-tag-prog {
  display: inline-block;
  background: var(--blue-dim); color: var(--blue);
  border: 1px solid rgba(59,130,246,.25);
  border-radius: var(--r-xs); padding: 1px 6px;
  font-size: 9.5px; font-weight: 600; margin-top: 3px;
}

/* ── BI / Deduções ───────────────────────────────────────────── */
.deducao-row {
  display: flex; justify-content: space-between; align-items: center;
  padding: .4rem 0; border-bottom: 1px solid var(--border); font-size: 13px;
}
.deducao-label    { color: var(--text-3); }
.deducao-val-neg  { color: var(--red);   font-weight: 600; font-family: 'Geist Mono', monospace; }
.deducao-val-pos  { color: var(--green); font-weight: 700; font-size: 1.05em; font-family: 'Geist Mono', monospace; }

.modelo-card {
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--r-s); padding: .9rem; margin-bottom: .4rem;
}
.modelo-header {
  font-size: 10px; font-weight: 700; text-transform: uppercase;
  letter-spacing: .08em; padding: .25rem .6rem;
  border-radius: var(--r-xs); margin-bottom: .4rem; display: inline-block;
}
.modelo-assoc { background: var(--blue-dim);   color: var(--blue);   }
.modelo-cons  { background: var(--orange-dim);  color: var(--orange); }
.modelo-auto  { background: var(--green-dim);   color: var(--green);  }

/* ── Calendar ────────────────────────────────────────────────── */
.cal-box {
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--r-l); padding: 1.1rem 1.2rem;
}
.cal-day-header {
  font-size: 9px; font-weight: 700; color: var(--text-4);
  text-transform: uppercase; letter-spacing: .12em;
  margin-bottom: .35rem; margin-top: .7rem; padding-bottom: .25rem;
  border-bottom: 1px solid var(--border);
}
.cal-event {
  background: var(--orange-dim); border-left: 2px solid var(--orange);
  border-radius: var(--r-xs); padding: 3px 7px; font-size: 11px;
  margin-bottom: 3px; color: #FBA672;
}
.cal-event-atrasado {
  background: var(--red-dim); border-left: 2px solid var(--red);
  border-radius: var(--r-xs); padding: 3px 7px; font-size: 11px;
  margin-bottom: 3px; color: #FDA4AF;
}
.cal-event-hoje {
  background: var(--green-dim); border-left: 2px solid var(--green);
  border-radius: var(--r-xs); padding: 3px 7px; font-size: 11px;
  margin-bottom: 3px; color: #6EE7B7;
}

/* ── Login ───────────────────────────────────────────────────── */
.login-page {
  min-height: 100vh;
  background: var(--bg);
  display: flex; align-items: center; justify-content: center;
}
.login-card {
  background: var(--surface);
  border: 1px solid var(--border-m);
  border-radius: var(--r-xl);
  padding: 2.8rem 2.5rem 2.4rem;
  box-shadow: var(--shadow-l), 0 0 80px rgba(243,110,33,.06);
  max-width: 400px; width: 100%;
}
.login-brand {
  font-family: 'Syne', sans-serif;
  font-size: 1.35rem; color: var(--text-1);
  font-weight: 700; letter-spacing: -.02em; margin-bottom: .1rem;
}
.login-sub {
  font-size: 11px; color: var(--text-3);
  text-transform: uppercase; letter-spacing: .1em;
  margin-bottom: 1.8rem;
}
.login-form-label {
  font-size: 10.5px; font-weight: 600; color: var(--text-3);
  text-transform: uppercase; letter-spacing: .1em; margin-bottom: .3rem;
}

/* ── Page header ─────────────────────────────────────────────── */
.page-header {
  display: flex; align-items: flex-end; justify-content: space-between;
  margin-bottom: 1.5rem; padding-bottom: .9rem;
  border-bottom: 1px solid var(--border);
}
.page-header-title {
  font-family: 'Syne', sans-serif;
  font-size: 1.65rem; color: var(--text-1); font-weight: 700;
  letter-spacing: -.03em; line-height: 1;
}
.page-header-sub { font-size: 12px; color: var(--text-3); margin-top: .3rem; }

/* ── v12 Gestor Dashboard ────────────────────────────────────── */
.gestor-analista-card {
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--r-l); padding: .8rem 1rem;
  margin-bottom: .5rem;
  transition: border-color .14s, box-shadow .14s;
}
.gestor-analista-card:hover {
  border-color: var(--border-m);
  box-shadow: var(--shadow);
}

/* ── v12 Auditoria / Medição ─────────────────────────────────── */
.audit-alert-crit {
  background: var(--red-dim);
  border-left: 3px solid var(--red);
  border-radius: var(--r); padding: .7rem .9rem;
  margin-bottom: .4rem; font-size: 12.5px; color: #FDA4AF;
}
.audit-alert-warn {
  background: var(--amber-dim);
  border-left: 3px solid var(--amber);
  border-radius: var(--r); padding: .7rem .9rem;
  margin-bottom: .4rem; font-size: 12.5px; color: #FCD34D;
}
.audit-alert-ok {
  background: var(--green-dim);
  border-left: 3px solid var(--green);
  border-radius: var(--r); padding: .7rem .9rem;
  margin-bottom: .4rem; font-size: 12.5px; color: #6EE7B7;
}

/* ── v12 Gantt ───────────────────────────────────────────────── */
.gantt-row-v12 {
  display: flex; align-items: center; gap: 8px;
  padding: 5px 0; border-bottom: 1px solid var(--border);
  font-size: 12.5px;
}
.gantt-track-v12 {
  flex: 1; height: 6px; background: var(--surface-3);
  border-radius: 3px; overflow: hidden;
}
.gantt-fill-v12 { height: 100%; border-radius: 3px; transition: width .3s ease; }

/* ── Sidebar badges ──────────────────────────────────────────── */
.sb-badge-warn {
  background: var(--red-dim); border: 1px solid rgba(244,63,94,.22);
  border-radius: var(--r-s); padding: 5px 10px;
  margin-bottom: 5px; font-size: 11.5px; color: #FDA4AF;
}
.sb-badge-info {
  background: var(--amber-dim); border: 1px solid rgba(245,158,11,.22);
  border-radius: var(--r-s); padding: 5px 10px;
  margin-bottom: 5px; font-size: 11.5px; color: #FCD34D;
}
</style>
"""
# ── PATHS ─────────────────────────────────────────────────────────────────────
DB              = "database"
USERS_FILE      = "users.json"
GER_FILE        = f"{DB}/geradores.json"
USI_FILE        = f"{DB}/usinas.json"
TASKS_FILE      = f"{DB}/tasks.json"
GERACAO_FILE    = f"{DB}/geracao_usinas.json"
BACKOFFICE_FILE = f"{DB}/backoffice.json"
RATEIO_FILE     = f"{DB}/historico_rateios.json"
ANALISES_FILE   = f"{DB}/historico_analises.json"
GOOGLE_FILE     = f"{DB}/google_tokens.json"

os.makedirs(DB, exist_ok=True)

# Tarifas base por enquadramento
TARIFAS_BASE = {"GD1": 0.8182, "GD2": 0.64788}

# Modelos de negócio (múltiplos relatórios de medição)
MODELOS_NEGOCIO = ["Associação", "Consórcio", "Autoconsumo"]

def _load(path, default):
    if not os.path.exists(path): _save(path, default)
    with open(path, encoding="utf-8") as f: return json.load(f)

def _save(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)

# ── AUTH ──────────────────────────────────────────────────────────────────────
def load_users():
    if not os.path.exists(USERS_FILE):
        _save(USERS_FILE, {"users": [
            {"name":"Milena",   "email":"milena@sunne.com.br",   "password":"sunne2026","role":"admin"},
            {"name":"Analista", "email":"analista@sunne.com.br", "password":"sunne2026","role":"user"},
        ]})
    with open(USERS_FILE) as f: return json.load(f).get("users", [])

def authenticate(email, password):
    for u in load_users():
        if u["email"].lower() == email.lower() and u["password"] == password:
            return u
    return None

# ── CRUD ──────────────────────────────────────────────────────────────────────
def load_geradores():   return _load(GER_FILE, [])
def save_geradores(d):  _save(GER_FILE, d)
def load_usinas():      return _load(USI_FILE, [])
def save_usinas(d):     _save(USI_FILE, d)
def load_tasks():       return _load(TASKS_FILE, [])
def save_tasks(d):      _save(TASKS_FILE, d)
def load_geracao():     return _load(GERACAO_FILE, [])
def save_geracao(d):    _save(GERACAO_FILE, d)
def load_backoffice():  return _load(BACKOFFICE_FILE, [])
def save_backoffice(d): _save(BACKOFFICE_FILE, d)
def load_rateios():     return _load(RATEIO_FILE, {})
def save_rateios(d):    _save(RATEIO_FILE, d)
def load_analises():    return _load(ANALISES_FILE, {})
def save_analises(d):   _save(ANALISES_FILE, d)
def load_google_tokens(): return _load(GOOGLE_FILE, {})
def save_google_tokens(d): _save(GOOGLE_FILE, d)

# ── TASK DEADLINE HELPERS ─────────────────────────────────────────────────────
def task_data_programada(t) -> datetime | None:
    """Retorna datetime da data programada ou None."""
    v = t.get("data_programada", "")
    if not v: return None
    for fmt in ["%d/%m/%Y %H:%M", "%d/%m/%Y"]:
        try: return datetime.strptime(str(v).strip(), fmt)
        except: pass
    return None

def task_esta_atrasada(t) -> bool:
    """True se tem data programada, passou, e não está concluída/cancelada."""
    if t.get("status") in ("Concluido", "Cancelado"): return False
    dp = task_data_programada(t)
    return dp is not None and dp.date() < datetime.now().date()

def task_vence_hoje(t) -> bool:
    if t.get("status") in ("Concluido", "Cancelado"): return False
    dp = task_data_programada(t)
    return dp is not None and dp.date() == datetime.now().date()

# ── TASK HELPERS ──────────────────────────────────────────────────────────────
TIPOS       = ["Avulsa","Análise de Faturamento","Rateio","Captura","Relatório","Auditoria"]
STATUS_LIST = ["Em aberto","Em andamento","Travado","Concluido","Cancelado"]
MOTIVO_OBRIG = {"Travado","Cancelado"}
KB_CSS = {"Em aberto":"kb-aberto","Em andamento":"kb-andamento","Travado":"kb-travado",
          "Concluido":"kb-concluido","Cancelado":"kb-cancelado"}

def new_task(titulo, usina, gerador, analista, tipo="Avulsa",
             agendamento="", descricao="", anexo_nome="", anexo_b64="",
             data_programada=""):
    tasks = load_tasks()
    tasks.append({
        "id":              datetime.now().strftime("%Y%m%d%H%M%S%f"),
        "titulo":          titulo, "usina":usina, "gerador":gerador,
        "analista":        analista, "tipo":tipo, "agendamento":agendamento,
        "data_programada": data_programada,
        "descricao":       descricao, "observacoes":"",
        "anexo_nome":      anexo_nome, "anexo_b64":anexo_b64,
        "status":          "Em aberto", "motivo_bloqueio":"",
        "criado_em":       datetime.now().strftime("%d/%m/%Y %H:%M"),
        "historico":       [],
    })
    save_tasks(tasks)

def update_task(tid, **kwargs):
    tasks = load_tasks()
    for t in tasks:
        if t["id"] == tid:
            if "status" in kwargs and kwargs["status"] != t["status"]:
                t.setdefault("historico", []).append({
                    "de":t["status"],"para":kwargs["status"],
                    "em":datetime.now().strftime("%d/%m/%Y %H:%M"),
                })
            t.update(kwargs)
    save_tasks(tasks)

def sla_days(t):
    try: return (datetime.now() - datetime.strptime(t["criado_em"], "%d/%m/%Y %H:%M")).days
    except: return 0

def sla_cls(d):
    if d <= 3: return "kb-sla-ok"
    if d <= 7: return "kb-sla-med"
    return "kb-sla-bad"

# ══════════════════════════════════════════════════════════════════════════════
# v12 · CRONOGRAMA MENSAL AUTOMÁTICO
# ══════════════════════════════════════════════════════════════════════════════

CRON_FILE = f"{DB}/cronograma_gerado.json"

def load_cronograma_gerado(): return _load(CRON_FILE, [])
def save_cronograma_gerado(d): _save(CRON_FILE, d)

CRON_SEMANA = {
    "relatorio_medicao": {"semana": 1, "dia_ini": 1,  "dia_fim": 7,  "label": "Relatório de Medição",  "tipo": "Relatório"},
    "fatura_ug":         {"semana": 1, "dia_ini": 1,  "dia_fim": 7,  "label": "Fatura da UG",          "tipo": "Captura"},
    "auditoria_ufv":     {"semana": 2, "dia_ini": 8,  "dia_fim": 21, "label": "Auditoria UFV",         "tipo": "Auditoria"},
    "auditoria_ucs":     {"semana": 2, "dia_ini": 8,  "dia_fim": 21, "label": "Auditoria UCs",         "tipo": "Auditoria"},
    "material_gerador":  {"semana": 3, "dia_ini": 15, "dia_fim": 21, "label": "Material do Gerador",   "tipo": "Relatório"},
    "checkpoint_s3":     {"semana": 3, "dia_ini": 15, "dia_fim": 21, "label": "Checkpoint Gerador S3", "tipo": "Avulsa"},
    "subir_rateio":      {"semana": 4, "dia_ini": 22, "dia_fim": 28, "label": "Subir Rateio",          "tipo": "Rateio"},
    "inadimplencia":     {"semana": 4, "dia_ini": 22, "dia_fim": 28, "label": "Levantar Inadimplência","tipo": "Análise de Faturamento"},
    "checkpoint_s4":     {"semana": 4, "dia_ini": 22, "dia_fim": 31, "label": "Checkpoint Gerador S4", "tipo": "Avulsa"},
}

def _data_programada_cron(ano: int, mes: int, dia_fim: int) -> str:
    import calendar
    ultimo = calendar.monthrange(ano, mes)[1]
    d = min(dia_fim, ultimo)
    return f"{d:02d}/{mes:02d}/{ano}"

def gerar_cronograma_mensal(ano: int, mes: int, analista: str) -> int:
    """
    Para cada usina do analista cria as tarefas mensais que ainda não existem.
    Retorna o nº de tarefas criadas.
    """
    usinas   = [u for u in load_usinas() if u.get("analista","").lower() == analista.lower()]
    geradores= load_geradores()
    tasks    = load_tasks()
    cron_log = load_cronograma_gerado()
    mes_str  = f"{mes:02d}/{ano}"

    # índice de tarefas existentes para evitar duplicatas
    existentes = {
        (t.get("tipo",""), t.get("usina",""), t.get("agendamento",""))
        for t in tasks
    }

    criadas = 0
    for u in usinas:
        uc       = str(u.get("uc","")).strip()
        ufv      = u.get("ufv","") or uc
        gerador  = u.get("gerador","")
        ger_obj  = next((g for g in geradores if g.get("gerador","").lower() == gerador.lower()), {})

        for chave, cfg in CRON_SEMANA.items():
            titulo = f"{cfg['label']} — {ufv}"
            dp     = _data_programada_cron(ano, mes, cfg["dia_fim"])
            chave_dup = (cfg["tipo"], uc, mes_str)

            # Só cria se ainda não existe tarefa do mesmo tipo/usina/mês
            if any(
                t.get("tipo") == cfg["tipo"]
                and t.get("usina") == uc
                and mes_str in t.get("data_programada","")
                and cfg["label"].split("—")[0].strip().lower() in t.get("titulo","").lower()
                for t in tasks
            ):
                continue

            new_task(
                titulo        = titulo,
                usina         = uc,
                gerador       = gerador,
                analista      = analista,
                tipo          = cfg["tipo"],
                agendamento   = mes_str,
                descricao     = f"Tarefa automática — cronograma {mes_str} · Semana {cfg['semana']}",
                data_programada = dp,
            )
            criadas += 1

    # Registra log de geração
    cron_log.append({
        "mes_ref":    mes_str,
        "analista":   analista,
        "criadas":    criadas,
        "gerado_em":  datetime.now().strftime("%d/%m/%Y %H:%M"),
    })
    save_cronograma_gerado(cron_log)
    return criadas


# ══════════════════════════════════════════════════════════════════════════════
# v12 · CRUZAMENTO EXTRATO vs RELATÓRIO DE MEDIÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def cruzar_medicao_extrato(df_extrato: "pd.DataFrame", lista_df_medicao: list, competencia_alvo: str) -> dict:
    """
    Racional Operacional Homologado:
    1. Busca no extrato apenas linhas onde 'Status de Pagamento' seja exatamente 'Pago'.
    2. Filtra pela data real contida em 'Data de Pagamento Boleto Sunne' usando o padrão brasileiro.
    3. Coleta os códigos contidos na coluna 'Nº do Cliente' de até 3 relatórios de medição.
    4. Cruza por UC, consolidando os valores reais a partir da coluna 'Total Pago'.
    """
    # Colunas exatas do Extrato Detalhado
    col_uc_e  = "Número da UC"
    col_st_e  = "Status de Pagamento"
    col_dt_e  = "Data de Pagamento Boleto Sunne"
    col_val_e = "Total Pago"

    if col_uc_e not in df_extrato.columns: return {"erro": f"Coluna '{col_uc_e}' não encontrada no extrato."}
    if col_st_e not in df_extrato.columns: return {"erro": f"Coluna '{col_st_e}' não encontrada no extrato."}
    if col_dt_e not in df_extrato.columns: return {"erro": f"Coluna '{col_dt_e}' não encontrada no extrato."}
    if col_val_e not in df_extrato.columns: return {"erro": f"Coluna '{col_val_e}' não encontrada no extrato."}

    try:
        tgt_mes, tgt_ano = competencia_alvo.strip().split("/")
        tgt_mes = int(tgt_mes)
        tgt_ano = int(tgt_ano)
    except:
        return {"erro": "Formato de competência inválido. Preencha como MM/AAAA (Ex: 05/2026)."}

    linhas_filtradas = []
    for _, row in df_extrato.iterrows():
        # Regra 1: Filtrar apenas faturas com status "Pago"
        if str(row[col_st_e]).strip().lower() != "pago":
            continue

        val_dt = row[col_dt_e]
        if pd.isna(val_dt) or str(val_dt).strip() == "":
            continue
            
        # Regra 2: Validar competência garantindo leitura correta de datas brasileiras
        match_date = False
        dt_parsed = pd.to_datetime(val_dt, dayfirst=True, errors='coerce')
        if pd.notna(dt_parsed):
            if dt_parsed.month == tgt_mes and dt_parsed.year == tgt_ano:
                match_date = True
        else:
            dt_str = str(val_dt).strip()
            import re
            digits = re.findall(r'\d+', dt_str)
            if len(digits) >= 3:
                if len(digits[0]) == 4: # Formato YYYY-MM-DD
                    if int(digits[1]) == tgt_mes and int(digits[0]) == tgt_ano: match_date = True
                elif len(digits[2]) >= 4: # Formato DD/MM/YYYY
                    if int(digits[1]) == tgt_mes and int(digits[2][:4]) == tgt_ano: match_date = True

        if match_date:
            linhas_filtradas.append(row)

    if not linhas_filtradas:
        return {"ok": [], "ausentes": [], "extras": [], "total_pago": 0.0, "total_ausente_valor": 0.0, "aviso": f"Nenhum pagamento real localizado no extrato para a competência {competencia_alvo}."}

    df_pago = pd.DataFrame(linhas_filtradas)
    df_pago["_uc_norm"] = df_pago[col_uc_e].apply(normalize_uc)

    # Consolida os IDs dos clientes de todos os relatórios de medição ativos enviados pela coluna "Nº do Cliente"
    clientes_na_medicao = set()
    for df_medicao in lista_df_medicao:
        if df_medicao is None or df_medicao.empty: continue
        col_uc_m = next((c for c in df_medicao.columns if "nº do cliente" in c.lower() or "numero do cliente" in c.lower() or "cliente" in c.lower()), None)
        
        if col_uc_m:
            for _, row_m in df_medicao.iterrows():
                uc_m_norm = normalize_uc(str(row_m[col_uc_m]))
                if uc_m_norm:
                    clientes_na_medicao.add(uc_m_norm)

    ok, ausentes = [], []
    for _, row in df_pago.iterrows():
        uc = row["_uc_norm"]
        val = clean_val(row[col_val_e]) 
        tit = str(row["Titular da Conta"]) if "Titular da Conta" in df_extrato.columns else "—"
        comp_fatura = str(row["Competência"]) if "Competência" in df_extrato.columns else "—"
        
        item = {"uc": row[col_uc_e], "competencia": comp_fatura, "valor": val, "titular": tit, "data_pagamento": row[col_dt_e]}
        
        # Como o relatório já é da competência correta, valida apenas a presença da UC nas tabelas B
        if uc in clientes_na_medicao:
            ok.append(item)
        else:
            ausentes.append(item)

    return {
        "ok": ok, "ausentes": ausentes, "extras": [],
        "total_pago": sum(i["valor"] for i in ok) + sum(i["valor"] for i in ausentes),
        "total_ausente_valor": sum(i["valor"] for i in ausentes),
    }
# ══════════════════════════════════════════════════════════════════════════════
# v12 · AUDITORIA UFV / UCS — PARSER DA PLANILHA NAFISAH
# ══════════════════════════════════════════════════════════════════════════════

def parse_auditoria_planilha(file_or_bytes) -> dict:
    """
    Lê a planilha de auditoria (formato Nafisah UFV7 / padrão Sunne).
    Retorna dict com:
      - usina_resumo    : {geracao, consumo, creditos, saldo_total, saldo_usina}
      - beneficiarios   : lista de {uc, saldo, consumo, autonomia, percentual, ancora, reprovado}
      - rateio_base     : lista de {uc, cnpj, percentual, ancora, reprovado}
      - alertas_ben     : lista de {uc, tipo, mensagem}
      - alertas_rateio  : lista de {tipo, mensagem}
      - meses_analisados: int
    """
    import openpyxl
    from openpyxl import load_workbook

    if isinstance(file_or_bytes, bytes):
        wb = load_workbook(io.BytesIO(file_or_bytes), read_only=True, data_only=True)
    else:
        wb = load_workbook(file_or_bytes, read_only=True, data_only=True)

    result = {
        "usina_resumo":     {},
        "beneficiarios":    [],
        "rateio_base":      [],
        "alertas_ben":      [],
        "alertas_rateio":   [],
        "meses_analisados": 0,
        "nome_planilha":    "",
        "erro":             "",
    }

    # ── Aba principal "Tabela" ──────────────────────────────────────────────
    sheet_tabela = None
    for nome in wb.sheetnames:
        if "tabela" in nome.lower() and "últim" not in nome.lower() and "base" not in nome.lower():
            sheet_tabela = wb[nome]
            result["nome_planilha"] = nome
            break
    if not sheet_tabela and wb.sheetnames:
        sheet_tabela = wb[wb.sheetnames[0]]

    if sheet_tabela:
        rows = list(sheet_tabela.iter_rows(values_only=True))
        geracao_row = consumo_row = creditos_row = saldo_total_row = saldo_usina_row = None

        for row in rows[:30]:
            lbl = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
            if "geração"   in lbl or "geracao" in lbl:  geracao_row    = row
            elif "consumo" in lbl:                       consumo_row    = row
            elif "créditos utiliz" in lbl:               creditos_row   = row
            elif "saldo acumulado total" in lbl:         saldo_total_row= row
            elif "saldo acumulado usina" in lbl:         saldo_usina_row= row

        def _row_vals(row):
            if row is None: return []
            return [clean_val(v) for v in row[2:] if v not in (None, "", "nan")]

        ger_vals   = _row_vals(geracao_row)
        cons_vals  = _row_vals(consumo_row)
        cred_vals  = _row_vals(creditos_row)
        sald_vals  = _row_vals(saldo_total_row)

        # Remove zeros no início (meses sem dados)
        def _strip_zeros(lst): return [v for v in lst if v > 0]
        ger_nz  = _strip_zeros(ger_vals)
        cons_nz = _strip_zeros(cons_vals)
        cred_nz = _strip_zeros(cred_vals)
        sald_nz = _strip_zeros(sald_vals)

        meses = max(len(ger_nz), len(cons_nz), len(cred_nz), 1)
        result["meses_analisados"] = meses
        result["usina_resumo"] = {
            "geracao_total":  sum(ger_nz),
            "consumo_total":  sum(cons_nz),
            "creditos_total": sum(cred_nz),
            "saldo_atual":    sald_nz[-1] if sald_nz else 0.0,
            "geracao_media":  sum(ger_nz) / len(ger_nz) if ger_nz else 0.0,
            "consumo_media":  sum(cons_nz) / len(cons_nz) if cons_nz else 0.0,
        }

        # ── Saúde do rateio: consumo > geração por 3+ meses ────────────────
        meses_cons_maior = sum(
            1 for g, c in zip(ger_nz, cons_nz) if c > g
        )
        if meses_cons_maior >= 3:
            result["alertas_rateio"].append({
                "tipo": "critico",
                "mensagem": f"🚨 Usina com consumo > geração por {meses_cons_maior} meses consecutivos. "
                            "Necessário remanejar beneficiários.",
            })
        elif meses_cons_maior >= 1:
            result["alertas_rateio"].append({
                "tipo": "atencao",
                "mensagem": f"⚠️ Usina com consumo > geração em {meses_cons_maior} mês(es). Monitorar.",
            })

        # ── Beneficiários: bloco "Autonomia" ──────────────────────────────
        # Localiza seção de beneficiários via cabeçalho "UC | Saldo | Consumo | Autonomia"
        in_ben = False
        for row in rows:
            cells = [str(c).strip().lower() if c else "" for c in row]
            if "autonomia" in cells and "saldo" in cells and "consumo" in cells:
                in_ben = True
                continue
            if not in_ben:
                continue
            # Linha de dados: primeiro campo parece UC (número grande)
            uc_val = row[0] if row else None
            if uc_val is None or str(uc_val).strip() in ("", "None", "nan"):
                continue
            uc_str    = str(uc_val).strip()
            saldo_v   = clean_val(row[1]) if len(row) > 1 else 0.0
            consumo_v = clean_val(row[2]) if len(row) > 2 else 0.0
            autonomia_raw = str(row[3]).strip() if len(row) > 3 and row[3] else "0"

            try:
                autonomia_v = float(str(autonomia_raw).replace(",","."))
            except Exception:
                autonomia_v = 0.0

            result["beneficiarios"].append({
                "uc":        uc_str,
                "saldo":     saldo_v,
                "consumo":   consumo_v,
                "autonomia": autonomia_v,
            })

            # Alertas por beneficiário
            if autonomia_v > 3:
                result["alertas_ben"].append({
                    "uc":      uc_str,
                    "tipo":    "saldo_alto",
                    "mensagem": f"💰 UC {uc_str}: saldo acumulado = {autonomia_v:.1f} meses — "
                                "reduzir % de rateio para queimar saldo.",
                })
            if consumo_v > 0 and saldo_v == 0 and autonomia_v < 0.5:
                result["alertas_ben"].append({
                    "uc":      uc_str,
                    "tipo":    "recebendo_menos",
                    "mensagem": f"📉 UC {uc_str}: recebendo menos do que consome — "
                                "aumentar % de rateio.",
                })

    # ── Aba "Base do último rateio" ─────────────────────────────────────────
    sheet_rateio = next((wb[n] for n in wb.sheetnames if "rateio" in n.lower() or "último" in n.lower()), None)
    if sheet_rateio:
        in_rateio = False
        for row in sheet_rateio.iter_rows(values_only=True):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if "número uc" in cells or "numero uc" in cells or ("uc" in cells and "percentual" in cells):
                in_rateio = True
                continue
            if not in_rateio:
                continue
            if not row[0]:
                continue
            uc_r   = str(row[0]).strip()
            cnpj_r = str(row[1]).strip() if len(row) > 1 and row[1] else "—"
            pct_r  = clean_val(row[2]) if len(row) > 2 else 0.0
            anc_r  = str(row[3]).strip() if len(row) > 3 and row[3] else "Não"
            rep_r  = str(row[4]).strip() if len(row) > 4 and row[4] else "Não"
            if uc_r and uc_r.lower() not in ("nan","none","número uc","numero uc"):
                result["rateio_base"].append({
                    "uc": uc_r, "cnpj": cnpj_r, "percentual": pct_r,
                    "ancora": anc_r, "reprovado": rep_r,
                })

    wb.close()
    return result


# ══════════════════════════════════════════════════════════════════════════════
# v12 · DASHBOARD DO GESTOR
# ══════════════════════════════════════════════════════════════════════════════

def page_gestor_dashboard():
    st.title("Painel do Gestor")
    st.caption("Visão consolidada de todos os analistas e cronogramas — acesso exclusivo Admin")

    users  = load_users()
    tasks  = load_tasks()
    agora  = datetime.now()

    # ── Filtros ───────────────────────────────────────────────────────────────
    analistas_nomes = sorted({u["name"] for u in users})
    fc1, fc2, fc3 = st.columns(3)
    f_an  = fc1.selectbox("Analista", ["Todos"] + analistas_nomes, key="gest_fan")
    meses_opts = sorted(
        {t.get("agendamento","") for t in tasks if t.get("agendamento","")},
        key=lambda x: (int(x.split("/")[1]) if "/" in x else 0,
                       int(x.split("/")[0]) if "/" in x else 0),
        reverse=True,
    )
    f_mes = fc2.selectbox("Mês ref.", ["Todos"] + meses_opts, key="gest_fmes")
    f_tipo= fc3.selectbox("Tipo", ["Todos"] + TIPOS, key="gest_ftipo")

    def _match_gest(t):
        if f_an  != "Todos" and t.get("analista","").lower() != f_an.lower():  return False
        if f_mes != "Todos" and t.get("agendamento","") != f_mes:              return False
        if f_tipo!= "Todos" and t.get("tipo","") != f_tipo:                    return False
        return True

    tasks_f = [t for t in tasks if _match_gest(t)]

    # ── KPIs globais ──────────────────────────────────────────────────────────
    st.markdown("<div class='sdiv'></div>", unsafe_allow_html=True)
    ab_g    = [t for t in tasks_f if t["status"] == "Em aberto"]
    and_g   = [t for t in tasks_f if t["status"] == "Em andamento"]
    trav_g  = [t for t in tasks_f if t["status"] == "Travado"]
    conc_g  = [t for t in tasks_f if t["status"] == "Concluido"]
    atras_g = [t for t in tasks_f if task_esta_atrasada(t)]
    ativos_g= [t for t in tasks_f if t["status"] not in ("Concluido","Cancelado")]
    sla_g   = round(sum(sla_days(t) for t in ativos_g)/len(ativos_g), 1) if ativos_g else 0

    kpi_cols = st.columns(7)
    for col, lbl, val, cor in zip(kpi_cols,
        ["Total","Em Aberto","Andamento","Travadas","Atrasadas","Concluídas","SLA Médio"],
        [len(tasks_f), len(ab_g), len(and_g), len(trav_g), len(atras_g), len(conc_g), f"{sla_g}d"],
        ["var(--rubi)","#A84010","#7A5010","#C41230","#C41230","#0B7A5F","var(--rubi)"],
    ):
        col.markdown(
            f'<div class="kpi-box"><div class="kpi-value" style="color:{cor};font-size:1.6rem">{val}</div>'
            f'<div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)

    st.write("")

    # ── Performance por analista ──────────────────────────────────────────────
    st.markdown('<div class="sec-head"><span class="sec-head-title">Performance por Analista</span>'
                '<div class="sec-head-line"></div></div>', unsafe_allow_html=True)

    analistas_data = []
    for u in users:
        an_name = u["name"]
        tt = [t for t in tasks if t.get("analista","").lower() == an_name.lower()
              and _match_gest(t)]
        if not tt: continue
        ab_a   = sum(1 for t in tt if t["status"] == "Em aberto")
        and_a  = sum(1 for t in tt if t["status"] == "Em andamento")
        trav_a = sum(1 for t in tt if t["status"] == "Travado")
        conc_a = sum(1 for t in tt if t["status"] == "Concluido")
        atras_a= sum(1 for t in tt if task_esta_atrasada(t))
        total_a= len(tt)
        pct_a  = round(conc_a / total_a * 100) if total_a > 0 else 0
        sla_a  = round(sum(sla_days(t) for t in tt if t["status"] not in ("Concluido","Cancelado"))
                       / max(1, sum(1 for t in tt if t["status"] not in ("Concluido","Cancelado"))), 1)
        analistas_data.append({
            "Analista": an_name, "Total": total_a,
            "Aberto": ab_a, "Andamento": and_a,
            "Travado": trav_a, "Atrasadas": atras_a,
            "Concluídas": conc_a, "% Concluído": pct_a, "SLA (d)": sla_a,
        })

    if analistas_data:
        for ad in sorted(analistas_data, key=lambda x: x["Atrasadas"], reverse=True):
            iniciais = "".join(p[0].upper() for p in ad["Analista"].split()[:2])
            cor_sla  = "#C41230" if ad["SLA (d)"] > 7 else ("#F36E21" if ad["SLA (d)"] > 3 else "#0B7A5F")
            pct_bar  = ad["% Concluído"]
            cor_bar  = "#0B7A5F" if pct_bar >= 70 else ("#F36E21" if pct_bar >= 40 else "#C41230")

            c_av, c_info, c_det = st.columns([0.7, 6, 2])
            with c_av:
                st.markdown(
                    f'<div style="width:38px;height:38px;border-radius:50%;background:var(--gold-l);'
                    f'display:flex;align-items:center;justify-content:center;font-weight:600;'
                    f'font-size:13px;color:var(--gold);margin-top:4px">{iniciais}</div>',
                    unsafe_allow_html=True)
            with c_info:
                st.markdown(
                    f'<div style="padding:8px 0">'
                    f'<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:4px">'
                    f'<b style="font-size:14px">{ad["Analista"]}</b>'
                    f'<span style="background:#FFF0F3;color:#8B1530;border:1px solid #FFC8D0;'
                    f'border-radius:4px;padding:1px 7px;font-size:11px;font-weight:700">'
                    f'⏰ {ad["Atrasadas"]} atrasadas</span>'
                    f'<span style="background:#FFFBEC;color:#7A5010;border:1px solid #FFE580;'
                    f'border-radius:4px;padding:1px 7px;font-size:11px">{ad["Aberto"]} abertas</span>'
                    f'<span style="background:#EDFCF6;color:#0A5040;border:1px solid #A0EDCE;'
                    f'border-radius:4px;padding:1px 7px;font-size:11px">{ad["Concluídas"]} concluídas</span>'
                    f'<span style="background:#EEF4FF;color:#1E3A8A;border:1px solid #B8D0FF;'
                    f'border-radius:4px;padding:1px 7px;font-size:11px">{ad["Travado"]} travadas</span>'
                    f'</div>'
                    f'<div style="display:flex;align-items:center;gap:8px">'
                    f'<div style="flex:1;height:6px;background:var(--border);border-radius:3px;overflow:hidden">'
                    f'<div style="width:{pct_bar}%;height:100%;background:{cor_bar};border-radius:3px"></div>'
                    f'</div>'
                    f'<span style="font-size:11px;color:var(--ink-l);white-space:nowrap">'
                    f'{pct_bar}% · SLA <b style="color:{cor_sla}">{ad["SLA (d)"]}d</b></span>'
                    f'</div></div>',
                    unsafe_allow_html=True)
            with c_det:
                if st.button(f"Ver tarefas de {ad['Analista']}", key=f"gest_det_{ad['Analista']}"):
                    st.session_state["gest_detalhe_analista"] = ad["Analista"]
                    st.session_state["page"] = "atividades"
                    st.rerun()
            st.markdown("<div class='sdiv' style='margin:.4rem 0'></div>", unsafe_allow_html=True)

    # ── Progresso do cronograma mensal ────────────────────────────────────────
    st.write("")
    st.markdown('<div class="sec-head"><span class="sec-head-title">Cronograma do Mês — Andamento</span>'
                '<div class="sec-head-line"></div></div>', unsafe_allow_html=True)

    mes_cron = f_mes if f_mes != "Todos" else agora.strftime("%m/%Y")
    tasks_mes = [t for t in tasks if t.get("agendamento","") == mes_cron]

    if not tasks_mes:
        st.info(f"Nenhuma tarefa gerada para {mes_cron}. "
                "Use a seção **Cronograma Automático** para gerar.")
    else:
        for chave, cfg in CRON_SEMANA.items():
            lbl   = cfg["label"]
            tt_cr = [t for t in tasks_mes if cfg["label"].split("—")[0].strip().lower()
                     in t.get("titulo","").lower()]
            total_cr = len(tt_cr)
            if total_cr == 0: continue
            conc_cr  = sum(1 for t in tt_cr if t["status"] == "Concluido")
            atras_cr = sum(1 for t in tt_cr if task_esta_atrasada(t))
            pct_cr   = round(conc_cr / total_cr * 100)
            cor_cr   = "#0B7A5F" if pct_cr >= 70 else ("#F36E21" if pct_cr >= 30 else "#C41230")

            c1, c2, c3 = st.columns([3, 5, 1.5])
            c1.markdown(f'<div style="font-size:13px;padding:4px 0">'
                        f'<b>S{cfg["semana"]}</b> · {lbl}</div>', unsafe_allow_html=True)
            with c2:
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:8px;padding:4px 0">'
                    f'<div style="flex:1;height:8px;background:var(--border);border-radius:4px;overflow:hidden">'
                    f'<div style="width:{pct_cr}%;height:100%;background:{cor_cr};border-radius:4px"></div></div>'
                    f'<span style="font-size:11px;color:var(--ink-l);white-space:nowrap">'
                    f'{conc_cr}/{total_cr} · {pct_cr}%</span></div>',
                    unsafe_allow_html=True)
            c3.markdown(
                f'<div style="padding:4px 0;font-size:11px">'
                f'{"🔴 " + str(atras_cr) + " atrasadas" if atras_cr else "✅"}</div>',
                unsafe_allow_html=True)

    # ── BI de setor ───────────────────────────────────────────────────────────
    st.write("")
    st.markdown('<div class="sec-head"><span class="sec-head-title">Distribuição de Status — Setor</span>'
                '<div class="sec-head-line"></div></div>', unsafe_allow_html=True)

    contagens = {s: sum(1 for t in tasks_f if t["status"] == s) for s in STATUS_LIST}
    df_status = pd.DataFrame([
        {"Status": s, "Qtd": contagens.get(s, 0)} for s in STATUS_LIST
    ])
    st.bar_chart(df_status.set_index("Status"), height=200)

    # ── Geração de cronograma ─────────────────────────────────────────────────
    st.write("")
    st.markdown('<div class="sec-head"><span class="sec-head-title">⚙️ Gerar Cronograma Mensal Automático</span>'
                '<div class="sec-head-line"></div></div>', unsafe_allow_html=True)

    with st.expander("Gerar tarefas mensais para um analista"):
        gc1, gc2, gc3 = st.columns(3)
        gc_an  = gc1.selectbox("Analista",  analistas_nomes, key="gc_an")
        gc_mes = gc2.number_input("Mês",  1, 12, agora.month, key="gc_mes")
        gc_ano = gc3.number_input("Ano",  2024, 2030, agora.year, key="gc_ano")
        if st.button("🗓️ Gerar Cronograma", key="btn_gc"):
            n = gerar_cronograma_mensal(int(gc_ano), int(gc_mes), gc_an)
            if n > 0:
                st.success(f"✅ {n} atividades criadas para {gc_an} — {gc_mes:02d}/{gc_ano}!")
            else:
                st.info("Nenhuma atividade nova — todas já existem para este período.")

        # Log de gerações
        cron_log = load_cronograma_gerado()
        if cron_log:
            st.caption("Histórico de gerações")
            st.dataframe(pd.DataFrame(cron_log[-10:]), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# v12 · PÁGINA: RELATÓRIO DE MEDIÇÃO — CRUZAMENTO
# ══════════════════════════════════════════════════════════════════════════════

def page_medicao_cruzamento():
    st.title("Relatório de Medição")
    st.caption("Cruzamento automático: faturas pagas no extrato vs. itens no relatório de medição.")

    tab_cruz, tab_bi = st.tabs(["🔍 Cruzar Extrato vs Medição", "📊 BI do Relatório"])

    with tab_cruz:
        st.markdown("#### Parâmetros Operacionais e Upload")
        comp_alvo = st.text_input("Competência de Pagamento para Filtrar (MM/AAAA)", value="05/2026", help="Filtra pagamentos realizados de 01 a 31 do mês informado", key="mc_comp_target_unique_v12")
        
        c1, c2 = st.columns(2)
        f_ext = c1.file_uploader("📄 Extrato Detalhado (xlsx/csv)", type=["xlsx","xls","csv"], key="mc_ext_uploader_v12")
        
        with c2:
            st.markdown("<p style='font-size:11.5px; font-weight:600; color:var(--text-3); text-transform:uppercase;'>Relatórios de Medição (Até 3 arquivos simultâneos)</p>", unsafe_allow_html=True)
            f_med1 = st.file_uploader("Relatório de Medição 1 (Obrigatório)", type=["xlsx","xls","csv"], key="mc_med1_v12")
            f_med2 = st.file_uploader("Relatório de Medição 2 (Opcional)", type=["xlsx","xls","csv"], key="mc_med2_v12")
            f_med3 = st.file_uploader("Relatório de Medição 3 (Opcional)", type=["xlsx","xls","csv"], key="mc_med3_v12")

        if f_ext and f_med1:
            if st.button("🔁 Cruzar agora", key="btn_executar_cruzamento_v12", use_container_width=True):
                with st.spinner("Limpando cabeçalhos e executando cruzamento analítico..."):
                    try:
                        # Varredura inteligente de cabeçalhos do Extrato Detalhado
                        df_e_raw = pd.read_excel(f_ext, header=None) if not f_ext.name.endswith(".csv") else pd.read_csv(f_ext, header=None, sep=None, engine="python")
                        df_e = df_e_raw.fillna("")
                        for i, row in df_e_raw.iterrows():
                            row_l = [str(c).strip().lower() for c in row]
                            hits = sum(1 for s in row_l if "uc" in s or "unidade" in s or "pagamento" in s or "status" in s or "total" in s)
                            if hits >= 2:
                                df_e.columns = [str(c).strip() for c in row]
                                df_e = df_e.iloc[i+1:].reset_index(drop=True)
                                break

                        # Varredura inteligente de até 3 relatórios de medição simultâneos
                        lista_m = []
                        for f_m in [f_med1, f_med2, f_med3]:
                            if f_m is not None:
                                df_m_raw = pd.read_excel(f_m, header=None) if not f_m.name.endswith(".csv") else pd.read_csv(f_m, header=None, sep=None, engine="python")
                                df_m = df_m_raw.fillna("")
                                for i, row in df_m_raw.iterrows():
                                    row_l = [str(c).strip().lower() for c in row]
                                    hits = sum(1 for s in row_l if "cliente" in s or "nº" in s or "valor" in s or "pago" in s or "uc" in s)
                                    if hits >= 2:
                                        df_m.columns = [str(c).strip() for c in row]
                                        df_m = df_m.iloc[i+1:].reset_index(drop=True)
                                        break
                                lista_m.append(df_m)

                        resultado = cruzar_medicao_extrato(df_e, lista_m, comp_alvo)
                        st.session_state["medicao_cruzamento"] = resultado
                    except Exception as e:
                        st.error(f"Erro: {e}")
                        st.code(traceback.format_exc())

        res = st.session_state.get("medicao_cruzamento")
        if res:
            if "erro" in res:
                st.error(res["erro"])
                return
            if "aviso" in res:
                st.warning(res["aviso"])
                return

            ausentes = res.get("ausentes", [])
            ok_list  = res.get("ok", [])
            extras   = res.get("extras", [])

            kc1, kc2, kc3, kc4 = st.columns(4)
            kc1.markdown(f'<div class="kpi-box"><div class="kpi-value" style="color:#0B7A5F">{len(ok_list)}</div><div class="kpi-label">Faturas OK</div></div>', unsafe_allow_html=True)
            kc2.markdown(f'<div class="kpi-box"><div class="kpi-value" style="color:#C41230">{len(ausentes)}</div><div class="kpi-label">Ausentes no Relatório</div></div>', unsafe_allow_html=True)
            kc3.markdown(f'<div class="kpi-box"><div class="kpi-value" style="color:#A84010">{len(extras)}</div><div class="kpi-label">No Rel. sem pagamento</div></div>', unsafe_allow_html=True)
            kc4.markdown(f'<div class="kpi-box"><div class="kpi-value">R$ {res.get("total_ausente_valor",0):,.2f}</div><div class="kpi-label">Valor ausente</div></div>', unsafe_allow_html=True)

            st.write("")

            if ausentes:
                st.markdown(f'<div class="alert alert-r">⚠️ <b>{len(ausentes)} fatura(s)</b> pagas no extrato de caixa filtrado NÃO constam nos relatórios de medição. É necessário incluí-las.</div>', unsafe_allow_html=True)
                with st.expander(f"Ver {len(ausentes)} faturas ausentes — R$ {res['total_ausente_valor']:,.2f}"):
                    df_aus = pd.DataFrame(ausentes)
                    df_aus["valor"] = df_aus["valor"].apply(lambda x: f"R$ {x:,.2f}")
                    df_aus.columns = [c.capitalize() for c in df_aus.columns]
                    st.dataframe(df_aus, use_container_width=True, hide_index=True)
                    st.download_button("⬇ Exportar ausentes (CSV)", df_aus.to_csv(index=False, sep=";").encode("utf-8-sig"), "faturas_ausentes.csv", "text/csv", key="btn_download_csv_excl_v12")
            else:
                st.markdown('<div class="alert alert-g">✅ Todas as faturas pagas no mês constam nos relatórios de medição informados!</div>', unsafe_allow_html=True)

            if ok_list:
                with st.expander(f"✅ {len(ok_list)} faturas OK"):
                    df_ok = pd.DataFrame(ok_list)
                    df_ok["valor"] = df_ok["valor"].apply(lambda x: f"R$ {x:,.2f}")
                    st.dataframe(df_ok, use_container_width=True, hide_index=True)

    with tab_bi:
        st.markdown("#### BI a partir do Relatório de Medição")
        f_bi = st.file_uploader("Relatório de Medição Sunne (.xlsx)", type=["xlsx","xls"], key="mc_bi_file_unique_v12")
        sel_ger_bi = st.selectbox("Gerador", ["—"] + sorted({g["gerador"] for g in load_geradores()}), key="mc_ger_bi_unique_v12")
        sel_comp_bi = st.text_input("Competência (MM/AAAA)", placeholder="04/2026", key="mc_comp_bi_unique_v12")

        if f_bi and sel_ger_bi != "—" and sel_comp_bi:
            if st.button("📊 Analisar", key="btn_bi_mc_click_v12"):
                with st.spinner("Analisando…"):
                    try:
                        medicao = parse_relatorio_medicao(f_bi.read())
                        ger_obj = next((g for g in load_geradores() if g.get("gerador","").lower() == sel_ger_bi.lower()), {})
                        gerador_cfg = {
                            "pct_desconto_gerador": clean_val(ger_obj.get("pct_desconto_gerador", 0.20)),
                            "pct_taxa_admin":       clean_val(ger_obj.get("pct_taxa_admin",       0.07)),
                            "enquadramento":        ger_obj.get("enquadramento", "GD1"),
                        }
                        ind = calcular_indicadores_bi(medicao, None, gerador_cfg)
                        st.session_state["medicao_bi_ind"]  = ind
                        st.session_state["medicao_bi_comp"] = sel_comp_bi
                        st.session_state["medicao_bi_ger"]  = sel_ger_bi

                        hist_an = load_analises()
                        hist_an.setdefault(sel_ger_bi, {})[sel_comp_bi] = {
                            "indicadores": ind,
                            "salvo_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        }
                        save_analises(hist_an)
                        st.success("Análise salva no histórico!")
                    except Exception as e:
                        st.error(str(e)); st.code(traceback.format_exc())

        ind_bi = st.session_state.get("medicao_bi_ind")
        if ind_bi:
            comp_lbl = st.session_state.get("medicao_bi_comp","—")
            ger_lbl  = st.session_state.get("medicao_bi_ger","—")
            st.markdown(f"#### {ger_lbl} · {comp_lbl}")
            c1,c2,c3,c4 = st.columns(4)
            for col, lbl, val, fmt in [
                (c1,"Fat. Bruto",   ind_bi.get("faturamento_bruto",0),   "R$ {:,.2f}"),
                (c2,"Fat. Líquido", ind_bi.get("faturamento_liquido",0),  "R$ {:,.2f}"),
                (c3,"Inadimplência",ind_bi.get("pct_inadimplencia",0),    "{:.1f}%"),
                (c4,"Efic. Rateio", ind_bi.get("eficiencia_rateio",0),    "{:.1f}%"),
            ]:
                col.markdown(f'<div class="kpi-box"><div class="kpi-value">{fmt.format(val)}</div><div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)
            st.write("")
            por_usina = ind_bi.get("por_usina",[])
            if por_usina:
                st.markdown("**Performance por Usina**")
                df_pu = pd.DataFrame(por_usina)
                df_pu.columns = ["Usina","Fat. Bruto","% Sunne","Tar. Banc.","Fat. Líquido","Conta Energia","Marketplace"]
                st.dataframe(df_pu, use_container_width=True, hide_index=True)
            insights = gerar_insights(ind_bi, None)
            for nivel, msg in insights:
                st.markdown(f'<div class="alert alert-{nivel}">{msg}</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# v12 · PÁGINA: AUDITORIA UFV / UCS
# ══════════════════════════════════════════════════════════════════════════════

def page_auditoria():
    st.title("Auditoria UFV / UCs")
    st.caption("Saúde dos beneficiários e saúde do rateio — alertas automáticos para rebalanceamento.")

    geradores    = load_geradores()
    usinas       = load_usinas()
    nomes_ger    = sorted({g["gerador"] for g in geradores})

    col_ger, col_usi, col_comp = st.columns(3)
    sel_ger  = col_ger.selectbox("Gerador", ["—"] + nomes_ger, key="aud_ger")
    usinas_g = [u for u in usinas if u.get("gerador","").lower() == sel_ger.lower()] if sel_ger != "—" else []
    nomes_usi= [u.get("ufv","") or u["uc"] for u in usinas_g]
    sel_usi  = col_usi.selectbox("Usina", ["—"] + nomes_usi, key="aud_usi") if nomes_usi else col_usi.selectbox("Usina", ["—"], key="aud_usi_vz")
    sel_comp = col_comp.text_input("Competência base (MM/AAAA)", placeholder="04/2026", key="aud_comp")

    f_aud = st.file_uploader(
        "📊 Planilha de Auditoria (.xlsx — formato Sunne)",
        type=["xlsx","xls"], key="aud_file",
        help="Suba a planilha de auditoria no padrão Sunne (Nafisah UFV7 / similar)")

    if f_aud:
        if st.button("🔬 Analisar Auditoria", key="btn_aud", use_container_width=True):
            with st.spinner("Processando planilha de auditoria…"):
                try:
                    resultado_aud = parse_auditoria_planilha(f_aud.read())
                    st.session_state["auditoria_resultado"] = resultado_aud
                    st.session_state["auditoria_usina"]     = sel_usi
                    st.session_state["auditoria_gerador"]   = sel_ger
                    st.session_state["auditoria_comp"]      = sel_comp
                except Exception as e:
                    st.error(str(e)); st.code(traceback.format_exc())

    res_aud = st.session_state.get("auditoria_resultado")
    if not res_aud:
        st.info("Faça upload da planilha e clique em Analisar para ver os resultados.")
        return

    ger_lbl  = st.session_state.get("auditoria_gerador","—")
    usi_lbl  = st.session_state.get("auditoria_usina","—")
    comp_lbl = st.session_state.get("auditoria_comp","—")

    st.markdown(f"### {ger_lbl} · {usi_lbl} · {comp_lbl}")
    st.caption(f"{res_aud['meses_analisados']} meses analisados")

    # ── KPIs da usina ─────────────────────────────────────────────────────────
    rs = res_aud.get("usina_resumo", {})
    kc1, kc2, kc3, kc4 = st.columns(4)
    for col, lbl, val, fmt in [
        (kc1, "Geração total",   rs.get("geracao_total",  0), "{:,.0f} kWh"),
        (kc2, "Consumo total",   rs.get("consumo_total",  0), "{:,.0f} kWh"),
        (kc3, "Créditos totais", rs.get("creditos_total", 0), "{:,.0f} kWh"),
        (kc4, "Saldo atual",     rs.get("saldo_atual",    0), "{:,.0f} kWh"),
    ]:
        col.markdown(
            f'<div class="kpi-box"><div class="kpi-value" style="font-size:1.4rem">{fmt.format(val)}</div>'
            f'<div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)

    st.write("")

    # ── Alertas do rateio ─────────────────────────────────────────────────────
    alertas_rat = res_aud.get("alertas_rateio", [])
    if alertas_rat:
        st.markdown('<div class="sec-head"><span class="sec-head-title">🏭 Saúde do Rateio</span><div class="sec-head-line"></div></div>', unsafe_allow_html=True)
        for alrt in alertas_rat:
            cls = "alert-r" if alrt["tipo"] == "critico" else "alert-y"
            st.markdown(f'<div class="alert {cls}">{alrt["mensagem"]}</div>', unsafe_allow_html=True)

        # Botão: criar tarefa de rateio automático
        if any(a["tipo"] == "critico" for a in alertas_rat):
            if st.button("📝 Criar tarefa de Atualização de Rateio", key="btn_aud_rateio"):
                analista = st.session_state["user"]["name"]
                uc_usina = next((u["uc"] for u in usinas if (u.get("ufv","") or u["uc"]) == usi_lbl), "")
                desc_auto = (
                    f"Auditoria {comp_lbl} — Alertas:\n"
                    + "\n".join(f"- {a['mensagem']}" for a in alertas_rat)
                )
                new_task(
                    titulo          = f"Atualizar Rateio — {usi_lbl} ({comp_lbl})",
                    usina           = uc_usina,
                    gerador         = ger_lbl,
                    analista        = analista,
                    tipo            = "Rateio",
                    agendamento     = comp_lbl,
                    descricao       = desc_auto,
                    data_programada = datetime.now().strftime("%d/%m/%Y"),
                )
                st.success("✅ Tarefa de atualização de rateio criada automaticamente!")
    else:
        st.markdown('<div class="alert alert-g">✅ Saúde do rateio OK — nenhuma anomalia detectada.</div>',
                    unsafe_allow_html=True)

    # ── Alertas dos beneficiários ──────────────────────────────────────────────
    alertas_ben = res_aud.get("alertas_ben", [])
    st.write("")
    st.markdown('<div class="sec-head"><span class="sec-head-title">👥 Saúde dos Beneficiários</span><div class="sec-head-line"></div></div>', unsafe_allow_html=True)

    if alertas_ben:
        alts_saldo  = [a for a in alertas_ben if a["tipo"] == "saldo_alto"]
        alts_menos  = [a for a in alertas_ben if a["tipo"] == "recebendo_menos"]

        if alts_saldo:
            st.markdown(f'<div class="alert alert-y">💰 <b>{len(alts_saldo)} beneficiário(s)</b> com saldo acumulado alto (> 3 meses). Reduzir % de rateio.</div>',
                        unsafe_allow_html=True)
            with st.expander("Ver UCs com saldo alto"):
                for a in alts_saldo:
                    st.markdown(f"· {a['mensagem']}")

        if alts_menos:
            st.markdown(f'<div class="alert alert-r">📉 <b>{len(alts_menos)} beneficiário(s)</b> recebendo menos energia do que consomem. Aumentar % de rateio.</div>',
                        unsafe_allow_html=True)
            with st.expander("Ver UCs recebendo menos"):
                for a in alts_menos:
                    st.markdown(f"· {a['mensagem']}")
    else:
        st.markdown('<div class="alert alert-g">✅ Todos os beneficiários dentro dos parâmetros esperados.</div>',
                    unsafe_allow_html=True)

    # ── Tabela completa de beneficiários ──────────────────────────────────────
    bens = res_aud.get("beneficiarios", [])
    if bens:
        st.write("")
        with st.expander(f"📋 Ver todos os {len(bens)} beneficiários"):
            df_ben = pd.DataFrame(bens)
            df_ben.columns = [c.capitalize() for c in df_ben.columns]
            if "Autonomia" in df_ben.columns:
                def _color_aut(val):
                    try:
                        v = float(val)
                        if v > 3:  return "background:#FFFBEC;color:#856404"
                        if v < 0.5 and v > 0: return "background:#FFF0F3;color:#8B1530"
                        return ""
                    except Exception: return ""
                st.dataframe(df_ben.style.applymap(_color_aut, subset=["Autonomia"]),
                             use_container_width=True, hide_index=True)
            else:
                st.dataframe(df_ben, use_container_width=True, hide_index=True)

    # ── Base do último rateio ─────────────────────────────────────────────────
    rateio_base = res_aud.get("rateio_base", [])
    if rateio_base:
        st.write("")
        with st.expander(f"📂 Base do último rateio ({len(rateio_base)} UCs)"):
            df_rb = pd.DataFrame(rateio_base)
            df_rb.columns = [c.capitalize() for c in df_rb.columns]
            st.dataframe(df_rb, use_container_width=True, hide_index=True)
            # Exportar planilha de rateio pré-preenchida
            st.download_button(
                "⬇ Exportar base do rateio (.xlsx)",
                df_to_excel_bytes(df_rb),
                f"rateio_base_{usi_lbl}_{comp_lbl.replace('/','_')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

# ── DATA UTILS ────────────────────────────────────────────────────────────────
def normalize_uc(val):
    """Remove tudo que não é dígito, ignora zeros à esquerda."""
    if not val: return ""
    s = "".join(filter(str.isdigit, str(val).strip().split('.')[0]))
    return s.lstrip("0") or s   # preserva "0" sozinho

def clean_val(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return 0.0
    s = str(v).replace("R$","").replace(" ","").replace("\xa0","").strip()
    if not s or s in ("-","—","nan","NaN","None",""): return 0.0
    s = s.replace("%","")
    if "," in s and "." in s:
        if s.index(".") < s.index(","): s = s.replace(".","").replace(",",".")
        else: s = s.replace(",","")
    elif "," in s: s = s.replace(",",".")
    try: return float(s)
    except: return 0.0

def df_to_excel_bytes(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False)
    return buf.getvalue()

def csv_from_list(rows, cols, headers):
    out = io.StringIO(); df = pd.DataFrame(rows)
    if not df.empty:
        ex = df[[c for c in cols if c in df.columns]]
        ex.columns = headers[:len(ex.columns)]
        ex.to_csv(out, index=False, sep=';', encoding='utf-8-sig')
    return out.getvalue().encode('utf-8-sig')

def comp_sort_key(comp_str):
    try:
        p = str(comp_str).strip().split("/")
        if len(p) == 2: return (int(p[1]), int(p[0]))
    except: pass
    return (9999, 99)

def load_planilha(file):
    if file is None: return None
    try:
        df = (pd.read_excel(file, header=None) if not file.name.endswith('.csv')
              else pd.read_csv(file, header=None, sep=None, engine='python'))
        for i, row in df.head(20).iterrows():
            row_l = [str(c).strip().lower() for c in row]
            if any("uc nova" in s or "número da uc" in s for s in row_l):
                df.columns = [str(c).strip() for c in row]
                df = df.iloc[i+1:].reset_index(drop=True)
                break
        df.columns = [str(c).strip() for c in df.columns]
        return df.dropna(how='all').fillna("")
    except: return None

def get_usinas_do_gerador(nome_ger):
    return [u for u in load_usinas()
            if u.get("gerador","").strip().lower() == nome_ger.strip().lower()]

def style_critical(row):
    if row.get('Dias de Atraso', 0) > 90:
        return ['background:#ffcccc;color:#990000;font-weight:bold'] * len(row)
    return ['background:#fff4cc;color:#856404;font-weight:bold'] * len(row)

# ── ANÁLISE DE FATURAMENTO (original) ────────────────────────────────────────
def analyze_performance(df_r, df_e):
    uc_r_col    = next((c for c in df_r.columns if "UC Nova"       in c), df_r.columns[0])
    uc_e_col    = next((c for c in df_e.columns if "Número da UC"  in c), df_e.columns[0])
    comp_col    = next((c for c in df_e.columns if "Competência"   in c), None)
    status_col  = next((c for c in df_e.columns if "Status"        in c), None)
    valor_col   = next((c for c in df_e.columns if "Total a Pagar" in c), None)
    titular_col = next((c for c in df_e.columns if "Titular"       in c), None)
    venc_col    = next((c for c in df_e.columns if "Vencimento"    in c), None)

    df_r["UC_NORM"] = df_r[uc_r_col].apply(normalize_uc)
    df_e["UC_NORM"] = df_e[uc_e_col].apply(normalize_uc)

    missing_res={}; inad_res={}; t_gerado={}; t_pago={}; t_vencido={}; critical_inad=[]
    hoje = datetime.now()
    if venc_col:
        df_e[venc_col] = pd.to_datetime(df_e[venc_col], errors='coerce', dayfirst=True)

    for _, row in df_e.iterrows():
        comp   = str(row[comp_col])           if comp_col   else "Geral"
        status = str(row[status_col]).lower() if status_col else ""
        valor  = clean_val(row[valor_col])
        venc   = row[venc_col]                if venc_col   else None

        t_gerado[comp] = t_gerado.get(comp, 0.0) + valor
        if "pago" in status:
            t_pago[comp] = t_pago.get(comp, 0.0) + valor
        if "vencido" in status:
            t_vencido[comp] = t_vencido.get(comp, 0.0) + valor
            item = {"uc":row[uc_e_col], "valor":valor,
                    "titular":row[titular_col] if titular_col else "—"}
            inad_res.setdefault(comp, []).append(item)
            if pd.notnull(venc):
                dias = (hoje - venc).days
                if dias > 60:
                    critical_inad.append({
                        "Titular":item["titular"], "UC":item["uc"],
                        "Vencimento":venc.strftime('%d/%m/%Y'),
                        "Dias de Atraso":dias, "Valor":valor, "Mês Ref":comp,
                    })

    critical_inad = sorted(critical_inad, key=lambda x: x['Dias de Atraso'], reverse=True)
    extrato_set = set(zip(df_e["UC_NORM"], df_e[comp_col].astype(str)))
    ucs_rateio  = df_r["UC_NORM"].unique()

    for comp in df_e[comp_col].unique():
        if not comp or str(comp).lower() == 'nan': continue
        for uc in ucs_rateio:
            if (uc, str(comp)) not in extrato_set:
                r = df_r[df_r["UC_NORM"] == uc].iloc[0]
                missing_res.setdefault(comp, []).append({
                    "uc":r[uc_r_col], "apelido":r.get("Apelido UC","—"), "usina":r.get("Usina","—"),
                })

    return {"missing":missing_res, "inad":inad_res, "t_gerado":t_gerado,
            "t_pago":t_pago, "t_vencido":t_vencido, "critical_inad":critical_inad}


# ── RELATÓRIO DE MEDIÇÃO PARSER ───────────────────────────────────────────────
def parse_relatorio_medicao(file_bytes_or_path):
    """Extrai dados estruturados do Relatório de Medição Sunne."""
    result = {
        "faturamento_bruto": 0.0,
        "percentual_sunne": 0.0,
        "tarifas_bancarias": 0.0,
        "faturamento_liquido": 0.0,
        "outros": 0.0,
        "marketplace": 0.0,
        "por_usina": [],
        "detalhado": [],
        "historico_geral": [],
        "historico_por_usina": [],
    }
    try:
        if isinstance(file_bytes_or_path, bytes):
            xl = pd.ExcelFile(io.BytesIO(file_bytes_or_path))
        else:
            xl = pd.ExcelFile(file_bytes_or_path)

        # ── Tabela Resumida ──
        if "Tabela Resumida" in xl.sheet_names:
            df_res = pd.read_excel(xl, sheet_name="Tabela Resumida", header=None)
            for _, row in df_res.iterrows():
                label = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ""
                val   = clean_val(row.iloc[1]) if len(row) > 1 else 0.0
                if "faturamento bruto" in label:       result["faturamento_bruto"] = val
                elif "percentual sunne" in label:      result["percentual_sunne"]  = val
                elif "tarifas bancárias" in label or "tarifas bancarias" in label:
                    result["tarifas_bancarias"] = val
                elif "faturamento líquido" in label or "faturamento liquido" in label:
                    result["faturamento_liquido"] = val
                elif label == "outros":                result["outros"] = val

        # ── Resumo por Usina ──
        if "Resumo - Visão Por Usina" in xl.sheet_names:
            df_u = pd.read_excel(xl, sheet_name="Resumo - Visão Por Usina")
            df_u.columns = [str(c).strip() for c in df_u.columns]
            rows = []
            for _, row in df_u.iterrows():
                usina = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                if not usina: continue
                rows.append({
                    "usina":    usina,
                    "fat_bruto": clean_val(row.iloc[1]) if len(row) > 1 else 0.0,
                    "pct_sunne": clean_val(row.iloc[2]) if len(row) > 2 else 0.0,
                    "tar_banc":  clean_val(row.iloc[3]) if len(row) > 3 else 0.0,
                    "fat_liq":   clean_val(row.iloc[4]) if len(row) > 4 else 0.0,
                    "conta_energia": clean_val(row.iloc[5]) if len(row) > 5 else 0.0,
                    "marketplace":   clean_val(row.iloc[6]) if len(row) > 6 else 0.0,
                })
            result["por_usina"] = rows

        # ── Tabela Detalhada ──
        if "Tabela Detalhada" in xl.sheet_names:
            df_det = pd.read_excel(xl, sheet_name="Tabela Detalhada")
            df_det.columns = [str(c).strip() for c in df_det.columns]
            rows = []
            for _, row in df_det.iterrows():
                rows.append({k: clean_val(v) if isinstance(v, (int, float)) else str(v)
                             for k, v in row.items()})
            result["detalhado"] = rows

        # ── Histórico Geral ──
        if "Histórico - Visão Geral" in xl.sheet_names:
            df_hg = pd.read_excel(xl, sheet_name="Histórico - Visão Geral")
            df_hg.columns = [str(c).strip() for c in df_hg.columns]
            result["historico_geral"] = df_hg.fillna(0).to_dict(orient="records")

        # ── Histórico Por Usina ──
        if "Histórico - Visão Por Usina" in xl.sheet_names:
            df_hu = pd.read_excel(xl, sheet_name="Histórico - Visão Por Usina")
            df_hu.columns = [str(c).strip() for c in df_hu.columns]
            result["historico_por_usina"] = df_hu.fillna(0).to_dict(orient="records")

    except Exception as e:
        result["erro"] = str(e)
    return result


# ── MOTOR BI: CÁLCULO DE INDICADORES ─────────────────────────────────────────
def calcular_indicadores_bi(medicao, extrato_df, gerador_cfg):
    """
    medicao      : dict do parse_relatorio_medicao
    extrato_df   : DataFrame do extrato detalhado (pode ser None)
    gerador_cfg  : dict com 'pct_desconto_gerador' e 'pct_taxa_admin'
    """
    ind = {}

    # Faturamento bruto do relatório de medição
    ind["faturamento_bruto"]   = medicao.get("faturamento_bruto", 0.0)
    ind["faturamento_liquido"] = medicao.get("faturamento_liquido", 0.0)
    ind["percentual_sunne"]    = medicao.get("percentual_sunne", 0.0)
    ind["tarifas_bancarias"]   = medicao.get("tarifas_bancarias", 0.0)

    # Retorno por usina
    ind["por_usina"] = medicao.get("por_usina", [])

    if extrato_df is not None and not extrato_df.empty:
        df = extrato_df.copy()
        # Normaliza colunas
        col_status  = next((c for c in df.columns if "Status"           in c), None)
        col_credito = next((c for c in df.columns if "Créditos"         in c and "Utiliz" in c), None)
        col_tarifa  = next((c for c in df.columns if "Tarifa Compensável" in c), None)
        col_ener    = next((c for c in df.columns if "Energia Utilizada" in c), None)
        col_venc    = next((c for c in df.columns if "Vencimento"        in c), None)
        col_valor   = next((c for c in df.columns if "Total a Pagar"     in c), None)

        # Converte numéricos
        for col in [col_credito, col_tarifa, col_ener, col_valor]:
            if col: df[col] = df[col].apply(clean_val)

        mask_ativo = pd.Series([True] * len(df))
        if col_status:
            mask_ativo = df[col_status].astype(str).str.lower().isin(["pago","vencido"])

        # Créditos utilizados
        if col_credito:
            ind["creditos_utilizados"] = float(df.loc[mask_ativo, col_credito].sum())
        else:
            ind["creditos_utilizados"] = 0.0

        # Energia utilizada na fatura
        if col_ener:
            ind["energia_utilizada_fatura"] = float(df.loc[mask_ativo, col_ener].sum())
        else:
            ind["energia_utilizada_fatura"] = 0.0

        # Média tarifa compensável
        if col_tarifa:
            vals = df.loc[mask_ativo, col_tarifa]
            vals = vals[vals > 0]
            ind["media_tarifa_compensavel"] = float(vals.mean()) if len(vals) > 0 else 0.0
        else:
            ind["media_tarifa_compensavel"] = 0.0

        # Inadimplência da competência
        if col_status and col_valor:
            mask_venc = df[col_status].astype(str).str.lower() == "vencido"
            ind["inadimplencia_valor"] = float(df.loc[mask_venc, col_valor].sum())
            ind["total_faturado"]      = float(df.loc[mask_ativo, col_valor].sum())
            ind["pct_inadimplencia"]   = (ind["inadimplencia_valor"] / ind["total_faturado"] * 100
                                          if ind["total_faturado"] > 0 else 0.0)
        else:
            ind["inadimplencia_valor"] = 0.0
            ind["total_faturado"]      = 0.0
            ind["pct_inadimplencia"]   = 0.0
    else:
        ind["creditos_utilizados"]      = 0.0
        ind["energia_utilizada_fatura"] = 0.0
        ind["media_tarifa_compensavel"] = 0.0
        ind["inadimplencia_valor"]      = 0.0
        ind["total_faturado"]           = 0.0
        ind["pct_inadimplencia"]        = 0.0

    # Eficiência de rateio (créditos / faturamento bruto proxy)
    if ind["faturamento_bruto"] > 0 and ind["creditos_utilizados"] > 0:
        ind["eficiencia_rateio"] = round(ind["creditos_utilizados"] / ind["faturamento_bruto"] * 100, 2)
    else:
        ind["eficiencia_rateio"] = 0.0

    # Cálculo de retorno estimado (Fórmula Sunne)
    pct_desc  = gerador_cfg.get("pct_desconto_gerador", 0.0)   # ex: 0.20
    pct_admin = gerador_cfg.get("pct_taxa_admin", 0.0)         # ex: 0.07
    tarifa_comp = ind["media_tarifa_compensavel"]

    # Tarifa Retorno = Tarifa Compensável * (1 - % Desconto Gerador) * (1 - % Taxa Admin)
    ind["tarifa_retorno"] = tarifa_comp * (1 - pct_desc) * (1 - pct_admin)

    # Retorno Bruto = Tarifa Retorno * Créditos Utilizados (ou faturamento bruto como proxy)
    base_kwh = ind["creditos_utilizados"] if ind["creditos_utilizados"] > 0 else ind["faturamento_bruto"]
    ind["retorno_bruto_estimado"] = ind["tarifa_retorno"] * base_kwh

    return ind


# ── INSIGHTS AUTOMÁTICOS ──────────────────────────────────────────────────────
def gerar_insights(ind_atual, ind_anterior):
    """Retorna lista de (nivel, mensagem) onde nivel é 'r', 'y', 'g'."""
    insights = []
    if ind_anterior:
        # Inadimplência subiu > 5%?
        inad_ant = ind_anterior.get("pct_inadimplencia", 0.0)
        inad_atu = ind_atual.get("pct_inadimplencia", 0.0)
        if inad_atu - inad_ant > 5:
            insights.append(("r", f"⚠️ Inadimplência subiu {inad_atu - inad_ant:.1f}pp vs mês anterior ({inad_ant:.1f}% → {inad_atu:.1f}%)"))

        # Faturamento caiu > 10%?
        fat_ant = ind_anterior.get("faturamento_bruto", 0.0)
        fat_atu = ind_atual.get("faturamento_bruto", 0.0)
        if fat_ant > 0 and (fat_ant - fat_atu) / fat_ant > 0.10:
            queda = (fat_ant - fat_atu) / fat_ant * 100
            insights.append(("y", f"📉 Faturamento bruto caiu {queda:.1f}% vs mês anterior (R$ {fat_ant:,.2f} → R$ {fat_atu:,.2f})"))

    # Vacância (eficiência < 95%)
    efic = ind_atual.get("eficiencia_rateio", 0.0)
    if 0 < efic < 95:
        insights.append(("y", f"🔄 Eficiência de rateio {efic:.1f}% < 95% — considere rebalanceamento urgente"))

    if ind_atual.get("pct_inadimplencia", 0.0) > 10:
        insights.append(("r", f"🚨 Inadimplência crítica: {ind_atual['pct_inadimplencia']:.1f}% do faturado"))

    if not insights:
        insights.append(("g", "✅ Todos os indicadores dentro dos parâmetros esperados"))

    return insights


# ═══════════════════════════════════════════════════════════════════
# DIALOGS (st.dialog — Streamlit ≥ 1.36)
# ═══════════════════════════════════════════════════════════════════

@st.dialog("📝 Nova Atividade", width="large")
def dialog_nova_atividade(uc_pre="", ger_pre=""):
    analista = st.session_state["user"]["name"]
    with st.form("form_nova_ativ", clear_on_submit=True):
        c1, c2 = st.columns(2)
        uv = c1.text_input("Usina",   value=uc_pre,  disabled=bool(uc_pre))
        gv = c2.text_input("Gerador", value=ger_pre, disabled=bool(ger_pre))
        titulo = st.text_input("Título da atividade *")
        d1, d2, d3 = st.columns(3)
        tipo  = d1.selectbox("Tipo", TIPOS)
        agend = d2.text_input("Agendamento", placeholder="20/05/2026 09:00")
        data_prog = d3.text_input("📅 Data Programada", placeholder="20/05/2026",
                                  help="Data limite para execução — o sistema alertará se passar sem conclusão")
        desc  = st.text_area("Descrição / Contexto", height=90,
                             placeholder="Links HubSpot, tickets, notas…")
        anex  = st.file_uploader("📎 Anexo (PDF/Excel — opcional)", type=["pdf","xlsx","xls"])
        col_s, col_c = st.columns(2)
        ok   = col_s.form_submit_button("✅ Criar Atividade", use_container_width=True)
        cncl = col_c.form_submit_button("Cancelar",           use_container_width=True)

    if ok:
        if not titulo.strip():
            st.warning("Título obrigatório.")
        else:
            an, ab = "", ""
            if anex:
                an = anex.name
                ab = base64.b64encode(anex.read()).decode()
            new_task(titulo.strip(), uv, gv, analista, tipo, agend, desc, an, ab,
                     data_programada=data_prog.strip())
            st.toast("✅ Atividade criada!", icon="🎉")
            st.success("Atividade criada! Acesse a seção **Atividades** no menu.")
            if st.button("📋 Ver em Atividades"):
                st.session_state["page"] = "atividades"
                st.rerun()
    if cncl:
        st.rerun()


@st.dialog("Detalhes da Atividade", width="large")
def dialog_task_detail(tid):
    tasks = load_tasks()
    t = next((x for x in tasks if x["id"] == tid), None)
    if not t:
        st.error("Tarefa não encontrada."); return

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    atrasada = task_esta_atrasada(t)
    hoje_tag  = task_vence_hoje(t)
    tag_html  = ""
    if atrasada:
        tag_html = '<span class="kb-tag-atrasada">⏰ TAREFA ATRASADA</span>'
    elif hoje_tag:
        tag_html = '<span class="kb-tag-hoje">🔔 VENCE HOJE</span>'

    st.markdown(f"### {t['titulo']} {tag_html}", unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    c1.markdown(f"**Usina** · {t.get('usina','—')}")
    c2.markdown(f"**Gerador** · {t.get('gerador','—')}")
    c3.markdown(f"**Tipo** · {t.get('tipo','—')}")
    d1, d2 = st.columns(2)
    d1.markdown(f"**Analista** · {t.get('analista','—')}")
    d2.markdown(f"**Criado** · {t.get('criado_em','—')}")

    dp = t.get("data_programada","")
    if dp:
        cor_dp = "#C41230" if atrasada else ("#F36E21" if hoje_tag else "#0B7A5F")
        st.markdown(f'<span style="color:{cor_dp};font-weight:600">📅 Data Programada · {dp}</span>',
                    unsafe_allow_html=True)

    if t.get("agendamento"): st.markdown(f"**⏰ Agendamento** · {t['agendamento']}")
    if t.get("motivo_bloqueio"): st.markdown(f"**🔒 Motivo** · {t['motivo_bloqueio']}")
    if t.get("descricao","").strip():
        with st.expander("Descrição original"):
            st.write(t["descricao"])

    # ── Tabs internas do card ─────────────────────────────────────────────────
    tc_log, tc_prog, tc_email = st.tabs(["📝 Log", "📅 Programação", "📧 E-mail"])

    with tc_log:
        nova_obs = st.text_area("Nova observação", height=80, key=f"obs_txt_{tid}",
                                placeholder="Ex: Ligou 14/05 — aguardando doc. Ticket: https://…",
                                label_visibility="collapsed")
        if st.button("Registrar observação", key=f"btn_obs_{tid}"):
            if nova_obs.strip():
                ts  = datetime.now().strftime("%d/%m/%Y %H:%M")
                an  = st.session_state["user"]["name"]
                log = t.get("observacoes","") + f"\n[{ts} — {an}]\n{nova_obs.strip()}\n"
                update_task(tid, observacoes=log)
                st.success("Observação salva."); st.rerun()
        if t.get("observacoes","").strip():
            with st.expander("Ver histórico"):
                st.text(t["observacoes"].strip())
        if t.get("historico"):
            with st.expander("🔄 Movimentações"):
                for h in t["historico"]:
                    st.write(f"· {h['em']}  `{h['de']}` → `{h['para']}`")
        if t.get("anexo_nome") and t.get("anexo_b64"):
            st.download_button(f"📎 {t['anexo_nome']}", base64.b64decode(t["anexo_b64"]),
                               t["anexo_nome"], key=f"dl_anex_{tid}")

    with tc_prog:
        st.markdown("**Alterar Data Programada**")
        nova_dp = st.text_input("Nova data (DD/MM/AAAA)", value=dp,
                                placeholder="20/05/2026", key=f"dp_inp_{tid}")
        if st.button("💾 Salvar data", key=f"dp_save_{tid}"):
            update_task(tid, data_programada=nova_dp.strip())
            st.success("Data programada atualizada!"); st.rerun()

        st.divider()
        st.markdown("**Integração Google Calendar**")
        goog = load_google_tokens()
        if goog.get("access_token"):
            st.success("✅ Google conectado")
            if st.button("📅 Criar evento no Google Calendar", key=f"gcal_{tid}"):
                _criar_evento_google_calendar(t, goog)
        else:
            st.info("Conecte o Google na seção **Integrações** do Dashboard para criar eventos no Calendar.")

    with tc_email:
        st.markdown("**Enviar e-mail para o Gerador**")
        goog = load_google_tokens()
        geradores = load_geradores()
        ger_obj = next((g for g in geradores
                        if g.get("gerador","").lower() == t.get("gerador","").lower()), None)
        email_dest_default = ger_obj.get("contato","") if ger_obj else ""

        if not goog.get("access_token"):
            st.info("Conecte o Google na seção **Integrações** para enviar e-mail pelo Hub.")
        else:
            ep1, ep2 = st.columns(2)
            email_para = ep1.text_input("Para", value=email_dest_default, key=f"email_para_{tid}")
            email_assunto = ep2.text_input("Assunto",
                value=f"[Sunne] {t.get('titulo','')} — {t.get('gerador','')}",
                key=f"email_assunto_{tid}")
            email_corpo = st.text_area("Corpo do e-mail", height=150, key=f"email_corpo_{tid}",
                placeholder="Olá,\n\nEntramos em contato a respeito de...")
            if st.button("📤 Enviar via Gmail", key=f"email_send_{tid}", use_container_width=True):
                if not email_para.strip():
                    st.warning("Informe o destinatário.")
                else:
                    ok_mail = _enviar_gmail(goog, email_para, email_assunto, email_corpo, tid)
                    if ok_mail:
                        ts  = datetime.now().strftime("%d/%m/%Y %H:%M")
                        an  = st.session_state["user"]["name"]
                        log = t.get("observacoes","") + f"\n[{ts} — {an}]\n📧 E-mail enviado para {email_para}: {email_assunto}\n"
                        update_task(tid, observacoes=log)
                        st.success(f"✅ E-mail enviado para {email_para}!")
                    else:
                        st.error("Falha ao enviar. Verifique a conexão Google.")

    # ── Mover status ──────────────────────────────────────────────────────────
    st.divider()
    st.markdown("**↗ Mover para:**")
    opcoes = [s for s in STATUS_LIST if s != t["status"]]
    dest   = st.selectbox("", opcoes, key=f"dest_d_{tid}", label_visibility="collapsed")
    mot    = ""
    if dest in MOTIVO_OBRIG:
        mot = st.text_area("Motivo obrigatório *", key=f"mot_d_{tid}")
    if st.button("Confirmar movimentação", key=f"mv_d_{tid}", use_container_width=True):
        if dest in MOTIVO_OBRIG and not mot.strip():
            st.warning("O motivo é obrigatório.")
        else:
            update_task(tid, status=dest, motivo_bloqueio=mot.strip())
            st.rerun()


@st.dialog("📊 Dashboard BI — Competência", width="large")
def dialog_bi_dashboard(analise_key, comp):
    hist = load_analises()
    dados = hist.get(analise_key, {})
    if not dados:
        st.error("Dados não encontrados."); return

    comp_data = dados.get(comp)
    if not comp_data:
        st.error(f"Sem dados para {comp}."); return

    ind = comp_data.get("indicadores", {})
    st.markdown(f"#### {analise_key} · {comp}")

    # KPIs principais
    col1, col2, col3, col4 = st.columns(4)
    for col, lbl, val, fmt in [
        (col1, "Faturamento Bruto",    ind.get("faturamento_bruto",0),     "R$ {:,.2f}"),
        (col2, "Faturamento Líquido",  ind.get("faturamento_liquido",0),    "R$ {:,.2f}"),
        (col3, "Inadimplência",        ind.get("pct_inadimplencia",0),      "{:.1f}%"),
        (col4, "Eficiência Rateio",    ind.get("eficiencia_rateio",0),      "{:.1f}%"),
    ]:
        col.markdown(
            f'<div class="kpi-box"><div class="kpi-value">{fmt.format(val)}</div>'
            f'<div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)

    st.write("")

    col5, col6, col7 = st.columns(3)
    for col, lbl, val, fmt in [
        (col5, "Créditos Utilizados",  ind.get("creditos_utilizados",0),    "{:,.2f} kWh"),
        (col6, "Média Tarifa Comp.",   ind.get("media_tarifa_compensavel",0),"R$ {:.6f}"),
        (col7, "Retorno Bruto Est.",   ind.get("retorno_bruto_estimado",0),  "R$ {:,.2f}"),
    ]:
        col.markdown(
            f'<div class="kpi-box"><div class="kpi-value">{fmt.format(val)}</div>'
            f'<div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)

    st.write("")

    # Por usina
    por_usina = ind.get("por_usina", [])
    if por_usina:
        st.markdown("**Performance por Usina**")
        df_pu = pd.DataFrame(por_usina)
        df_pu.columns = ["Usina","Fat. Bruto","% Sunne","Tar. Banc.","Fat. Líquido","Conta Energia","Marketplace"]
        st.dataframe(df_pu, use_container_width=True, hide_index=True)

    # Insights
    prev_comp = comp_data.get("indicadores_anterior")
    insights  = gerar_insights(ind, prev_comp)
    st.markdown("**Insights Automáticos**")
    for nivel, msg in insights:
        st.markdown(f'<div class="alert alert-{nivel}">{msg}</div>', unsafe_allow_html=True)

    # Export PDF
    st.divider()
    if st.button("📄 Exportar Relatório Excel"):
        rows = [
            {"Indicador":"Faturamento Bruto",         "Valor":f"R$ {ind.get('faturamento_bruto',0):,.2f}"},
            {"Indicador":"Faturamento Líquido",        "Valor":f"R$ {ind.get('faturamento_liquido',0):,.2f}"},
            {"Indicador":"Créditos Utilizados",        "Valor":f"{ind.get('creditos_utilizados',0):,.2f} kWh"},
            {"Indicador":"Eficiência de Rateio",       "Valor":f"{ind.get('eficiencia_rateio',0):.2f}%"},
            {"Indicador":"Média Tarifa Compensável",   "Valor":f"R$ {ind.get('media_tarifa_compensavel',0):.6f}"},
            {"Indicador":"Inadimplência (R$)",         "Valor":f"R$ {ind.get('inadimplencia_valor',0):,.2f}"},
            {"Indicador":"Inadimplência (%)",          "Valor":f"{ind.get('pct_inadimplencia',0):.1f}%"},
            {"Indicador":"Tarifa Retorno Estimada",    "Valor":f"R$ {ind.get('tarifa_retorno',0):.6f}"},
            {"Indicador":"Retorno Bruto Estimado",     "Valor":f"R$ {ind.get('retorno_bruto_estimado',0):,.2f}"},
        ]
        df_exp = pd.DataFrame(rows)
        if por_usina:
            df_pu2 = pd.DataFrame(por_usina)
            df_pu2.columns = ["Usina","Fat. Bruto","% Sunne","Tar. Banc.","Fat. Líquido","Conta Energia","Marketplace"]
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as w:
                df_exp.to_excel(w, index=False, sheet_name="Indicadores")
                df_pu2.to_excel(w, index=False, sheet_name="Por Usina")
            st.download_button("⬇ Baixar Excel", buf.getvalue(),
                               f"relatorio_bi_{analise_key}_{comp.replace('/','-')}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.download_button("⬇ Baixar Excel", df_to_excel_bytes(df_exp),
                               f"relatorio_bi_{analise_key}_{comp.replace('/','-')}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════════
# GOOGLE INTEGRATIONS (Calendar + Gmail)
# ═══════════════════════════════════════════════════════════════════

def _google_auth_url() -> str:
    """Gera URL de autorização Google. Client ID configurado 1x nos secrets pela TI."""
    import urllib.parse
    try:
        client_id = st.secrets["google"]["client_id"]
    except Exception:
        client_id = os.environ.get("GOOGLE_CLIENT_ID","")
    if not client_id:
        return ""
    redirect = os.environ.get("GOOGLE_REDIRECT_URI","https://performance-sunne.streamlit.app")
    scopes = " ".join([
        "openid",
        "https://www.googleapis.com/auth/userinfo.email",
        "https://www.googleapis.com/auth/userinfo.profile",
        "https://www.googleapis.com/auth/calendar.readonly",
        "https://www.googleapis.com/auth/gmail.send",
    ])
    params = urllib.parse.urlencode({
        "client_id": client_id, "redirect_uri": redirect,
        "response_type": "code", "scope": scopes,
        "access_type": "offline", "prompt": "consent",
    })
    return f"https://accounts.google.com/o/oauth2/v2/auth?{params}"

def _google_exchange_code(code: str) -> dict:
    """Troca o code de autorização por access_token + busca nome/email do usuário."""
    import urllib.request, urllib.parse
    try:
        client_id     = st.secrets["google"]["client_id"]
        client_secret = st.secrets["google"]["client_secret"]
    except Exception:
        return {}
    redirect = os.environ.get("GOOGLE_REDIRECT_URI","https://performance-sunne.streamlit.app")
    data = urllib.parse.urlencode({
        "code": code, "client_id": client_id, "client_secret": client_secret,
        "redirect_uri": redirect, "grant_type": "authorization_code"
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data, method="POST")
    req.add_header("Content-Type","application/x-www-form-urlencoded")
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            tokens = json.loads(r.read())
        # Busca nome + email do usuário Google para exibir na UI
        try:
            uinfo_req = urllib.request.Request("https://www.googleapis.com/oauth2/v2/userinfo")
            uinfo_req.add_header("Authorization", f"Bearer {tokens['access_token']}")
            with urllib.request.urlopen(uinfo_req, timeout=10) as ur:
                uinfo = json.loads(ur.read())
                tokens["_google_name"]  = uinfo.get("name","")
                tokens["_google_email"] = uinfo.get("email","")
                tokens["_google_pic"]   = uinfo.get("picture","")
        except Exception:
            pass
        return tokens
    except Exception:
        return {}

def _criar_evento_google_calendar(task: dict, tokens: dict) -> bool:
    import urllib.request
    dp = task_data_programada(task)
    if not dp: dp = datetime.now() + timedelta(days=1)
    evento = {
        "summary": f"[Sunne] {task.get('titulo','')}",
        "description": task.get("descricao","") or task.get("titulo",""),
        "start": {"dateTime": dp.strftime("%Y-%m-%dT09:00:00"), "timeZone": "America/Fortaleza"},
        "end":   {"dateTime": dp.strftime("%Y-%m-%dT10:00:00"), "timeZone": "America/Fortaleza"},
    }
    req = urllib.request.Request(
        "https://www.googleapis.com/calendar/v3/calendars/primary/events",
        data=json.dumps(evento).encode(), method="POST")
    req.add_header("Authorization", f"Bearer {tokens['access_token']}")
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=10): return True
    except Exception: return False

def _enviar_gmail(tokens: dict, para: str, assunto: str, corpo: str, tid: str="") -> bool:
    import urllib.request, base64 as b64
    msg_raw = (f"To: {para}\nSubject: {assunto}\nContent-Type: text/plain; charset=utf-8\n\n{corpo}")
    encoded = b64.urlsafe_b64encode(msg_raw.encode()).decode()
    payload = json.dumps({"raw": encoded}).encode()
    req = urllib.request.Request(
        "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
        data=payload, method="POST")
    req.add_header("Authorization", f"Bearer {tokens['access_token']}")
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=10): return True
    except Exception: return False

def _buscar_eventos_google_semana(tokens: dict) -> list:
    """Retorna eventos dos próximos 7 dias do Google Calendar."""
    import urllib.request, urllib.parse
    agora  = datetime.now()
    fim    = agora + timedelta(days=7)
    params = urllib.parse.urlencode({
        "timeMin": agora.strftime("%Y-%m-%dT00:00:00Z"),
        "timeMax": fim.strftime("%Y-%m-%dT23:59:59Z"),
        "singleEvents": "true", "orderBy": "startTime", "maxResults": "30",
    })
    req = urllib.request.Request(
        f"https://www.googleapis.com/calendar/v3/calendars/primary/events?{params}")
    req.add_header("Authorization", f"Bearer {tokens['access_token']}")
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            return data.get("items", [])
    except Exception:
        return []


# ═══════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════

def render_sidebar():
    with st.sidebar:
        u    = st.session_state["user"]
        goog = load_google_tokens()

        # ── Brand ─────────────────────────────────────────────────────────────
        st.markdown(
            '<div class="sb-brand">'
            '<div class="sb-brand-name">Sunne</div>'
            '<div class="sb-brand-sub">Hub Operacional</div>'
            '</div>',
            unsafe_allow_html=True)

        # ── User card ─────────────────────────────────────────────────────────
        role_label = {"admin":"Administrador","user":"Analista"}.get(u.get("role","user"),"Analista")
        st.markdown(
            f'<div class="sb-user">'
            f'<div class="sb-user-name">{u["name"]}</div>'
            f'<div class="sb-user-role">{role_label}</div>'
            f'</div>', unsafe_allow_html=True)

        # Google badge
        if goog.get("access_token"):
            nome_g = goog.get("_google_name","Google")
            st.markdown(
                f'<div class="sb-google-badge">✅ {nome_g}</div>',
                unsafe_allow_html=True)

        st.session_state.setdefault("page", "dash")

        role = u.get("role", "user")

        # ── Alertas rápidos na sidebar ────────────────────────────────────────
        tasks_sb  = load_tasks()
        minhas_sb = [t for t in tasks_sb if t.get("analista","").lower() == u["name"].lower()]
        atras_sb  = [t for t in minhas_sb if task_esta_atrasada(t)]
        hoje_sb   = [t for t in minhas_sb if task_vence_hoje(t)]
        if atras_sb:
            st.markdown(
                f'<div style="background:#FFF0F3;border:1px solid #FFC8D0;border-radius:8px;'
                f'padding:6px 10px;margin-bottom:6px;font-size:12px;color:#8B1530">'
                f'⏰ <b>{len(atras_sb)}</b> tarefa(s) atrasada(s)</div>',
                unsafe_allow_html=True)
        if hoje_sb:
            st.markdown(
                f'<div style="background:#FFFBEC;border:1px solid #FFE580;border-radius:8px;'
                f'padding:6px 10px;margin-bottom:6px;font-size:12px;color:#7A5010">'
                f'📅 <b>{len(hoje_sb)}</b> vence(m) hoje</div>',
                unsafe_allow_html=True)

        # ── Nav ADMIN — seção Gestão primeiro ─────────────────────────────────
        if role == "admin":
            st.markdown('<div class="sec-label">👔 Gestão</div>', unsafe_allow_html=True)
            if st.button("Painel do Gestor", key="nav_gestor_dash", use_container_width=True):
                st.session_state["page"] = "gestor_dash"
                st.rerun()

        # ── Nav ANALISTA / COMUM ──────────────────────────────────────────────
        st.markdown('<div class="sec-label">📋 Operação</div>', unsafe_allow_html=True)
        for key, label in [
            ("dash",       "Dashboard / Agenda"),
            ("atividades", "Atividades"),
            ("geradores",  "Geradores"),
            ("usinas",     "Usinas"),
        ]:
            if st.button(label, key=f"nav_{key}", use_container_width=True):
                st.session_state["page"] = key
                st.rerun()

        st.markdown('<div class="sec-label">🔍 Análise Mensal</div>', unsafe_allow_html=True)
        for key, label in [
            ("medicao",   "Rel. de Medição"),
            ("auditoria", "Auditoria UFV / UCs"),
        ]:
            if st.button(label, key=f"nav_{key}", use_container_width=True):
                st.session_state["page"] = key
                st.rerun()

        st.markdown('<div class="sec-label">⚡ Energia</div>', unsafe_allow_html=True)
        for key, label in [
            ("geracao",    "Geração"),
            ("backoffice", "Backoffice"),
        ]:
            if st.button(label, key=f"nav_{key}", use_container_width=True):
                st.session_state["page"] = key
                st.rerun()

        st.markdown('<div class="sec-label">💰 Financeiro</div>', unsafe_allow_html=True)
        for key, label in [
            ("rateio",      "Rateio"),
            ("faturamento", "Faturamento"),
            ("bi_analise",  "Análise BI"),
        ]:
            if st.button(label, key=f"nav_{key}", use_container_width=True):
                st.session_state["page"] = key
                st.rerun()

        st.markdown('<div class="sec-label">🤖 Automação</div>', unsafe_allow_html=True)
        if st.button("Captura Automática", key="nav_automacao", use_container_width=True):
            st.session_state["page"] = "automacao"
            st.rerun()

        st.markdown('<div class="sec-label">⚙️ Config</div>', unsafe_allow_html=True)
        if st.button("Integrações Google", key="nav_integracoes", use_container_width=True):
            st.session_state["page"] = "integracoes"
            st.rerun()

        # ── Divider + Sair ────────────────────────────────────────────────────
        st.markdown('<div class="sb-divider"></div>', unsafe_allow_html=True)
        if st.button("Sair", key="nav_sair", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


# ═══════════════════════════════════════════════════════════════════
# PAGES
# ═══════════════════════════════════════════════════════════════════

def page_dashboard():
    user = st.session_state["user"]; an = user["name"]
    agora = datetime.now()
    dia_semana = ["Segunda","Terça","Quarta","Quinta","Sexta","Sábado","Domingo"][agora.weekday()]
    st.markdown(
        f'<div style="margin-bottom:0.2rem"><h1 style="font-size:1.6rem;font-weight:500;margin:0">Olá, {an.split()[0]} 👋</h1><p style="color:var(--ink-l);font-size:13px;margin:2px 0 0">{dia_semana}, {agora.strftime("%d/%m/%Y")}</p></div>',
        unsafe_allow_html=True)

    # ── Troca OAuth code se presente na URL (retorno do Google) ──────────────
    params = st.query_params
    if "code" in params and not load_google_tokens().get("access_token"):
        with st.spinner("🔗 Conectando com o Google…"):
            tokens = _google_exchange_code(params["code"])
        if tokens.get("access_token"):
            save_google_tokens(tokens)
            st.query_params.clear()
            nome_g = tokens.get("_google_name","")
            st.toast(f"✅ Google conectado! Olá, {nome_g}! 👋", icon="🎉")
            st.rerun()
        else:
            st.query_params.clear()
            st.error("Não foi possível conectar com o Google. Tente novamente em Integrações.")

    gers    = [g for g in load_geradores() if g.get("analista","").lower() == an.lower()]
    usis    = [u for u in load_usinas()    if u.get("analista","").lower() == an.lower()]
    tasks   = load_tasks()
    geracao = load_geracao()
    bo      = load_backoffice()

    # ── Notificação de atividades do dia / atrasadas ──────────────────────────
    minhas_tasks = [t for t in tasks if t.get("analista","").lower() == an.lower()]
    vence_hoje   = [t for t in minhas_tasks if task_vence_hoje(t)]
    atrasadas    = [t for t in minhas_tasks if task_esta_atrasada(t)]
    if vence_hoje or atrasadas:
        linhas = []
        if vence_hoje:
            linhas.append(f"🔔 <b>{len(vence_hoje)}</b> atividade(s) programada(s) para <b>hoje</b>: "
                          + ", ".join(f"<i>{t['titulo']}</i>" for t in vence_hoje[:3]))
        if atrasadas:
            linhas.append(f"⏰ <b>{len(atrasadas)}</b> atividade(s) <b>atrasada(s)</b>: "
                          + ", ".join(f"<i>{t['titulo']}</i>" for t in atrasadas[:3]))
        st.markdown(f'<div class="notif-banner">{"<br>".join(linhas)}</div>',
                    unsafe_allow_html=True)

    # ── v12: AGENDA DO DIA ────────────────────────────────────────────────────
    agora_d = datetime.now()
    st.markdown("<div class='sdiv'></div>", unsafe_allow_html=True)
    col_ag, col_kpis = st.columns([3, 2], gap="large")

    with col_ag:
        st.markdown('<div class="sec-head"><span class="sec-head-title">📅 Agenda de Hoje</span>'
                    '<div class="sec-head-line"></div></div>', unsafe_allow_html=True)
        if atrasadas:
            st.markdown(
                f'<div class="alert alert-r" style="margin-bottom:8px">'
                f'⏰ <b>{len(atrasadas)} tarefa(s) atrasada(s)</b> — atenção imediata</div>',
                unsafe_allow_html=True)
            for t in sorted(atrasadas, key=lambda x: sla_days(x), reverse=True)[:5]:
                dias_at = sla_days(t)
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:8px;padding:6px 0;'
                    f'border-bottom:0.5px solid var(--border)">'
                    f'<div style="width:8px;height:8px;border-radius:50%;background:#C41230;flex-shrink:0"></div>'
                    f'<div style="flex:1;min-width:0">'
                    f'<div style="font-size:13px;font-weight:500;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{t["titulo"]}</div>'
                    f'<div style="font-size:11px;color:var(--ink-l)">{t.get("tipo","—")} · {t.get("usina","—")}</div>'
                    f'</div>'
                    f'<span style="background:#FFF0F3;color:#8B1530;border-radius:4px;'
                    f'padding:2px 7px;font-size:11px;font-weight:600;white-space:nowrap">{dias_at}d atraso</span>'
                    f'</div>', unsafe_allow_html=True)

        if vence_hoje:
            st.markdown('<div style="font-size:12px;font-weight:600;color:var(--ink-l);'
                        'text-transform:uppercase;letter-spacing:.06em;margin-top:10px;margin-bottom:4px">'
                        'Para hoje</div>', unsafe_allow_html=True)
            for t in vence_hoje[:5]:
                dp = t.get("data_programada","—")
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:8px;padding:6px 0;'
                    f'border-bottom:0.5px solid var(--border)">'
                    f'<div style="width:8px;height:8px;border-radius:50%;background:#F36E21;flex-shrink:0"></div>'
                    f'<div style="flex:1;min-width:0">'
                    f'<div style="font-size:13px;font-weight:500;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{t["titulo"]}</div>'
                    f'<div style="font-size:11px;color:var(--ink-l)">{t.get("tipo","—")} · {t.get("usina","—")}</div>'
                    f'</div>'
                    f'<span style="background:#FFF5EC;color:#A84010;border-radius:4px;padding:2px 7px;font-size:11px">Hoje</span>'
                    f'</div>', unsafe_allow_html=True)

        proximas = sorted(
            [t for t in minhas_tasks
             if t["status"] not in ("Concluido","Cancelado")
             and not task_esta_atrasada(t) and not task_vence_hoje(t)],
            key=lambda t: (datetime.strptime(t["data_programada"],"%d/%m/%Y")
                           if t.get("data_programada") and "/" in str(t.get("data_programada",""))
                           else datetime(2099,1,1)))[:5]

        if proximas:
            st.markdown('<div style="font-size:12px;font-weight:600;color:var(--ink-l);'
                        'text-transform:uppercase;letter-spacing:.06em;margin-top:10px;margin-bottom:4px">'
                        'Próximas</div>', unsafe_allow_html=True)
            for t in proximas:
                dp = t.get("data_programada","—")
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:8px;padding:6px 0;'
                    f'border-bottom:0.5px solid var(--border)">'
                    f'<div style="width:8px;height:8px;border-radius:50%;background:#A0C4E8;flex-shrink:0"></div>'
                    f'<div style="flex:1;min-width:0">'
                    f'<div style="font-size:13px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{t["titulo"]}</div>'
                    f'<div style="font-size:11px;color:var(--ink-l)">{t.get("tipo","—")}</div>'
                    f'</div>'
                    f'<span style="font-size:11px;color:var(--ink-l);white-space:nowrap">{dp}</span>'
                    f'</div>', unsafe_allow_html=True)

        if not atrasadas and not vence_hoje:
            st.markdown('<div class="alert alert-g">✅ Nenhuma tarefa urgente para hoje!</div>',
                        unsafe_allow_html=True)

    with col_kpis:
        st.markdown('<div class="sec-head"><span class="sec-head-title">Seus KPIs</span>'
                    '<div class="sec-head-line"></div></div>', unsafe_allow_html=True)
        kp_ab   = sum(1 for t in minhas_tasks if t["status"] == "Em aberto")
        kp_and  = sum(1 for t in minhas_tasks if t["status"] == "Em andamento")
        kp_trav = sum(1 for t in minhas_tasks if t["status"] == "Travado")
        kp_conc = sum(1 for t in minhas_tasks if t["status"] == "Concluido")
        kp_tot  = len(minhas_tasks)
        kp_pct  = round(kp_conc / kp_tot * 100) if kp_tot else 0
        for lbl, val, cor in [
            ("Atrasadas",    len(atrasadas), "#C41230"),
            ("Em aberto",    kp_ab,          "#A84010"),
            ("Em andamento", kp_and,         "#185FA5"),
            ("Travadas",     kp_trav,        "#6B1A8A"),
            ("Concluídas",   kp_conc,        "#0B7A5F"),
        ]:
            st.markdown(
                f'<div class="kpi-box" style="margin-bottom:6px;text-align:left;display:flex;'
                f'align-items:center;gap:10px;padding:.6rem .8rem">'
                f'<div class="kpi-value" style="font-size:1.5rem;color:{cor};min-width:32px">{val}</div>'
                f'<div class="kpi-label" style="text-align:left;font-size:12px">{lbl}</div>'
                f'</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div style="margin-top:8px">'
            f'<div style="font-size:11px;color:var(--ink-l);margin-bottom:3px">Progresso do mês · {kp_pct}%</div>'
            f'<div style="height:8px;background:var(--border);border-radius:4px;overflow:hidden">'
            f'<div style="width:{kp_pct}%;height:100%;background:#F36E21;border-radius:4px"></div>'
            f'</div></div>', unsafe_allow_html=True)

        mes_atual_str = agora_d.strftime("%m/%Y")
        tasks_mes = [t for t in minhas_tasks if t.get("agendamento","") == mes_atual_str]
        if not tasks_mes:
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            st.markdown(
                f'<div class="alert alert-y" style="font-size:12px">'
                f'⚠️ Cronograma de {mes_atual_str} não gerado.<br>'
                f'Peça ao gestor para gerar via <b>Painel do Gestor</b>.</div>',
                unsafe_allow_html=True)

    st.markdown("<div class='sdiv'></div>", unsafe_allow_html=True)

    ab    = [t for t in tasks if t["status"] == "Em aberto"]
    and_  = [t for t in tasks if t["status"] == "Em andamento"]
    trav  = [t for t in tasks if t["status"] == "Travado"]
    ativos= [t for t in tasks if t["status"] not in ("Concluido","Cancelado")]
    sla   = round(sum(sla_days(t) for t in ativos)/len(ativos), 1) if ativos else 0

    cols = st.columns(6)
    for col, lbl, val in zip(cols,
        ["Geradores","Usinas","Em Aberto","Em Andamento","Travadas","SLA Médio"],
        [len(gers), len(usis), len(ab), len(and_), len(trav), f"{sla}d"]):
        col.markdown(
            f'<div class="kpi-box"><div class="kpi-value">{val}</div>'
            f'<div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)

    st.write("")

    # ── Layout: alertas + calendário ─────────────────────────────────────────
    col_main, col_cal = st.columns([3, 2])

    with col_main:
        st.markdown("#### Alertas Operacionais")
        mes = datetime.now().strftime("%m/%Y")
        uc_ger_mes = {g["uc"] for g in geracao if g.get("competencia","") == mes}
        usi_ids    = {u["uc"] for u in usis}
        alerts = []
        for uc in usi_ids - uc_ger_mes:
            ui = next((x for x in usis if x["uc"] == uc), {})
            alerts.append(("r", f"⚡ Usina <b>{ui.get('ufv',uc)}</b> sem geração registrada em {mes}."))
        for uc in uc_ger_mes:
            r = next((x for x in bo if str(x.get("uc","")) == str(uc)), None)
            if r:
                c = clean_val(r.get("consumo_total",0)); s = clean_val(r.get("saldo_credito",0))
                if c > 0 and s/c > 6:
                    ui = next((x for x in usis if str(x["uc"]) == str(uc)), {})
                    alerts.append(("y", f"💰 Usina <b>{ui.get('ufv',uc)}</b>: saldo acumulado > 6 meses."))
        if not alerts:
            st.markdown('<div class="alert alert-g">✅ Nenhum alerta operacional.</div>', unsafe_allow_html=True)
        else:
            for k, m in alerts:
                st.markdown(f'<div class="alert alert-{k}">{m}</div>', unsafe_allow_html=True)

        st.write("")
        st.markdown("#### Tarefas Abertas")
        pend = ab + and_
        if pend:
            df_p = pd.DataFrame(pend)
            cols_p = [c for c in ["titulo","usina","gerador","analista","status","data_programada","criado_em"] if c in df_p.columns]
            df_p = df_p[cols_p]
            df_p.columns = [{"titulo":"Título","usina":"Usina","gerador":"Gerador",
                              "analista":"Analista","status":"Status",
                              "data_programada":"Data Prog.","criado_em":"Criado em"}.get(c,c)
                            for c in cols_p]
            st.dataframe(df_p, use_container_width=True, hide_index=True)
        else:
            st.success("Nenhuma tarefa pendente.")

    with col_cal:
        st.markdown("#### 📅 Agenda da Semana")
        # Atividades programadas do Hub
        hoje = datetime.now().date()
        dias = [hoje + timedelta(days=i) for i in range(7)]
        goog = load_google_tokens()
        eventos_google = _buscar_eventos_google_semana(goog) if goog.get("access_token") else []

        st.markdown('<div class="cal-box">', unsafe_allow_html=True)
        for dia in dias:
            dia_str   = dia.strftime("%d/%m")
            dia_label = dia.strftime("%a %d/%m").upper()
            # Tarefas do Hub neste dia
            tasks_dia = [t for t in minhas_tasks
                         if task_data_programada(t) and
                         task_data_programada(t).date() == dia]
            # Eventos Google neste dia
            ev_google = [e for e in eventos_google
                         if (e.get("start",{}).get("dateTime","") or
                             e.get("start",{}).get("date","")).startswith(dia.strftime("%Y-%m-%d"))]

            if not tasks_dia and not ev_google: continue
            st.markdown(f'<div class="cal-day-header">{dia_label}</div>', unsafe_allow_html=True)
            for t in tasks_dia:
                cls = "cal-event-atrasado" if task_esta_atrasada(t) else \
                      ("cal-event-hoje"     if dia == hoje else "cal-event")
                st.markdown(f'<div class="{cls}">📌 {t["titulo"]}</div>', unsafe_allow_html=True)
            for ev in ev_google:
                titulo_ev = ev.get("summary","(sem título)")
                hora_ev   = (ev.get("start",{}).get("dateTime","") or "")
                hora_fmt  = hora_ev[11:16] if len(hora_ev) > 15 else ""
                st.markdown(f'<div class="cal-event">🗓 {hora_fmt} {titulo_ev}</div>',
                            unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        if not goog.get("access_token"):
            auth_url = _google_auth_url()
            st.markdown(
                f'<a href="{auth_url}" target="_self" style="display:inline-block;margin-top:.6rem;'
                f'background:#F36E21;color:white;padding:.45rem 1.1rem;border-radius:8px;'
                f'font-size:13px;font-weight:500;text-decoration:none">🔗 Conectar Google Calendar</a>',
                unsafe_allow_html=True)


def page_geradores():
    user = st.session_state["user"]; an = user["name"]
    st.title("Geradores")
    tc, ti, tm_add = st.tabs(["Carteira","Importar","Adicionar Manual"])

    with tc:
        minha = [g for g in load_geradores() if g.get("analista","").lower() == an.lower()]
        if not minha:
            st.info("Nenhum gerador cadastrado. Use as abas **Importar** ou **Adicionar Manual**.")
        else:
            conc = {g.get("concessionaria","") for g in minha if g.get("concessionaria","")}
            tot  = sum(int(g.get("usinas",0) or 0) for g in minha)
            c1, c2, c3 = st.columns(3)
            for col, lbl, val in zip([c1,c2,c3],
                ["Geradores","Concessionárias","Usinas Totais"],[len(minha),len(conc),tot]):
                col.markdown(
                    f'<div class="kpi-box"><div class="kpi-value">{val}</div>'
                    f'<div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)
            st.write("")
            filtro = st.selectbox("Concessionária", ["Todas"] + sorted(conc))
            lista  = minha if filtro == "Todas" else [g for g in minha if g.get("concessionaria") == filtro]
            df_s   = pd.DataFrame(lista)
            cols_s = [c for c in ["gerador","contato","concessionaria","usinas","porte","origem",
                                   "enquadramento","pct_desconto_gerador","pct_taxa_admin"] if c in df_s.columns]
            df_s   = df_s[cols_s]
            df_s.columns = [{"pct_desconto_gerador":"% Desc.","pct_taxa_admin":"% Admin",
                              "enquadramento":"GD"}.get(c, c.capitalize()) for c in cols_s]
            st.dataframe(df_s, use_container_width=True, hide_index=True)

    with ti:
        st.markdown("#### Importar Planilha de Geradores")
        st.markdown(
            "Colunas obrigatórias: **Gerador · Analista**  \n"
            "Opcionais: Contato · Concessionária · Usinas · Porte · Origem · "
            "Enquadramento (GD1/GD2) · pct_desconto_gerador · pct_taxa_admin")
        st.write("")   # garante render antes do uploader
        arquivo_ger = st.file_uploader(
            "Selecione o arquivo (.xlsx ou .csv)",
            type=["xlsx","xls","csv"],
            key="up_ger_v10",
            help="A planilha deve ter pelo menos a coluna 'Gerador'")
        if arquivo_ger is not None:
            st.success(f"📄 Arquivo carregado: **{arquivo_ger.name}**")
            if st.button("✅ Importar Geradores", key="btn_sg_v10", use_container_width=True):
                try:
                    if arquivo_ger.name.endswith(".csv"):
                        df = pd.read_csv(arquivo_ger, dtype=str)
                    else:
                        df = pd.read_excel(arquivo_ger, dtype=str)
                    df.columns = df.columns.str.strip().str.lower()
                    df.rename(columns={"concessionária":"concessionaria"}, inplace=True)
                    df = df.fillna("")
                    ex = load_geradores()
                    nex = {g["gerador"].lower() for g in ex}
                    n = 0
                    for _, row in df.iterrows():
                        nm = str(row.get("gerador","")).strip()
                        if nm and nm.lower() not in nex:
                            rec = {k: str(row.get(k,"")) for k in
                                   ["gerador","contato","analista","concessionaria","usinas","porte","origem"]}
                            rec["enquadramento"]        = str(row.get("enquadramento","GD1"))
                            rec["pct_desconto_gerador"] = str(row.get("pct_desconto_gerador","0.20"))
                            rec["pct_taxa_admin"]       = str(row.get("pct_taxa_admin","0.07"))
                            if not rec["analista"].strip():
                                rec["analista"] = an
                            ex.append(rec)
                            nex.add(nm.lower())
                            n += 1
                    save_geradores(ex)
                    st.success(f"✅ {n} gerador(es) importado(s) com sucesso!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao importar: {e}")
                    st.code(traceback.format_exc())

    with tm_add:
        st.markdown("#### Adicionar Gerador Manualmente")
        with st.form("form_add_ger", clear_on_submit=True):
            ag1, ag2 = st.columns(2)
            nm_ger  = ag1.text_input("Nome do Gerador *")
            contato = ag2.text_input("E-mail de contato")
            ag3, ag4, ag5 = st.columns(3)
            concess = ag3.text_input("Concessionária")
            porte   = ag4.text_input("Porte")
            origem  = ag5.text_input("Origem")
            ag6, ag7, ag8 = st.columns(3)
            enq_g  = ag6.selectbox("Enquadramento", list(TARIFAS_BASE.keys()))
            desc_g = ag7.number_input("% Desconto Gerador", 0.0, 1.0, 0.20, 0.01, format="%.2f")
            taxa_g = ag8.number_input("% Taxa Admin Sunne",  0.0, 1.0, 0.07, 0.01, format="%.2f")
            salvar_g = st.form_submit_button("💾 Salvar Gerador", use_container_width=True)
        if salvar_g:
            if not nm_ger.strip():
                st.warning("Nome do gerador obrigatório.")
            else:
                ex = load_geradores()
                if nm_ger.strip().lower() in {g["gerador"].lower() for g in ex}:
                    st.warning("Gerador já cadastrado.")
                else:
                    ex.append({"gerador":nm_ger.strip(),"contato":contato,"analista":an,
                                "concessionaria":concess,"usinas":"0","porte":porte,"origem":origem,
                                "enquadramento":enq_g,
                                "pct_desconto_gerador":float(desc_g),
                                "pct_taxa_admin":float(taxa_g)})
                    save_geradores(ex)
                    st.success(f"✅ Gerador '{nm_ger}' cadastrado!")
                    st.rerun()


def page_usinas():
    user = st.session_state["user"]; an = user["name"]
    st.title("Usinas")

    ca, cb, _ = st.columns([1.3, 1.7, 6])
    if ca.button("Adicionar"):
        st.session_state["show_add_usi"] = not st.session_state.get("show_add_usi", False)
    if cb.button("Importar Planilha"):
        st.session_state["show_imp_usi"] = not st.session_state.get("show_imp_usi", False)

    if st.session_state.get("show_add_usi"):
        with st.form("fau", clear_on_submit=True):
            st.markdown("**Nova Usina**")
            a1, a2 = st.columns(2)
            uc    = a1.text_input("UC *"); ger = a2.text_input("Gerador")
            b1, b2, b3 = st.columns(3)
            ufv   = b1.text_input("UFV"); ativa = b2.selectbox("Ativa", ["Sim","Não"])
            gest  = b3.number_input("Geração estimada kWh", min_value=0.0, step=1.0)
            # Campos de contrato
            c1, c2, c3 = st.columns(3)
            enquad = c1.selectbox("Enquadramento", list(TARIFAS_BASE.keys()),
                                  help="GD1 = tarifa 0.8182 · GD2 = 0.64788")
            pct_desc_u = c2.number_input("% Desconto Gerador", 0.0, 1.0, 0.20, 0.01,
                                         format="%.2f", help="Ex: 0.20 = 20%")
            pct_taxa_u = c3.number_input("% Taxa Admin Sunne",  0.0, 1.0, 0.07, 0.01,
                                         format="%.2f", help="Ex: 0.07 = 7%")
            s, c  = st.columns(2)
            sal   = s.form_submit_button("Salvar"); canc = c.form_submit_button("Cancelar")
        if sal:
            if not uc.strip(): st.warning("UC obrigatória.")
            else:
                us = load_usinas()
                us.append({"uc":str(uc).strip(),"gerador":ger.strip(),"ufv":ufv.strip(),
                            "analista":an,"ativa":ativa,"geracao_estimada":gest,
                            "enquadramento":enquad,
                            "pct_desconto_gerador":float(pct_desc_u),
                            "pct_taxa_admin":float(pct_taxa_u),
                            "criado_em":datetime.now().strftime("%d/%m/%Y")})
                save_usinas(us); st.session_state["show_add_usi"] = False; st.rerun()
        if canc: st.session_state["show_add_usi"] = False; st.rerun()

    if st.session_state.get("show_imp_usi"):
        st.caption("Colunas: UC · Gerador · UFV · Analista · Ativa · Geração")
        f = st.file_uploader("Arquivo", type=["xlsx","xls","csv"], key="up_usi")
        i1, i2 = st.columns([1, 5])
        if i1.button("Importar", key="btn_iusi") and f:
            try:
                df = pd.read_excel(f, dtype=str) if not f.name.endswith(".csv") else pd.read_csv(f, dtype=str)
                df.columns = df.columns.str.strip().str.lower(); df = df.fillna("")
                for col in list(df.columns):
                    cl = col.lower()
                    if "gerador" in cl and col != "gerador": df.rename(columns={col:"gerador"}, inplace=True)
                    if cl in ("geração","geracao","geração (kwh)","geracao (kwh)"):
                        df.rename(columns={col:"geracao_estimada"}, inplace=True)
                us = load_usinas(); uc_ex = {u["uc"] for u in us}; n = 0
                for _, row in df.iterrows():
                    uv = str(row.get("uc","")).strip()
                    if uv and uv not in uc_ex:
                        us.append({"uc":uv,"gerador":str(row.get("gerador","")),
                            "ufv":str(row.get("ufv","")),"analista":an,
                            "ativa":str(row.get("ativa","Sim")),
                            "geracao_estimada":clean_val(row.get("geracao_estimada",0)),
                            "enquadramento":str(row.get("enquadramento","GD1")),
                            "pct_desconto_gerador":clean_val(row.get("pct_desconto_gerador",0.20)),
                            "pct_taxa_admin":clean_val(row.get("pct_taxa_admin",0.07)),
                            "criado_em":datetime.now().strftime("%d/%m/%Y")})
                        uc_ex.add(uv); n += 1
                save_usinas(us); st.session_state["show_imp_usi"] = False
                st.success(f"{n} usina(s) importada(s)."); st.rerun()
            except Exception as e: st.error(str(e))
        if i2.button("Fechar", key="fusi"):
            st.session_state["show_imp_usi"] = False; st.rerun()

    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    usinas = load_usinas()
    minhas = [u for u in usinas if u.get("analista","").lower() == an.lower()]
    if not minhas: st.info("Nenhuma usina cadastrada."); return

    ger_opts = ["Todos"] + sorted({u.get("gerador","—") for u in minhas})
    filtro   = st.selectbox("Filtrar por Gerador", ger_opts)
    lista    = minhas if filtro == "Todos" else [u for u in minhas if u.get("gerador") == filtro]

    h = st.columns([1.2, 1.6, 2.5, 0.6, 0.6, 0.6, 0.8])
    for col, txt in zip(h, ["UC","Gerador","UFV","GD","Ativa","Desc%","Atividade"]):
        col.markdown(f"**{txt}**")
    st.markdown("<hr style='margin:4px 0 8px;border:none;border-top:1px solid rgba(51,0,26,.08)'>",
                unsafe_allow_html=True)

    for idx, u in enumerate(lista):
        r = st.columns([1.2, 1.6, 2.5, 0.6, 0.6, 0.6, 0.8])
        r[0].write(str(u.get("uc","—")))
        r[1].write(u.get("gerador","—"))
        r[2].write(u.get("ufv","—"))
        r[3].write(u.get("enquadramento","—"))
        r[4].write("✅" if u.get("ativa","Sim") == "Sim" else "❌")
        pct = clean_val(u.get("pct_desconto_gerador",0))
        r[5].write(f"{pct*100:.0f}%" if pct <= 1 else f"{pct:.0f}%")
        if r[6].button("📝 Nova", key=f"ativ_{idx}_{u['uc']}"):
            dialog_nova_atividade(uc_pre=str(u.get("uc","")), ger_pre=u.get("gerador",""))


def page_atividades():
    st.title("Atividades")

    hc1, _ = st.columns([2, 8])
    if hc1.button("➕ Nova Atividade"):
        dialog_nova_atividade()

    fc1, fc2 = st.columns(2)
    f_an = fc1.text_input("Filtrar Analista", placeholder="todos")
    f_ge = fc2.text_input("Filtrar Gerador",  placeholder="todos")

    def match(t):
        if f_an and f_an.lower() not in t.get("analista","").lower(): return False
        if f_ge and f_ge.lower() not in t.get("gerador","").lower():  return False
        return True

    tasks = [t for t in load_tasks() if match(t)]
    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    kb_cols = st.columns(5)

    for col_ui, status in zip(kb_cols, STATUS_LIST):
        css   = KB_CSS[status]
        grupo = [t for t in tasks if t["status"] == status]
        with col_ui:
            count_badge = (
                f'<span style="background:rgba(0,0,0,.08);border-radius:12px;'
                f'padding:1px 8px">{len(grupo)}</span>')
            st.markdown(
                f'<div class="kb-col-head {css}"><span>{status}</span>{count_badge}</div>',
                unsafe_allow_html=True)
            st.markdown('<div class="kb-wrap">', unsafe_allow_html=True)
            if not grupo:
                st.markdown('<div class="kb-empty">—</div>', unsafe_allow_html=True)

            for t in grupo:
                tid  = t["id"]
                dias = sla_days(t)
                sc   = sla_cls(dias)
                mot_h = (f'<div class="kb-motivo">🔒 {t["motivo_bloqueio"]}</div>'
                         if t.get("motivo_bloqueio") else "")
                n_obs = len([l for l in t.get("observacoes","").split("\n")
                              if l.startswith("[")]) if t.get("observacoes") else 0
                obs_h = f'<span class="kb-obs-badge">📝 {n_obs}</span>' if n_obs > 0 else ""

                # Tags de prazo
                prazo_tag = ""
                if task_esta_atrasada(t):
                    prazo_tag = '<span class="kb-tag-atrasada">⏰ ATRASADA</span>'
                elif task_vence_hoje(t):
                    prazo_tag = '<span class="kb-tag-hoje">🔔 HOJE</span>'
                elif t.get("data_programada"):
                    prazo_tag = f'<span class="kb-tag-prog">📅 {t["data_programada"]}</span>'

                st.markdown(f"""
                <div class="kb-card">
                  <div class="kb-card-title">{t['titulo']}</div>
                  <div class="kb-card-meta">
                    🏭 {t.get('usina','—')}<br>
                    ⚡ {t.get('gerador','—')}<br>
                    👤 {t.get('analista','—')}<br>
                    📅 {t.get('criado_em','')}
                  </div>
                  <div class="{sc}">⏱ SLA: {dias}d</div>
                  {prazo_tag}{obs_h}{mot_h}
                </div>""", unsafe_allow_html=True)

                if st.button("Abrir", key=f"open_{tid}", use_container_width=True):
                    dialog_task_detail(tid)

                opcoes = [s for s in STATUS_LIST if s != status]
                dest   = st.selectbox("", opcoes, key=f"dest_{tid}", label_visibility="collapsed")
                if st.button("↗", key=f"mv_{tid}", help="Mover para status selecionado"):
                    if dest in MOTIVO_OBRIG:
                        st.session_state[f"pm_{tid}"] = dest
                    else:
                        update_task(tid, status=dest, motivo_bloqueio=""); st.rerun()

                if st.session_state.get(f"pm_{tid}"):
                    dv  = st.session_state[f"pm_{tid}"]
                    mot = st.text_area(f"Motivo para '{dv}' *", key=f"mot_{tid}")
                    mc1, mc2 = st.columns(2)
                    if mc1.button("OK", key=f"conf_{tid}"):
                        if mot.strip():
                            update_task(tid, status=dv, motivo_bloqueio=mot.strip())
                            del st.session_state[f"pm_{tid}"]; st.rerun()
                        else: st.warning("Motivo obrigatório.")
                    if mc2.button("✕", key=f"cno_{tid}"):
                        del st.session_state[f"pm_{tid}"]; st.rerun()

                st.markdown("<hr style='margin:4px 0;border:none;border-top:1px solid rgba(51,0,26,.06)'>",
                            unsafe_allow_html=True)
            if grupo:
                med = round(sum(sla_days(t) for t in grupo)/len(grupo), 1)
                st.markdown(f'<div class="kb-metric">⏱ Média {med}d</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)


def page_geracao():
    st.title("Geração das Usinas")
    tm, ti = st.tabs(["Lançamento Manual","Importar Excel"])

    def _upsert_geracao(ger_list: list, novo: dict) -> list:
        """Substitui registro com mesma UC+competência; adiciona se não existe."""
        uc_n   = normalize_uc(novo["uc"])
        comp_n = str(novo.get("competencia","")).strip()
        ger_list = [g for g in ger_list if not (
            normalize_uc(g.get("uc","")) == uc_n and
            str(g.get("competencia","")).strip() == comp_n
        )]
        ger_list.append(novo)
        return ger_list

    with tm:
        uc_inp = st.text_input("UC da Usina *", key="ger_uc_inp", placeholder="Digite a UC…")
        nome_auto = ""; gerador_auto = ""
        if uc_inp.strip():
            usinas = load_usinas()
            m = next((u for u in usinas if normalize_uc(u["uc"]) == normalize_uc(uc_inp)), None)
            if m:
                nome_auto = m.get("ufv",""); gerador_auto = m.get("gerador","")
                st.success(f"✅ **{nome_auto}** · Gerador: **{gerador_auto}**")
            else:
                st.caption(f"UC {uc_inp} não encontrada no cadastro — preencha o nome manualmente.")

        with st.form("form_geracao", clear_on_submit=True):
            g1, g2 = st.columns(2)
            nome_g = g1.text_input("Nome Usina", value=nome_auto)
            comp_g = g2.text_input("Competência (MM/AAAA)", value=datetime.now().strftime("%m/%Y"))
            g3, g4 = st.columns(2)
            inj_g  = g3.number_input("Energia Injetada (kWh)", min_value=0.0, step=0.1)
            sld_g  = g4.number_input("Saldo (kWh)",             min_value=0.0, step=0.1)
            ok = st.form_submit_button("Salvar Geração", use_container_width=True)

        if ok:
            if not uc_inp.strip(): st.warning("UC obrigatória.")
            else:
                usinas = load_usinas()
                m2 = next((u for u in usinas if normalize_uc(u["uc"]) == normalize_uc(uc_inp)), None)
                nf = m2.get("ufv","") if m2 else nome_g
                gd = m2.get("gerador","") if m2 else gerador_auto
                ger = load_geracao()
                novo = {"uc":str(uc_inp).strip(),"nome_usina":nf,"gerador":gd,
                        "competencia":comp_g,"energia_injetada":inj_g,"saldo":sld_g,
                        "registrado_em":datetime.now().strftime("%d/%m/%Y %H:%M")}
                ger = _upsert_geracao(ger, novo)
                save_geracao(ger)
                st.success(f"✅ Geração registrada — **{nf or uc_inp}** · {comp_g}")

    with ti:
        st.caption("Colunas: Nome da Usina · Número da UG · Competência · Energia Injetada · Saldo")
        st.info("💡 Registros com mesma UC + Competência serão **substituídos** (sem duplicatas).")
        f = st.file_uploader("Arquivo", type=["xlsx","xls","csv"], key="up_ger2")
        if f and st.button("Importar", key="btn_iger"):
            try:
                df = pd.read_excel(f, dtype=str) if not f.name.endswith(".csv") else pd.read_csv(f, dtype=str)
                df.columns = df.columns.str.strip().str.lower(); df = df.fillna("")
                df.rename(columns={"nome da usina":"nome_usina","número da ug":"uc",
                                   "numero da ug":"uc","energia injetada":"energia_injetada",
                                   "número da uc":"uc","numero da uc":"uc"}, inplace=True)
                usinas = load_usinas(); ger = load_geracao(); n = 0
                for _, row in df.iterrows():
                    uv = str(row.get("uc","")).strip()
                    if not uv: continue
                    m  = next((u for u in usinas if normalize_uc(u["uc"]) == normalize_uc(uv)), None)
                    nm = m.get("ufv","") if m else str(row.get("nome_usina",""))
                    gd = m.get("gerador","") if m else ""
                    comp_raw = str(row.get("competência", row.get("competencia",
                                  row.get("data","")))).strip()
                    # Normaliza competência para MM/AAAA
                    comp_fmt = comp_raw
                    try:
                        if "-" in comp_raw or len(comp_raw) > 7:
                            dt_c = pd.to_datetime(comp_raw, errors="coerce")
                            if pd.notna(dt_c): comp_fmt = dt_c.strftime("%m/%Y")
                    except: pass
                    novo = {"uc":str(uv),"nome_usina":nm,"gerador":gd,
                            "competencia":comp_fmt,
                            "energia_injetada":clean_val(row.get("energia_injetada",0)),
                            "saldo":clean_val(row.get("saldo",0)),
                            "registrado_em":datetime.now().strftime("%d/%m/%Y %H:%M")}
                    ger = _upsert_geracao(ger, novo); n += 1
                save_geracao(ger); st.success(f"✅ {n} registro(s) importado(s) sem duplicatas.")
            except Exception as e: st.error(str(e))

    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    st.markdown("#### Registros de Geração")
    ger = load_geracao()
    if not ger: st.info("Nenhum registro ainda."); return

    df_g = pd.DataFrame(ger)
    for cn in ["uc","nome_usina","gerador","competencia","energia_injetada","saldo"]:
        if cn not in df_g.columns: df_g[cn] = ""

    fa, fb, _ = st.columns([2, 2, 3])
    gers_list  = sorted({str(x) for x in df_g["gerador"].unique() if x})
    ger_filtro = fa.selectbox("Gerador", ["Todos"] + gers_list, key="gf_ger")
    usi_list   = []
    if ger_filtro != "Todos":
        usi_list = sorted({str(r) for r in df_g[df_g["gerador"]==ger_filtro]["nome_usina"].unique() if r})
    usi_filtro = fb.selectbox("Usina", ["Todas"] + usi_list, key="gf_usi",
                              disabled=(ger_filtro=="Todos"))

    df_f = df_g.copy()
    if ger_filtro != "Todos":  df_f = df_f[df_f["gerador"] == ger_filtro]
    if usi_filtro != "Todas" and ger_filtro != "Todos":
        df_f = df_f[df_f["nome_usina"] == usi_filtro]

    df_f = df_f.copy()
    df_f["_sort"] = df_f["competencia"].apply(comp_sort_key)
    df_f = df_f.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    cols_show = [c for c in ["uc","nome_usina","gerador","competencia","energia_injetada","saldo","registrado_em"] if c in df_f.columns]
    st.dataframe(df_f[cols_show], use_container_width=True, hide_index=True)


def page_backoffice():
    st.title("Backoffice · Captura de Consumo")
    geradores = load_geradores(); nomes = [g["gerador"] for g in geradores]
    if not nomes: st.warning("Cadastre geradores primeiro."); return

    bc1, bc2 = st.columns(2)
    sel_ger  = bc1.selectbox("Vincular ao Gerador", sorted(set(nomes)), key="bo_ger")
    comp_bo  = bc2.text_input("Competência (MM/AAAA)",
                               value=datetime.now().strftime("%m/%Y"), key="bo_comp",
                               help="Competência referente a este extrato — salva o histórico por mês")

    f = st.file_uploader("Extrato Detalhado", type=["xlsx","xls","csv"], key="up_bo")
    if f and st.button("Processar Extrato"):
        try:
            df = pd.read_excel(f, header=None) if not f.name.endswith(".csv") else pd.read_csv(f, header=None, sep=None, engine='python')
            for i, row in df.head(20).iterrows():
                if any("número da uc" in str(c).lower() for c in row):
                    df.columns = [str(c).strip() for c in row]
                    df = df.iloc[i+1:].reset_index(drop=True); break
            df.columns = [str(c).strip() for c in df.columns]; df = df.fillna("")
            uc_c  = next((c for c in df.columns if "Número da UC"   in c), None)
            co_c  = next((c for c in df.columns if "Consumo"        in c), None)
            sa_c  = next((c for c in df.columns if "Saldo"          in c), None)
            ti_c  = next((c for c in df.columns if "Tipo"           in c or "Instalação" in c), None)
            tt_c  = next((c for c in df.columns if "Titular"        in c), None)
            cred_c= next((c for c in df.columns if "Créditos"       in c and "Utiliz" in c), None)
            tarf_c= next((c for c in df.columns if "Tarifa Compensável" in c), None)
            if not uc_c: st.error("Coluna 'Número da UC' não encontrada."); return
            bo = load_backoffice(); nn = 0
            for _, row in df.iterrows():
                uv = str(row[uc_c]).strip()
                if not uv: continue
                # SEMPRE adiciona — histórico por competência
                rec = {
                    "uc":            uv,
                    "gerador":       sel_ger,
                    "competencia":   comp_bo.strip(),
                    "titular":       str(row[tt_c])   if tt_c   else "—",
                    "consumo_total": clean_val(row[co_c]) if co_c else 0,
                    "saldo_credito": clean_val(row[sa_c]) if sa_c else 0,
                    "creditos_utilizados": clean_val(row[cred_c]) if cred_c else 0,
                    "tarifa_compensavel":  clean_val(row[tarf_c]) if tarf_c else 0,
                    "tipo_instalacao":str(row[ti_c]) if ti_c else "—",
                    "importado_em":  datetime.now().strftime("%d/%m/%Y %H:%M"),
                }
                bo.append(rec); nn += 1
            save_backoffice(bo)
            st.success(f"✅ {nn} registros adicionados ao histórico — Gerador: **{sel_ger}** · Competência: **{comp_bo}**")
        except Exception as e: st.error(str(e))

    bo = load_backoffice()
    if bo:
        st.markdown("#### Histórico de Consumo")
        bf1, bf2, bf3 = st.columns(3)
        gf = bf1.selectbox("Filtrar Gerador",
                           ["Todos"] + sorted({b.get("gerador","—") for b in bo}), key="bo_f")
        comps_bo = sorted({b.get("competencia","") for b in bo if b.get("competencia","")},
                          key=comp_sort_key, reverse=True)
        cf = bf2.selectbox("Competência", ["Todas"] + comps_bo, key="bo_cf")
        uc_f = bf3.text_input("UC", placeholder="todos", key="bo_uc_f")

        bof = bo
        if gf != "Todos":   bof = [b for b in bof if b.get("gerador","") == gf]
        if cf != "Todas":   bof = [b for b in bof if b.get("competencia","") == cf]
        if uc_f.strip():    bof = [b for b in bof if uc_f.strip() in str(b.get("uc",""))]

        st.caption(f"{len(bof)} registros")
        df_bo = pd.DataFrame(bof)
        cols_bo = [c for c in ["competencia","uc","gerador","titular","consumo_total",
                                "saldo_credito","creditos_utilizados","tarifa_compensavel",
                                "tipo_instalacao","importado_em"] if c in df_bo.columns]
        st.dataframe(df_bo[cols_bo], use_container_width=True, hide_index=True)
    else:
        st.info("Nenhum dado ainda.")


def page_rateio():
    st.title("Rateio")
    ta, tb, tc, td, te = st.tabs(["Rebalancear","Atualizar Vigente","Consultar","Buscar UC","🔄 Rateios Sunne"])
    geradores = load_geradores()
    nomes_ger = sorted({g["gerador"] for g in geradores})

    with ta:
        st.markdown("#### Rebalanceamento de Rateio")
        if not nomes_ger: st.info("Cadastre geradores."); return
        sg = st.selectbox("Gerador", nomes_ger, key="rb_ger")
        ug = get_usinas_do_gerador(sg)
        if not ug: st.warning(f"Nenhuma usina vinculada a **{sg}**."); return
        nms = [u.get("ufv") or u["uc"] for u in ug]
        su  = st.selectbox("Usina", nms, key="rb_usi")
        su_obj = ug[nms.index(su)]; uc_u = str(su_obj["uc"])

        gm  = load_geracao(); mes = datetime.now().strftime("%m/%Y")
        rg  = next((g for g in gm if normalize_uc(str(g.get("uc",""))) == normalize_uc(uc_u)
                    and g.get("competencia","") == mes), None)
        if not rg:
            st.warning(f"Geração de **{su}** para {mes} não registrada. Registre em Geração primeiro."); return

        gkwh = clean_val(rg.get("energia_injetada",0))
        st.info(f"Geração {mes}: **{gkwh:,.2f} kWh**")
        aut = st.number_input("Meses de autonomia desejada", 1, 12, 3, key="rb_aut")

        ra1, ra2 = st.columns(2)
        f_rat = ra1.file_uploader("Rateio Vigente",    type=["xlsx","xls","csv"], key="up_rat")
        f_ext = ra2.file_uploader("Extrato Detalhado", type=["xlsx","xls","csv"], key="up_ext_rat")

        st.markdown("**UCs Saindo:**")
        ns = st.number_input("Qtd", 0, 30, 0, step=1, key="rb_ns")
        ucs_s = []
        for i in range(int(ns)):
            c1, c2 = st.columns(2)
            us2 = c1.text_input(f"UC #{i+1}", key=f"us_{i}")
            ms2 = c2.text_input(f"Motivo #{i+1}", key=f"ms_{i}")
            if us2: ucs_s.append({"uc":str(us2).strip(),"motivo":ms2})

        st.markdown("**Novas UCs:**")
        nn2 = st.number_input("Qtd", 0, 30, 0, step=1, key="rb_nn")
        ucs_n = []
        for i in range(int(nn2)):
            c1, c2, c3 = st.columns(3)
            un = c1.text_input(f"UC #{i+1}", key=f"un_{i}")
            an2= c2.text_input(f"Apelido #{i+1}", key=f"an2_{i}")
            cn = c3.number_input(f"Consumo kWh #{i+1}", 0.0, step=1.0, key=f"cn_{i}")
            if un: ucs_n.append({"uc":str(un).strip(),"apelido":an2,"consumo":cn})

        ancora = st.text_input("Unidade Âncora (opcional)", key="rb_anc")

        if st.button("Calcular", key="btn_rb"):
            if not f_rat or not f_ext: st.error("Faça upload dos dois arquivos."); return
            try:
                df_rat = pd.read_excel(f_rat, dtype=str) if not f_rat.name.endswith(".csv") else pd.read_csv(f_rat, dtype=str)
                df_rat.columns = df_rat.columns.str.strip(); df_rat = df_rat.fillna("")
                df_ext = load_planilha(f_ext)
                if df_ext is None: st.error("Erro no extrato."); return

                uc_e = next((c for c in df_ext.columns if "Número da UC" in c), df_ext.columns[0])
                co_c = next((c for c in df_ext.columns if "Consumo"      in c), None)
                sa_c = next((c for c in df_ext.columns if "Saldo"        in c), None)
                ti_c = next((c for c in df_ext.columns if "Tipo"         in c or "Instalação" in c), None)
                df_ext["UC_NORM"] = df_ext[uc_e].apply(normalize_uc)

                uc_r = next((c for c in df_rat.columns if "UC" in c), df_rat.columns[0])
                ap_c = next((c for c in df_rat.columns if "Apelido" in c), None)
                cn_c = next((c for c in df_rat.columns if "CNPJ"    in c), None)
                df_rat["UC_NORM"] = df_rat[uc_r].apply(normalize_uc)

                sn = {normalize_uc(x["uc"]) for x in ucs_s}
                df_rat = df_rat[~df_rat["UC_NORM"].isin(sn)].copy()

                res = []
                for _, rr in df_rat.iterrows():
                    un = rr["UC_NORM"]
                    ap = rr[ap_c] if ap_c else "—"; cn = rr[cn_c] if cn_c else "—"
                    er = df_ext[df_ext["UC_NORM"] == un]
                    if er.empty: ct=sa=0; tp="monofasico"
                    else:
                        er = er.iloc[0]
                        ct = clean_val(er[co_c]) if co_c else 0
                        sa = clean_val(er[sa_c]) if sa_c else 0
                        tp = str(er[ti_c]).lower() if ti_c else "monofasico"
                    cc  = max(ct-100, 0) if "trif" in tp else max(ct-30, 0)
                    nec = cc*0.20 if ct > 0 and sa/ct > aut else cc
                    res.append({"UC_NORM":un,"UC":rr[uc_r],"Apelido":ap,"CNPJ":cn,
                                "Consumo Comp.":round(cc,2),"Saldo":round(sa,2),"Necessidade":round(nec,2)})

                for nu_item in ucs_n:
                    cc = max(nu_item["consumo"]-30, 0)
                    res.append({"UC_NORM":normalize_uc(nu_item["uc"]),"UC":nu_item["uc"],
                                "Apelido":nu_item.get("apelido","—"),"CNPJ":"—",
                                "Consumo Comp.":round(cc,2),"Saldo":0,"Necessidade":round(cc,2)})

                tn = sum(r["Necessidade"] for r in res)
                for r in res:
                    r["Rateio %"] = round((r["Necessidade"]/gkwh*100), 4) if gkwh > 0 else 0

                soma  = sum(r["Rateio %"] for r in res)
                sobra = round(100.0 - soma, 4)
                if sobra != 0:
                    anorm = normalize_uc(ancora) if ancora else None
                    ia = next((i for i,r in enumerate(res) if r["UC_NORM"]==anorm), None) if anorm else None
                    if ia is not None:
                        res[ia]["Rateio %"] = round(res[ia]["Rateio %"] + sobra, 4)
                    elif tn > 0:
                        for r in res:
                            r["Rateio %"] = round(r["Rateio %"] + sobra*(r["Necessidade"]/tn), 4)

                df_res  = pd.DataFrame(res)
                df_res.insert(0,"#",range(1,len(df_res)+1))
                cols_f  = ["#","UC","Apelido","CNPJ","Consumo Comp.","Saldo","Rateio %"]
                df_show = df_res[[c for c in cols_f if c in df_res.columns]]
                total   = {"#":"TOTAL","UC":"","Apelido":"","CNPJ":"",
                           "Consumo Comp.":df_show["Consumo Comp."].sum(),
                           "Saldo":df_show["Saldo"].sum(),"Rateio %":df_show["Rateio %"].sum()}
                df_show = pd.concat([df_show, pd.DataFrame([total])], ignore_index=True)
                st.success(f"Soma: {df_res['Rateio %'].sum():.4f}% · Geração: {gkwh:,.2f} kWh")
                st.dataframe(df_show, use_container_width=True, hide_index=True)
                st.download_button("Exportar Excel", df_to_excel_bytes(df_show),
                    f"rateio_{sg}_{su}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(str(e)); st.code(traceback.format_exc())

    with tb:
        st.markdown("#### Atualizar Rateio Vigente")
        if not nomes_ger: st.info("Cadastre geradores."); return
        sg2 = st.selectbox("Gerador", nomes_ger, key="upd_ger")
        ug2 = get_usinas_do_gerador(sg2)
        if not ug2: st.warning(f"Nenhuma usina em **{sg2}**."); return
        nm2 = [u.get("ufv") or u["uc"] for u in ug2]
        su2 = st.selectbox("Usina", nm2, key="upd_usi")
        uc2 = ug2[nm2.index(su2)]["uc"]
        gv2 = clean_val(ug2[nm2.index(su2)].get("geracao_estimada",0))
        f_u = st.file_uploader("Planilha do novo rateio", type=["xlsx","xls","csv"], key="up_upd")
        if f_u and st.button("Salvar como Vigente"):
            try:
                df_u = pd.read_excel(f_u, dtype=str) if not f_u.name.endswith(".csv") else pd.read_csv(f_u, dtype=str)
                df_u.columns = df_u.columns.str.strip(); df_u = df_u.fillna("")
                hist = load_rateios(); ch = f"{sg2}||{uc2}"
                hist.setdefault(ch, [])
                hist[ch].append({"versao":len(hist[ch])+1,
                                 "salvo_em":datetime.now().strftime("%d/%m/%Y %H:%M"),
                                 "gerador":sg2,"usina_nome":su2,"uc":uc2,
                                 "geracao_kwh":gv2,"dados":df_u.to_dict(orient="records")})
                save_rateios(hist)
                st.success(f"Rateio v{len(hist[ch])} salvo — **{su2}**")
            except Exception as e: st.error(str(e))

    with tc:
        st.markdown("#### Consultar Rateio")
        if not nomes_ger: st.info("Cadastre geradores."); return
        sg3 = st.selectbox("Gerador", nomes_ger, key="cons_ger")
        ug3 = get_usinas_do_gerador(sg3)
        if not ug3: st.warning(f"Nenhuma usina em **{sg3}**."); return
        nm3 = [u.get("ufv") or u["uc"] for u in ug3]
        su3 = st.selectbox("Usina", nm3, key="cons_usi")
        uc3 = ug3[nm3.index(su3)]["uc"]
        ch3 = f"{sg3}||{uc3}"
        hist = load_rateios(); vers = hist.get(ch3,[])
        if not vers: st.info(f"Nenhum rateio salvo para **{su3}**."); return
        opts = [f"v{v['versao']} — {v['salvo_em']}" for v in vers]
        sv   = st.selectbox("Versão", opts, index=len(opts)-1, key="cons_v")
        iv   = opts.index(sv); vd = vers[iv]
        lbl  = "✅ Vigente" if iv==len(vers)-1 else f"📁 Histórico v{vd['versao']}"
        st.markdown(f"**{lbl}** · {vd['salvo_em']}")
        if vd.get("geracao_kwh"): st.info(f"Geração: **{vd['geracao_kwh']:,.2f} kWh**")
        df_v = pd.DataFrame(vd["dados"])
        st.dataframe(df_v, use_container_width=True, hide_index=True)
        st.download_button(f"Exportar Excel — {su3}", df_to_excel_bytes(df_v),
            f"rateio_vigente_{sg3}_{su3}_v{vd['versao']}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with td:
        st.markdown("#### Buscar Beneficiário por UC")
        col_inp, col_btn = st.columns([4, 1])
        uc_b   = col_inp.text_input("Número da UC", placeholder="Ex: 65447775",
                                    label_visibility="collapsed", key="uc_busca_td")
        buscar = col_btn.button("🔍 Buscar", key="btn_busca_td", use_container_width=True)
        if buscar:
            if not uc_b.strip(): st.warning("Digite uma UC.")
            else:
                unb   = normalize_uc(uc_b)
                hist  = load_rateios(); bo = load_backoffice(); found = False
                for ch, vers in hist.items():
                    if not vers: continue
                    vig = vers[-1]; gen, uc_key = ch.split("||") if "||" in ch else (ch,"")
                    for row in vig["dados"]:
                        for k, v in row.items():
                            if normalize_uc(str(v)) == unb:
                                found = True
                                ap_k = next((k2 for k2 in row if any(x in k2.lower()
                                              for x in ["apelido","titular","nome"])), None)
                                nome_c = row.get(ap_k,"—") if ap_k else "—"
                                pt_k   = next((k2 for k2 in row if "%" in k2 or "rateio" in k2.lower() or "verificado" in k2.lower()), None)
                                pct    = row.get(pt_k,"—") if pt_k else "—"
                                usinas = load_usinas()
                                um     = next((u for u in usinas if normalize_uc(u["uc"])==normalize_uc(uc_key)),{})
                                unome  = um.get("ufv", uc_key)
                                bo_r   = next((b for b in bo if normalize_uc(str(b.get("uc","")))==unb), None)
                                st.success(f"UC **{uc_b}** encontrada.")
                                st.markdown(f"| Campo | Valor |\n|---|---|\n| **UC** | `{uc_b}` |\n| **Nome / Apelido** | {nome_c} |\n| **Gerador** | {gen} |\n| **Usina** | {unome} |\n| **% Rateio** | {pct} |\n| **Versão** | v{vig['versao']} · {vig['salvo_em']} |")
                                if bo_r:
                                    b1, b2, b3 = st.columns(3)
                                    b1.metric("Consumo", f"{bo_r.get('consumo_total',0):,.1f} kWh")
                                    b2.metric("Saldo", f"{bo_r.get('saldo_credito',0):,.1f} kWh")
                                    b3.metric("Tipo Inst.", bo_r.get("tipo_instalacao","—"))
                                break
                        if found: break
                    if found: break
                if not found:
                    st.warning(f"UC **{uc_b}** não encontrada em nenhum rateio ativo.")

    # ══ TAB E — RATEIOS SUNNE ══════════════════════════════════════════════════
    with te:
        st.markdown("#### 🔄 Rateios Ativos — Portal rateios.sunne.com.br")
        st.caption("O robô acessa o portal diariamente às 07h e sincroniza os rateios ativos de cada UG.")

        if not RATEIO_BOT_OK:
            st.error("⚠️ Módulo rateio_sunne_bot não disponível.")
            with st.expander("Detalhes"):
                st.code(_rateio_bot_err_msg)
        else:
            sched_r  = _rateio_bot.load_sync_schedule()
            ultima_r = sched_r.get("ultima","—")

            # Status + config agendamento
            rc1, rc2, rc3 = st.columns(3)
            rc1.markdown(f'<div class="kpi-box"><div class="kpi-value" style="font-size:1.1rem">'
                         f'{sched_r.get("hora","07:00")}</div>'
                         f'<div class="kpi-label">Horário Sync</div></div>', unsafe_allow_html=True)
            rc2.markdown(f'<div class="kpi-box"><div class="kpi-value" style="font-size:1.1rem">'
                         f'{"✅ Ativo" if sched_r.get("auto",True) else "⏸ Pausado"}</div>'
                         f'<div class="kpi-label">Agendamento</div></div>', unsafe_allow_html=True)
            rc3.markdown(f'<div class="kpi-box"><div class="kpi-value" style="font-size:1.1rem">'
                         f'{ultima_r}</div>'
                         f'<div class="kpi-label">Última Sync</div></div>', unsafe_allow_html=True)

            st.write("")
            # Configuração agendamento
            with st.expander("⚙️ Configurar Agendamento"):
                with st.form("form_rateio_sched"):
                    rs1, rs2 = st.columns(2)
                    hora_r = rs1.text_input("Horário (HH:MM)", value=sched_r.get("hora","07:00"))
                    auto_r = rs2.checkbox("Sincronização automática diária",
                                          value=sched_r.get("auto",True))
                    if st.form_submit_button("Salvar"):
                        sched_r["hora"] = hora_r.strip()
                        sched_r["auto"] = auto_r
                        _rateio_bot.save_sync_schedule(sched_r)
                        st.success("✅ Agendamento salvo."); st.rerun()
                st.markdown("**Secrets necessários:**")
                st.code("""[sunne_rateios]
email    = "milena.braga@sunne.com.br"
password = "Milena1968@" """, language="toml")

            # Botão manual
            rodando_r = st.session_state.get("rateio_bot_rodando", False)
            if st.button("🔄 Sincronizar Agora", disabled=rodando_r, key="btn_sync_rateio"):
                st.session_state["rateio_bot_rodando"] = True
                st.rerun()

            if st.session_state.get("rateio_bot_rodando"):
                st.markdown("---")
                prog_r  = st.progress(0.0)
                stat_r  = st.empty()
                log_r_t = st.empty()
                log_live_r = []

                def _on_prog_r(pct, msg):
                    prog_r.progress(min(pct,1.0))
                    stat_r.markdown(f'<div class="alert alert-b">🔄 {msg}</div>',
                                    unsafe_allow_html=True)

                def _on_log_r(entry):
                    log_live_r.append(entry)
                    df_r = pd.DataFrame(log_live_r)
                    log_r_t.dataframe(df_r, use_container_width=True, hide_index=True)

                try:
                    res_r = _rateio_bot.executar_sync_rateios(
                        progress_cb=_on_prog_r, log_cb=_on_log_r)
                    st.session_state["rateio_bot_ultimo"] = res_r
                except Exception as ex_r:
                    st.error(f"Erro: {ex_r}")
                finally:
                    st.session_state["rateio_bot_rodando"] = False
                    prog_r.progress(1.0)
                    stat_r.markdown('<div class="alert alert-g">✅ Sincronização concluída!</div>',
                                    unsafe_allow_html=True)

            # Tabela de rateios ativos
            st.markdown("---")
            st.markdown("#### Rateios Ativos Sincronizados")
            ativos = _rateio_bot.load_rateios_ativos()
            if not ativos:
                st.info("Nenhum rateio sincronizado ainda. Clique em **Sincronizar Agora**.")
            else:
                # Filtros
                rf1, rf2 = st.columns(2)
                ugs_disp = sorted({str(r.get("ug","")) for r in ativos})
                ug_filt  = rf1.selectbox("Filtrar UG", ["Todas"] + ugs_disp, key="rateio_ug_filt")
                status_filt = rf2.selectbox("Status", ["Todos","sincronizado","não_encontrado","erro"],
                                             key="rateio_status_filt")
                ativos_f = ativos
                if ug_filt != "Todas":         ativos_f = [r for r in ativos_f if str(r.get("ug","")) == ug_filt]
                if status_filt != "Todos":     ativos_f = [r for r in ativos_f if r.get("status","") == status_filt]

                df_at = pd.DataFrame(ativos_f)
                cols_at = [c for c in ["ug","status","percentuais_encontrados","coletado_em","texto_bruto"] if c in df_at.columns]
                st.dataframe(df_at[cols_at], use_container_width=True, hide_index=True)

                # Alerta de rebalanceamento
                sincs = [r for r in ativos if r.get("status") == "sincronizado"]
                if sincs:
                    st.markdown("---")
                    st.markdown("**💡 Análise de Necessidade de Rebalanceamento**")
                    usinas_hub = _load(f"{DB}/usinas.json", []) if os.path.exists(f"{DB}/usinas.json") else []
                    ger_list   = load_geracao()
                    mes_atual  = datetime.now().strftime("%m/%Y")
                    for r in sincs:
                        ug_r = str(r.get("ug",""))
                        # Busca geração do mês atual
                        ger_ug = next((g for g in ger_list
                                       if normalize_uc(str(g.get("uc",""))) == normalize_uc(ug_r)
                                       and g.get("competencia","") == mes_atual), None)
                        if ger_ug:
                            inj   = clean_val(ger_ug.get("energia_injetada",0))
                            saldo = clean_val(ger_ug.get("saldo",0))
                            if inj > 0 and saldo/inj > 0.05:
                                nome_u = ger_ug.get("nome_usina", ug_r)
                                st.markdown(
                                    f'<div class="alert alert-y">🚨 UG <b>{ug_r}</b> ({nome_u}): '
                                    f'saldo {saldo:,.0f} kWh = {saldo/inj*100:.1f}% da injeção. '
                                    f'Considere rebalancear o rateio.</div>',
                                    unsafe_allow_html=True)

                # Log de sync
                with st.expander("📋 Log de Sincronizações"):
                    log_s = _rateio_bot.load_sync_log()
                    if log_s:
                        st.dataframe(pd.DataFrame(log_s[:50]), use_container_width=True, hide_index=True)
                    else:
                        st.info("Nenhum log ainda.")

                # Export
                st.download_button("📥 Exportar Rateios CSV",
                    df_at.to_csv(index=False, sep=";").encode("utf-8-sig"),
                    f"rateios_ativos_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv", key="dl_rateio_csv")


def page_faturamento():
    st.title("Faturamento")
    t1, t2, t3 = st.tabs(["Importar","Captura","Inadimplência"])
    with t1:
        c1, c2 = st.columns(2)
        f_r = c1.file_uploader("Rateio")
        f_e = c2.file_uploader("Extrato")
        if f_r and f_e and st.button("Rodar Análise"):
            st.session_state["results"] = analyze_performance(
                load_planilha(f_r), load_planilha(f_e))
            st.success("Análise concluída.")
    res = st.session_state.get("results")
    if res:
        with t2:
            if res["missing"]:
                for comp, items in res["missing"].items():
                    with st.expander(f"⚠️ {comp} — {len(items)} faltantes"):
                        st.download_button(
                            f"Baixar CSV ({comp})",
                            csv_from_list(items,["uc","apelido","usina"],["UC","Apelido","Usina"]),
                            f"faltantes_{comp.replace('/','-')}.csv","text/csv")
                        st.table(pd.DataFrame(items))
            else:
                st.success("Nenhuma fatura faltante.")
        with t3:
            for comp, rows in res["inad"].items():
                g = res["t_gerado"].get(comp, 0.0); v = res["t_vencido"].get(comp, 0.0)
                taxa = (v/g*100) if g > 0 else 0
                st.markdown(f"#### {comp}")
                c1, c2, c3 = st.columns(3)
                c1.metric("Gerado", f"R$ {g:,.2f}"); c2.metric("Vencido", f"R$ {v:,.2f}")
                c3.metric("Inadimplência", f"{taxa:.1f}%")
                with st.expander(f"Clientes ({comp})"):
                    st.table(pd.DataFrame(rows))
            st.markdown("---")
            st.markdown("#### Inadimplência Crítica > 60 dias")
            if res["critical_inad"]:
                st.dataframe(
                    pd.DataFrame(res["critical_inad"]).style.apply(style_critical, axis=1),
                    use_container_width=True, hide_index=True)
            else:
                st.success("Nenhuma.")


# ═══════════════════════════════════════════════════════════════════
# PÁGINA: ANÁLISE BI — multi-modelo (v8)
# ═══════════════════════════════════════════════════════════════════

def _kpi(col, lbl, val, fmt="R$ {:,.2f}", delta=None):
    delta_html = ""
    if delta is not None:
        sinal = "▲" if delta >= 0 else "▼"
        cls   = "kpi-up" if delta >= 0 else "kpi-down"
        delta_html = f'<div class="kpi-delta {cls}">{sinal} {abs(delta):.1f}%</div>'
    col.markdown(
        f'<div class="kpi-box"><div class="kpi-value">{fmt.format(val)}</div>'
        f'<div class="kpi-label">{lbl}</div>{delta_html}</div>',
        unsafe_allow_html=True)

def _calcular_multi_modelo(medicoes_por_modelo: dict, extrato_df, gerador_cfg: dict) -> dict:
    """
    medicoes_por_modelo : {"Associação": {parsed_dict}, "Consórcio": {...}, ...}
    Retorna indicadores consolidados + breakdown por modelo.
    """
    MODELOS_CSS = {"Associação":"modelo-assoc","Consórcio":"modelo-cons","Autoconsumo":"modelo-auto"}
    ind_total = {
        "faturamento_bruto":0.0, "faturamento_liquido":0.0,
        "percentual_sunne":0.0,  "tarifas_bancarias":0.0,
        "outros":0.0, "por_modelo":{}, "por_usina":[],
        "creditos_utilizados":0.0, "inadimplencia_valor":0.0,
        "total_faturado":0.0, "pct_inadimplencia":0.0,
        "media_tarifa_compensavel":0.0, "eficiencia_rateio":0.0,
        "retorno_bruto_estimado":0.0, "tarifa_retorno":0.0,
        "energia_utilizada_fatura":0.0,
    }

    pct_desc  = gerador_cfg.get("pct_desconto_gerador", 0.0)
    pct_admin = gerador_cfg.get("pct_taxa_admin", 0.0)
    enquad    = gerador_cfg.get("enquadramento", "GD1")
    tarifa_b  = TARIFAS_BASE.get(enquad, 0.8182)

    tarifa_retorno_base = tarifa_b * (1 - pct_desc) * (1 - pct_admin)

    tarifas_soma = []

    for modelo, med in medicoes_por_modelo.items():
        m_ind = calcular_indicadores_bi(med, extrato_df if modelo == list(medicoes_por_modelo.keys())[0] else None, gerador_cfg)
        ind_total["por_modelo"][modelo] = {
            "faturamento_bruto":   med.get("faturamento_bruto",0),
            "faturamento_liquido": med.get("faturamento_liquido",0),
            "percentual_sunne":    med.get("percentual_sunne",0),
            "tarifas_bancarias":   med.get("tarifas_bancarias",0),
            "outros":              med.get("outros",0),
            "css":                 MODELOS_CSS.get(modelo,"modelo-assoc"),
        }
        ind_total["faturamento_bruto"]   += med.get("faturamento_bruto",0)
        ind_total["faturamento_liquido"] += med.get("faturamento_liquido",0)
        ind_total["percentual_sunne"]    += med.get("percentual_sunne",0)
        ind_total["tarifas_bancarias"]   += med.get("tarifas_bancarias",0)
        ind_total["outros"]              += med.get("outros",0)
        ind_total["por_usina"].extend(med.get("por_usina",[]))

        if m_ind.get("media_tarifa_compensavel",0) > 0:
            tarifas_soma.append(m_ind["media_tarifa_compensavel"])

    # Extrato (só cruza uma vez — carregado pelo caller)
    if extrato_df is not None and not extrato_df.empty:
        df = extrato_df.copy()
        col_status = next((c for c in df.columns if "Status"           in c), None)
        col_cred   = next((c for c in df.columns if "Créditos"         in c and "Utiliz" in c), None)
        col_tarifa = next((c for c in df.columns if "Tarifa Compensável" in c), None)
        col_ener   = next((c for c in df.columns if "Energia Utilizada" in c), None)
        col_valor  = next((c for c in df.columns if "Total a Pagar"     in c), None)
        for col in [col_cred, col_tarifa, col_ener, col_valor]:
            if col: df[col] = df[col].apply(clean_val)
        mask_ativo = df[col_status].astype(str).str.lower().isin(["pago","vencido"]) if col_status else pd.Series([True]*len(df))
        mask_venc  = df[col_status].astype(str).str.lower() == "vencido" if col_status else pd.Series([False]*len(df))

        if col_cred:  ind_total["creditos_utilizados"]      = float(df.loc[mask_ativo, col_cred].sum())
        if col_ener:  ind_total["energia_utilizada_fatura"] = float(df.loc[mask_ativo, col_ener].sum())
        if col_valor:
            ind_total["total_faturado"]      = float(df.loc[mask_ativo, col_valor].sum())
            ind_total["inadimplencia_valor"] = float(df.loc[mask_venc,  col_valor].sum())
        if col_tarifa:
            vals = df.loc[mask_ativo, col_tarifa]
            vals = vals[vals > 0]
            ind_total["media_tarifa_compensavel"] = float(vals.mean()) if len(vals) > 0 else 0.0
        if ind_total["total_faturado"] > 0:
            ind_total["pct_inadimplencia"] = ind_total["inadimplencia_valor"] / ind_total["total_faturado"] * 100

    # Tarifa retorno — usa média do extrato ou a base calculada
    tarifa_comp = ind_total["media_tarifa_compensavel"] if ind_total["media_tarifa_compensavel"] > 0 else tarifa_retorno_base
    ind_total["tarifa_retorno"] = tarifa_comp * (1 - pct_desc) * (1 - pct_admin) if ind_total["media_tarifa_compensavel"] > 0 else tarifa_retorno_base

    base_kwh = ind_total["creditos_utilizados"] if ind_total["creditos_utilizados"] > 0 else ind_total["faturamento_bruto"]
    ind_total["retorno_bruto_estimado"] = ind_total["tarifa_retorno"] * base_kwh

    if ind_total["faturamento_bruto"] > 0 and ind_total["creditos_utilizados"] > 0:
        ind_total["eficiencia_rateio"] = round(ind_total["creditos_utilizados"] / ind_total["faturamento_bruto"] * 100, 2)

    # Previsão pela injeção (tarifa base)
    ind_total["tarifa_retorno_base"]   = tarifa_retorno_base
    ind_total["enquadramento"]         = enquad
    ind_total["tarifa_base"]           = tarifa_b

    return ind_total


def _render_deducoes(ind: dict):
    """Waterfall de deduções do faturamento bruto → líquido."""
    fat_bruto = ind.get("faturamento_bruto", 0.0)
    pct_sunne = ind.get("percentual_sunne", 0.0)
    tar_banc  = ind.get("tarifas_bancarias", 0.0)
    outros    = ind.get("outros", 0.0)
    fat_liq   = ind.get("faturamento_liquido", 0.0)

    st.markdown("**📊 Composição do Faturamento Líquido**")
    itens = [
        ("Faturamento Bruto",       fat_bruto,            False),
        ("(−) % Sunne",             -abs(pct_sunne),      True),
        ("(−) Tarifas Bancárias",   -abs(tar_banc),       True),
        ("(−) Outros",              -abs(outros),         True),
        ("= Faturamento Líquido",   fat_liq,              False),
    ]
    for label, val, is_ded in itens:
        cls_val = "deducao-val-neg" if is_ded else "deducao-val-pos"
        fmt_val = f"R$ {val:,.2f}"
        st.markdown(
            f'<div class="deducao-row">'
            f'<span class="deducao-label">{label}</span>'
            f'<span class="{cls_val}">{fmt_val}</span>'
            f'</div>', unsafe_allow_html=True)

    # Por modelo
    por_modelo = ind.get("por_modelo", {})
    if por_modelo:
        st.markdown("<br>**Por Modelo de Negócio:**", unsafe_allow_html=True)
        for modelo, md in por_modelo.items():
            css = md.get("css","modelo-assoc")
            st.markdown(f'<div class="modelo-card">'
                        f'<div class="modelo-header {css}">{modelo}</div>', unsafe_allow_html=True)
            sub_itens = [
                ("Bruto",       md.get("faturamento_bruto",0),    False),
                ("(−) Sunne",   -abs(md.get("percentual_sunne",0)), True),
                ("(−) Banc.",   -abs(md.get("tarifas_bancarias",0)), True),
                ("Líquido",     md.get("faturamento_liquido",0),   False),
            ]
            for sl, sv, sd in sub_itens:
                cls_sv = "deducao-val-neg" if sd else "deducao-val-pos"
                st.markdown(f'<div class="deducao-row" style="font-size:12px">'
                            f'<span class="deducao-label">{sl}</span>'
                            f'<span class="{cls_sv}">R$ {sv:,.2f}</span>'
                            f'</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)


def page_bi_analise():
    st.title("📊 Análise de Faturamento · BI")

    geradores     = load_geradores()
    nomes_ger     = sorted({g["gerador"] for g in geradores})
    hist_analises = load_analises()

    tab_importar, tab_historico = st.tabs(["Importar & Calcular","📁 Histórico & Relatórios"])

    # ══ IMPORTAR ══════════════════════════════════════════════════════════════
    with tab_importar:
        st.markdown("#### Configuração do Período")
        col_ger, col_comp, col_enq, col_desc, col_taxa = st.columns(5)

        if not nomes_ger:
            st.info("Cadastre geradores primeiro.")
            return

        ger_sel  = col_ger.selectbox("Gerador", nomes_ger, key="bi_ger_sel")
        comp_sel = col_comp.text_input("Competência (MM/AAAA)",
                                       value=datetime.now().strftime("%m/%Y"), key="bi_comp")
        enquad   = col_enq.selectbox("Enquadramento", list(TARIFAS_BASE.keys()), key="bi_enq",
                                     help="GD1=0.8182 / GD2=0.64788")
        pct_desc = col_desc.number_input("% Desconto Gerador", 0.0, 1.0, 0.20, 0.01,
                                         format="%.2f", key="bi_desc")
        pct_taxa = col_taxa.number_input("% Taxa Admin Sunne", 0.0, 1.0, 0.07, 0.01,
                                         format="%.2f", key="bi_taxa")

        tarifa_b = TARIFAS_BASE[enquad]
        tarifa_r = tarifa_b * (1 - pct_desc) * (1 - pct_taxa)
        st.caption(f"🔢 Tarifa Base {enquad}: R$ {tarifa_b:.5f} → Tarifa Retorno: **R$ {tarifa_r:.5f}**/kWh")

        # ── Múltiplos relatórios ──────────────────────────────────────────────
        st.markdown("#### Relatórios de Medição por Modelo de Negócio")
        st.caption("Suba até 3 relatórios — um por modelo de contrato.")

        medicao_uploads = {}
        for i in range(3):
            mc1, mc2 = st.columns([2, 5])
            modelo_esc = mc1.selectbox(
                f"Modelo #{i+1}", ["— não usar —"] + MODELOS_NEGOCIO,
                index=i+1 if i < len(MODELOS_NEGOCIO) else 0, key=f"bi_modelo_{i}")
            if modelo_esc != "— não usar —":
                arq = mc2.file_uploader(f"📋 Relatório {modelo_esc}",
                                        type=["xlsx","xls"], key=f"bi_med_{i}")
                if arq:
                    medicao_uploads[modelo_esc] = arq

        st.markdown("#### Extrato Detalhado *(opcional)*")
        f_extrato = st.file_uploader("📄 Extrato Detalhado (xlsx/csv)",
                                     type=["xlsx","xls","csv"], key="bi_extrato")

        chave     = f"{ger_sel}||{comp_sel}"
        ja_existe = chave in hist_analises and comp_sel in hist_analises.get(chave, {})
        if ja_existe:
            st.markdown('<div class="alert alert-b">📂 Análise já processada. Carregada do histórico.</div>',
                        unsafe_allow_html=True)

        btn_calc = st.button("⚡ Calcular Indicadores", type="primary", key="btn_bi_calc",
                             disabled=(not medicao_uploads and not ja_existe))

        if btn_calc:
            with st.spinner("Processando…"):
                try:
                    gerador_cfg = {"pct_desconto_gerador": float(pct_desc),
                                   "pct_taxa_admin": float(pct_taxa), "enquadramento": enquad}
                    medicoes_parsed = {}
                    if medicao_uploads:
                        for modelo, arq in medicao_uploads.items():
                            medicoes_parsed[modelo] = parse_relatorio_medicao(arq.read())
                    elif ja_existe:
                        medicoes_parsed = hist_analises[chave][comp_sel].get("medicoes_parsed", {})

                    extrato_df = None
                    if f_extrato:
                        extrato_df = load_planilha(f_extrato)
                        if extrato_df is None or extrato_df.empty: extrato_df = None

                    ind = _calcular_multi_modelo(medicoes_parsed, extrato_df, gerador_cfg)

                    ind_anterior = None
                    try:
                        mm, aa = comp_sel.split("/")
                        comp_ant = f"{int(mm)-1:02d}/{aa}" if int(mm) > 1 else f"12/{int(aa)-1}"
                        if chave in hist_analises and comp_ant in hist_analises[chave]:
                            ind_anterior = hist_analises[chave][comp_ant].get("indicadores")
                    except: pass
                    ind["indicadores_anterior"] = ind_anterior

                    hist_analises.setdefault(chave, {})
                    hist_analises[chave][comp_sel] = {
                        "gerador": ger_sel, "competencia": comp_sel,
                        "processado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "gerador_cfg": gerador_cfg, "medicoes_parsed": medicoes_parsed,
                        "modelos": list(medicoes_parsed.keys()), "indicadores": ind,
                    }
                    save_analises(hist_analises)
                    st.session_state["bi_resultado_chave"] = chave
                    st.session_state["bi_resultado_comp"]  = comp_sel
                    st.success("✅ Análise calculada e salva!"); st.rerun()
                except Exception as e:
                    st.error(f"Erro: {e}"); st.code(traceback.format_exc())

        # ── Exibir resultado ──────────────────────────────────────────────────
        res_chave = st.session_state.get("bi_resultado_chave")
        res_comp  = st.session_state.get("bi_resultado_comp")
        if not (res_chave and res_comp and res_chave in hist_analises): return
        dados_comp = hist_analises[res_chave].get(res_comp)
        if not dados_comp: return
        ind     = dados_comp.get("indicadores", {})
        ind_ant = ind.get("indicadores_anterior") or {}

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown(f"### Resultado · {dados_comp['gerador']} · {res_comp}")
        st.caption(f"Processado em: {dados_comp.get('processado_em','—')} · "
                   f"Modelos: {', '.join(dados_comp.get('modelos',[]))}")

        def _delta_pct(atual, ant, campo):
            a = ant.get(campo, 0)
            return ((atual - a) / a * 100) if a > 0 else None

        k1,k2,k3,k4,k5,k6 = st.columns(6)
        _kpi(k1,"Faturamento Bruto",  ind.get("faturamento_bruto",0),   "R$ {:,.2f}",
             _delta_pct(ind.get("faturamento_bruto",0), ind_ant, "faturamento_bruto"))
        _kpi(k2,"Faturamento Líquido",ind.get("faturamento_liquido",0), "R$ {:,.2f}",
             _delta_pct(ind.get("faturamento_liquido",0), ind_ant, "faturamento_liquido"))
        _kpi(k3,"Créditos Utilizados",ind.get("creditos_utilizados",0), "{:,.1f} kWh")
        _kpi(k4,"Eficiência Rateio",  ind.get("eficiencia_rateio",0),   "{:.1f}%")
        _kpi(k5,"Inadimplência",       ind.get("pct_inadimplencia",0),   "{:.1f}%",
             _delta_pct(ind.get("pct_inadimplencia",0), ind_ant, "pct_inadimplencia"))
        _kpi(k6,"Retorno Estimado",    ind.get("retorno_bruto_estimado",0),"R$ {:,.2f}")

        st.write("")
        col_ded, col_ins = st.columns([2, 3])
        with col_ded:
            _render_deducoes(ind)
        with col_ins:
            st.markdown("**🚨 Insights Automáticos**")
            for nivel, msg in gerar_insights(ind, ind_ant if ind_ant else None):
                st.markdown(f'<div class="alert alert-{nivel}">{msg}</div>', unsafe_allow_html=True)
            st.write("")
            gcfg = dados_comp.get("gerador_cfg",{})
            enq  = gcfg.get("enquadramento","GD1")
            tb   = TARIFAS_BASE.get(enq, 0.8182)
            tr   = tb*(1-gcfg.get("pct_desconto_gerador",0))*(1-gcfg.get("pct_taxa_admin",0))
            st.markdown("**📐 Composição da Tarifa**")
            st.markdown(
                f'<div class="deducao-row"><span class="deducao-label">Enquadramento</span><span style="font-weight:600">{enq}</span></div>'
                f'<div class="deducao-row"><span class="deducao-label">Tarifa Base</span><span style="font-weight:600">R$ {tb:.5f}</span></div>'
                f'<div class="deducao-row"><span class="deducao-label">(−) Desconto Gerador</span><span class="deducao-val-neg">{gcfg.get("pct_desconto_gerador",0)*100:.1f}%</span></div>'
                f'<div class="deducao-row"><span class="deducao-label">(−) Taxa Admin Sunne</span><span class="deducao-val-neg">{gcfg.get("pct_taxa_admin",0)*100:.1f}%</span></div>'
                f'<div class="deducao-row"><span class="deducao-label">= Tarifa Retorno</span><span class="deducao-val-pos">R$ {tr:.5f}/kWh</span></div>',
                unsafe_allow_html=True)

        por_usina = ind.get("por_usina",[])
        if por_usina:
            st.markdown("**Performance por Usina**")
            df_pu = pd.DataFrame(por_usina)
            if len(df_pu.columns) >= 7:
                df_pu.columns = ["Usina","Fat. Bruto","% Sunne","Tar. Banc.",
                                 "Fat. Líquido","Conta Energia","Marketplace"]
            st.dataframe(df_pu, use_container_width=True, hide_index=True)

        st.markdown("---")
        exp_c1, _ = st.columns([2,4])
        if exp_c1.button("📥 Exportar Excel (Investidor)", key="btn_exp_bi"):
            buf = io.BytesIO()
            rows_kpi = [
                {"Indicador":"Faturamento Bruto",        "Valor":ind.get("faturamento_bruto",0)},
                {"Indicador":"Faturamento Líquido",      "Valor":ind.get("faturamento_liquido",0)},
                {"Indicador":"Deducao % Sunne",          "Valor":ind.get("percentual_sunne",0)},
                {"Indicador":"Tarifas Bancárias",        "Valor":ind.get("tarifas_bancarias",0)},
                {"Indicador":"Outros",                   "Valor":ind.get("outros",0)},
                {"Indicador":"Créditos Utilizados kWh",  "Valor":ind.get("creditos_utilizados",0)},
                {"Indicador":"Eficiência Rateio %",      "Valor":ind.get("eficiencia_rateio",0)},
                {"Indicador":"Inadimplência R$",         "Valor":ind.get("inadimplencia_valor",0)},
                {"Indicador":"Inadimplência %",          "Valor":ind.get("pct_inadimplencia",0)},
                {"Indicador":"Tarifa Retorno R$/kWh",    "Valor":ind.get("tarifa_retorno",0)},
                {"Indicador":"Retorno Bruto Estimado",   "Valor":ind.get("retorno_bruto_estimado",0)},
            ]
            rows_mod = [{"Modelo":m,"Fat. Bruto":md.get("faturamento_bruto",0),
                         "% Sunne":md.get("percentual_sunne",0),
                         "Tar. Banc.":md.get("tarifas_bancarias",0),
                         "Fat. Líquido":md.get("faturamento_liquido",0)}
                        for m, md in ind.get("por_modelo",{}).items()]
            with pd.ExcelWriter(buf, engine='openpyxl') as w:
                pd.DataFrame(rows_kpi).to_excel(w, index=False, sheet_name="Indicadores")
                if rows_mod: pd.DataFrame(rows_mod).to_excel(w, index=False, sheet_name="Por Modelo")
                if por_usina: pd.DataFrame(por_usina).to_excel(w, index=False, sheet_name="Por Usina")
            st.download_button("⬇ Baixar Excel", buf.getvalue(),
                f"bi_{ger_sel}_{res_comp.replace('/','-')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_bi_excel")

    # ══ HISTÓRICO ══════════════════════════════════════════════════════════════
    with tab_historico:
        st.markdown("#### Histórico de Análises por Gerador")
        if not hist_analises:
            st.info("Nenhuma análise processada ainda."); return

        ger_hist_opts = sorted({(ch.split("||")[0] if "||" in ch else ch) for ch in hist_analises})
        ger_hist_sel  = st.selectbox("Gerador", ger_hist_opts, key="hist_ger_sel")
        chaves_ger    = [ch for ch in hist_analises
                         if (ch.split("||")[0] if "||" in ch else ch) == ger_hist_sel]
        if not chaves_ger: st.info(f"Nenhuma análise para {ger_hist_sel}."); return

        todas_comps = [(ch, comp, dados)
                       for ch in chaves_ger
                       for comp, dados in hist_analises[ch].items()]
        todas_comps.sort(key=lambda x: comp_sort_key(x[1]), reverse=True)
        st.markdown(f"**{len(todas_comps)} competência(s)**")

        for chave, comp, dados in todas_comps:
            ind  = dados.get("indicadores",{})
            mods = dados.get("modelos",[])
            col_info, col_kpis, col_btn = st.columns([2,5,1])
            col_info.markdown(f"**{comp}**")
            if mods: col_info.caption(" · ".join(mods))
            col_info.caption(f"{dados.get('processado_em','—')}")
            with col_kpis:
                kk1,kk2,kk3,kk4 = st.columns(4)
                for kc,klbl,kval,kfmt in [
                    (kk1,"Bruto",  ind.get("faturamento_bruto",0),   "R$ {:,.0f}"),
                    (kk2,"Líquido",ind.get("faturamento_liquido",0), "R$ {:,.0f}"),
                    (kk3,"Inadimp",ind.get("pct_inadimplencia",0),   "{:.1f}%"),
                    (kk4,"Efic.",  ind.get("eficiencia_rateio",0),   "{:.1f}%"),
                ]:
                    kc.markdown(f'<div class="kpi-box" style="padding:.5rem .7rem">'
                                f'<div class="kpi-value" style="font-size:.95rem">{kfmt.format(kval)}</div>'
                                f'<div class="kpi-label">{klbl}</div></div>', unsafe_allow_html=True)
            if col_btn.button("🔍", key=f"hist_ver_{chave}_{comp}"):
                dialog_bi_dashboard(chave, comp)
            st.markdown("<hr style='margin:6px 0;border:none;border-top:1px solid rgba(51,0,26,.06)'>",
                        unsafe_allow_html=True)

        if len(todas_comps) >= 2:
            st.markdown("---")
            st.markdown("#### Tendência Consolidada")
            df_tend = pd.DataFrame([
                {"Competência":comp,
                 "Fat. Bruto":d.get("indicadores",{}).get("faturamento_bruto",0),
                 "Fat. Líquido":d.get("indicadores",{}).get("faturamento_liquido",0),
                 "Inadimpl. %":d.get("indicadores",{}).get("pct_inadimplencia",0),
                 "Eficiência %":d.get("indicadores",{}).get("eficiencia_rateio",0),
                 "Retorno Est.":d.get("indicadores",{}).get("retorno_bruto_estimado",0),
                 "Modelos":", ".join(d.get("modelos",[]))}
                for _,comp,d in sorted(todas_comps, key=lambda x: comp_sort_key(x[1]))
            ])
            st.dataframe(df_tend, use_container_width=True, hide_index=True)
            buf2 = io.BytesIO()
            with pd.ExcelWriter(buf2, engine='openpyxl') as w:
                df_tend.to_excel(w, index=False, sheet_name="Tendência")
            st.download_button("📥 Exportar Tendência Excel", buf2.getvalue(),
                f"tendencia_{ger_hist_sel}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def page_automacao():
    st.title("🤖 Automação · Captura de Faturas")

    # ── Aviso se Selenium não disponível ─────────────────────────────────────
    if not ROBOT_DISPONIVEL:
        st.error("⚠️ Módulo de automação não disponível.")
        with st.expander("Detalhes do erro de importação"):
            st.code(_robot_err_msg if "ROBOT_DISPONIVEL" in dir() else "robot_captura.py não encontrado")
        st.markdown("**Para ativar, adicione ao seu repositório:**")
        col1, col2 = st.columns(2)
        col1.markdown("**`packages.txt`**")
        col1.code("chromium-driver\nchromium", language="text")
        col2.markdown("**`requirements.txt`**")
        col2.code("selenium>=4.18.0\nwebdriver-manager>=4.0.1\npdfplumber>=0.11.0", language="text")
        st.markdown("**`.streamlit/secrets.toml`**")
        st.code('[sunne_portal]\nemail    = "milena.braga@sunne.com.br"\npassword = "Milena@2025"', language="toml")
        return

    usinas   = load_usinas()
    sched    = _robot.load_schedule()
    log_hist = _robot.load_log()

    # ── TABS principais ───────────────────────────────────────────────────────
    tab_painel, tab_config, tab_log, tab_template = st.tabs([
        "🚀 Painel de Controle", "⏰ Agendamento", "📋 Log de Execuções", "🧠 Ensinar Robô"
    ])

    # ══════════════════════════════════════════════════════
    # TAB 1 – PAINEL DE CONTROLE
    # ══════════════════════════════════════════════════════
    with tab_painel:

        # KPIs rápidos
        total_ucs = len(usinas)
        ger_hoje  = [e for e in log_hist if e.get("ts","").startswith(datetime.now().strftime("%d/%m/%Y"))]
        baixados  = len([e for e in ger_hoje if e.get("status") == "baixado"])
        nd        = len([e for e in ger_hoje if e.get("status") == "não_disponível"])
        erros     = len([e for e in ger_hoje if e.get("status") in ("erro","erro_geral")])
        pendentes = len([e for e in ger_hoje if e.get("status") == "precisa_template"])

        k1, k2, k3, k4, k5 = st.columns(5)
        for col, lbl, val, cor in [
            (k1, "UCs Cadastradas",  total_ucs,   "#33001A"),
            (k2, "Baixadas Hoje",    baixados,    "#0B7A5F"),
            (k3, "Não Disponíveis",  nd,          "#7A5010"),
            (k4, "Precisa Template", pendentes,   "#A84010"),
            (k5, "Erros Hoje",       erros,       "#C41230"),
        ]:
            col.markdown(
                f'<div class="kpi-box"><div class="kpi-value" style="color:{cor}">{val}</div>'
                f'<div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)

        st.write("")

        # Status agendamento
        auto_on = sched.get("auto", False)
        hora    = sched.get("hora", "08:00")
        ultima  = sched.get("ultima_execucao", "—")
        if auto_on:
            st.markdown(
                f'<div class="alert alert-g">⏰ <b>Agendamento ativo</b> — roda todos os dias às <b>{hora}</b> '
                f'· Última execução: <b>{ultima}</b></div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="alert alert-y">⏸ Agendamento desativado. '
                'Configure na aba <b>Agendamento</b>.</div>', unsafe_allow_html=True)

        st.write("")
        st.markdown("#### Configurar Varredura")

        c1, c2 = st.columns([2, 3])
        mes_atual = _robot.MESES_PT[datetime.now().month]
        comp_sel  = c1.selectbox("Competência", list(_robot.MESES_PT.values()),
                                  index=list(_robot.MESES_PT.values()).index(mes_atual),
                                  key="auto_comp")

        ger_opts  = sorted({u.get("gerador","—") for u in usinas})
        ger_filtro= c2.multiselect("Filtrar por Gerador (vazio = todos)", ger_opts, key="auto_ger")

        ucs_alvo = []
        if ger_filtro:
            ucs_alvo = [u["uc"] for u in usinas if u.get("gerador","") in ger_filtro]
            st.caption(f"🎯 {len(ucs_alvo)} UC(s) selecionada(s) dos geradores filtrados")
        else:
            st.caption(f"🎯 Todas as {total_ucs} UC(s) cadastradas serão varridas")

        # ── Botão principal ───────────────────────────────────────────────────
        rodando = st.session_state.get("robot_rodando", False)

        btn_col, status_col = st.columns([2, 5])
        iniciar = btn_col.button(
            "🚀 Iniciar Varredura" if not rodando else "⏳ Executando…",
            disabled=rodando, use_container_width=True, key="btn_iniciar_robot"
        )

        # Auto-trigger via agendamento
        if st.session_state.pop("robot_auto_trigger", False):
            iniciar = True

        if iniciar and not rodando:
            st.session_state["robot_rodando"] = True
            st.session_state["robot_log_live"] = []
            st.rerun()

        # ── Execução (quando rodando=True) ────────────────────────────────────
        if rodando:
            st.markdown("---")
            st.markdown("#### ⚡ Execução em Andamento")

            prog_bar  = st.progress(0.0)
            status_tx = st.empty()
            log_table = st.empty()

            log_live: list = st.session_state.get("robot_log_live", [])

            def on_progress(pct, msg):
                prog_bar.progress(min(pct, 1.0))
                status_tx.markdown(
                    f'<div class="alert alert-b">🔄 {msg}</div>', unsafe_allow_html=True)

            def on_log(entry):
                log_live.append(entry)
                st.session_state["robot_log_live"] = log_live
                _render_log_table(log_table, log_live)

            try:
                resultado = _robot.executar_varredura(
                    progress_cb    = on_progress,
                    log_cb         = on_log,
                    competencia_mes= comp_sel,
                    ucs_alvo       = ucs_alvo if ucs_alvo else None,
                )
                st.session_state["robot_ultimo_resultado"] = resultado
            except Exception as ex:
                st.error(f"Erro fatal na varredura: {ex}")
                st.code(traceback.format_exc())
            finally:
                st.session_state["robot_rodando"] = False
                prog_bar.progress(1.0)
                status_tx.markdown(
                    '<div class="alert alert-g">✅ Varredura finalizada!</div>',
                    unsafe_allow_html=True)
                st.rerun()

        # ── Resultado da última varredura ─────────────────────────────────────
        ultimo = st.session_state.get("robot_ultimo_resultado")
        if ultimo:
            st.markdown("---")
            st.markdown("#### Resultado da Última Varredura")
            _render_log_table(st, ultimo)
            if st.button("🔄 Ir para Geração (atualizado)", key="btn_ver_geracao"):
                st.session_state["page"] = "geracao"
                st.rerun()

    # ══════════════════════════════════════════════════════
    # TAB 2 – AGENDAMENTO
    # ══════════════════════════════════════════════════════
    with tab_config:
        st.markdown("#### ⏰ Agendamento Diário Automático")
        st.caption("O robô verificará o horário a cada interação com o sistema. "
                   "Para garantia em produção, considere um cron externo ou Streamlit Community Cloud com auto-refresh.")

        with st.form("form_schedule"):
            fc1, fc2 = st.columns(2)
            hora_nova = fc1.text_input("Horário de execução (HH:MM)",
                                       value=sched.get("hora","08:00"),
                                       placeholder="08:00")
            auto_nova = fc2.checkbox("Ativar execução automática diária",
                                     value=sched.get("auto", False))

            st.markdown("**Configurações de escopo**")
            sc1, sc2 = st.columns(2)
            comp_auto = sc1.selectbox("Competência padrão para varredura automática",
                                      list(_robot.MESES_PT.values()),
                                      index=list(_robot.MESES_PT.values()).index(
                                          _robot.MESES_PT[datetime.now().month]))
            _ = sc2.info("A varredura automática sempre usa todas as UCs cadastradas.")

            salvar_sched = st.form_submit_button("💾 Salvar Configuração", use_container_width=True)

        if salvar_sched:
            import re
            if not re.match(r"^\d{2}:\d{2}$", hora_nova.strip()):
                st.error("Formato inválido. Use HH:MM (ex: 08:00).")
            else:
                sched["hora"]  = hora_nova.strip()
                sched["auto"]  = auto_nova
                sched["comp_auto"] = comp_auto
                _robot.save_schedule(sched)
                st.success(f"✅ Agendamento {'ativado' if auto_nova else 'desativado'} às {hora_nova}.")
                st.rerun()

        st.markdown("---")
        st.markdown("**Credenciais do Portal** *(gerenciadas via `st.secrets`)*")
        st.code("""# .streamlit/secrets.toml
[sunne_portal]
email    = "milena.braga@sunne.com.br"
password = "Milena@2025"
""", language="toml")
        st.caption("No Streamlit Cloud: Settings → Secrets → cole o bloco acima.")

        st.markdown("**Dependências necessárias**")
        col_dep1, col_dep2 = st.columns(2)
        col_dep1.markdown("**`packages.txt`**")
        col_dep1.code("chromium-driver\nchromium", language="text")
        col_dep2.markdown("**`requirements.txt`** *(adicionar)*")
        col_dep2.code("selenium>=4.18.0\nwebdriver-manager>=4.0.1\npdfplumber>=0.11.0",
                      language="text")

    # ══════════════════════════════════════════════════════
    # TAB 3 – LOG COMPLETO
    # ══════════════════════════════════════════════════════
    with tab_log:
        st.markdown("#### 📋 Histórico de Execuções")

        log_hist = _robot.load_log()  # recarrega
        if not log_hist:
            st.info("Nenhuma execução registrada ainda.")
        else:
            # Filtros
            lc1, lc2, lc3 = st.columns(3)
            status_opts = ["Todos"] + sorted({e.get("status","") for e in log_hist})
            filt_status = lc1.selectbox("Status", status_opts, key="log_filt_status")
            filt_uc     = lc2.text_input("Filtrar UC", placeholder="todos", key="log_filt_uc")
            filt_comp   = lc3.text_input("Filtrar Competência", placeholder="ex: 04/2026", key="log_filt_comp")

            filtrado = log_hist
            if filt_status != "Todos":
                filtrado = [e for e in filtrado if e.get("status") == filt_status]
            if filt_uc.strip():
                filtrado = [e for e in filtrado if filt_uc.strip() in str(e.get("uc",""))]
            if filt_comp.strip():
                filtrado = [e for e in filtrado if filt_comp.strip() in str(e.get("comp",""))]

            st.caption(f"{len(filtrado)} de {len(log_hist)} registros")

            df_log = pd.DataFrame(filtrado)
            if not df_log.empty:
                cols_show = [c for c in ["ts","uc","comp","status","injetada","saldo","obs"] if c in df_log.columns]
                df_log = df_log[cols_show].copy()
                df_log.columns = [{"ts":"Timestamp","uc":"UC","comp":"Competência",
                                   "status":"Status","injetada":"Injetada (kWh)",
                                   "saldo":"Saldo (kWh)","obs":"Observação"}.get(c,c)
                                  for c in cols_show]

                def _colorir_status(val):
                    cores = {"baixado":"background:#EDFCF6;color:#0A5040",
                             "não_disponível":"background:#FFFBEC;color:#6B4A00",
                             "precisa_template":"background:#FFF3EC;color:#A84010",
                             "erro":"background:#FFF0F3;color:#8B1530",
                             "erro_geral":"background:#FFF0F3;color:#8B1530"}
                    return cores.get(str(val), "")

                st.dataframe(
                    df_log.style.applymap(_colorir_status, subset=["Status"]
                                          if "Status" in df_log.columns else []),
                    use_container_width=True, hide_index=True)

                st.download_button(
                    "📥 Exportar Log CSV",
                    df_log.to_csv(index=False, sep=";").encode("utf-8-sig"),
                    f"log_captura_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv", key="dl_log_csv")

            # Pendentes de template
            pendentes_log = [e for e in log_hist if e.get("status") == "precisa_template"]
            if pendentes_log:
                st.markdown("---")
                st.markdown(
                    f'<div class="alert alert-y">⚠️ <b>{len(pendentes_log)}</b> UC(s) com fatura baixada mas layout desconhecido. '
                    f'Acesse a aba <b>🧠 Ensinar Robô</b> para treinar a extração.</div>',
                    unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # TAB 4 – ENSINAR ROBÔ (template learning)
    # ══════════════════════════════════════════════════════
    with tab_template:
        st.markdown("#### 🧠 Ensinar Robô · Layout Desconhecido")
        st.markdown(
            "Quando o robô baixa uma fatura mas não consegue extrair os dados automaticamente, "
            "você pode subir o PDF manualmente e informar os valores corretos. "
            "O sistema aprenderá o layout para extrações futuras.")

        st.markdown('<div class="alert alert-b">💡 O aprendizado é salvo por <b>concessionária</b>. '
                    'Uma vez ensinado, todas as UCs da mesma concessionária serão extraídas automaticamente.</div>',
                    unsafe_allow_html=True)

        # Mostra PDFs pendentes salvos pelo robô
        pdfs_pendentes = []
        for f in os.listdir("database") if os.path.exists("database") else []:
            if f.startswith("pdf_pendente_") and f.endswith(".pdf"):
                uc_pend = f.replace("pdf_pendente_","").replace(".pdf","")
                pdfs_pendentes.append(uc_pend)

        if pdfs_pendentes:
            st.markdown(f"**📂 {len(pdfs_pendentes)} PDF(s) pendente(s) de treinamento:**")
            uc_treinar = st.selectbox("Selecionar UC pendente", pdfs_pendentes, key="uc_treinar_sel")
            pdf_path   = f"database/pdf_pendente_{uc_treinar}.pdf"

            if os.path.exists(pdf_path):
                with open(pdf_path, "rb") as fp:
                    pdf_bytes_pend = fp.read()
                st.download_button(f"📄 Visualizar fatura UC {uc_treinar}",
                                   pdf_bytes_pend, f"fatura_{uc_treinar}.pdf",
                                   "application/pdf", key="dl_pend_pdf")

                st.markdown("**Informe os valores corretos desta fatura:**")
                tp1, tp2, tp3, tp4 = st.columns(4)
                conc_treinar = tp1.text_input("Concessionária", placeholder="ex: ENEL CE", key="conc_treinar")
                inj_treinar  = tp2.number_input("Energia Injetada (kWh)", min_value=0.0, step=0.1, key="inj_treinar")
                sal_treinar  = tp3.number_input("Saldo kWh", min_value=0.0, step=0.1, key="sal_treinar")
                con_treinar  = tp4.number_input("Consumo kWh", min_value=0.0, step=0.1, key="con_treinar")

                if st.button("✅ Confirmar e Treinar", key="btn_treinar_pend", use_container_width=True):
                    if not conc_treinar.strip():
                        st.warning("Informe a concessionária.")
                    else:
                        template = _robot.aprender_template_de_pdf(
                            pdf_bytes_pend, uc_treinar, conc_treinar.strip(),
                            inj_treinar, sal_treinar, con_treinar)
                        # Salva geração já com os valores informados
                        _robot.salvar_geracao_hub(uc_treinar, {
                            "energia_injetada": inj_treinar,
                            "saldo":            sal_treinar,
                            "consumo":          con_treinar,
                            "concessionaria":   conc_treinar,
                            "confianca":        1.0,
                        }, datetime.now().strftime("%m/%Y"))
                        # Remove PDF pendente
                        try: os.remove(pdf_path)
                        except Exception: pass
                        st.success(f"✅ Template salvo para '{conc_treinar}' · Geração registrada no Hub!")
                        st.toast("🧠 Robô treinado com sucesso!", icon="🎉")
                        st.rerun()
            st.markdown("---")

        # Upload manual de nova fatura para ensinar
        st.markdown("**Ou suba uma fatura manualmente para treinar:**")
        mu1, mu2 = st.columns(2)
        uc_manual    = mu1.text_input("UC da fatura", key="uc_manual_template")
        conc_manual  = mu2.text_input("Concessionária", placeholder="ex: EQUATORIAL PI", key="conc_manual_template")

        pdf_manual = st.file_uploader("📎 Fatura PDF", type=["pdf"], key="pdf_manual_upload")

        if pdf_manual:
            # Preview texto extraído
            try:
                import pdfplumber as _pp
                with _pp.open(io.BytesIO(pdf_manual.read())) as _pdf:
                    texto_preview = "\n".join(p.extract_text() or "" for p in _pdf.pages[:2])
                pdf_manual.seek(0)
                with st.expander("👁 Texto extraído do PDF (primeiras 2 páginas)"):
                    st.text(texto_preview[:2000])
            except Exception:
                pass

            mv1, mv2, mv3 = st.columns(3)
            inj_m = mv1.number_input("Energia Injetada (kWh)", min_value=0.0, step=0.1, key="inj_m")
            sal_m = mv2.number_input("Saldo kWh",              min_value=0.0, step=0.1, key="sal_m")
            con_m = mv3.number_input("Consumo kWh",            min_value=0.0, step=0.1, key="con_m")

            if st.button("🧠 Treinar com este PDF", key="btn_treinar_manual", use_container_width=True):
                if not uc_manual.strip() or not conc_manual.strip():
                    st.warning("Informe UC e Concessionária.")
                else:
                    pdf_bytes_m = pdf_manual.read() if pdf_manual.tell() == 0 else None
                    if not pdf_bytes_m:
                        pdf_manual.seek(0); pdf_bytes_m = pdf_manual.read()
                    template = _robot.aprender_template_de_pdf(
                        pdf_bytes_m, uc_manual.strip(), conc_manual.strip(),
                        inj_m, sal_m, con_m)
                    _robot.salvar_geracao_hub(uc_manual.strip(), {
                        "energia_injetada": inj_m,
                        "saldo":            sal_m,
                        "consumo":          con_m,
                        "concessionaria":   conc_manual.strip(),
                        "confianca":        1.0,
                    }, datetime.now().strftime("%m/%Y"))
                    st.success(f"✅ Template '{conc_manual}' aprendido · UC {uc_manual} salva em Geração!")
                    st.toast("🧠 Treinamento concluído!", icon="✅")

        # Templates já salvos
        templates = _robot.load_templates()
        if templates:
            st.markdown("---")
            st.markdown(f"**📚 {len(templates)} template(s) aprendido(s):**")
            rows_t = []
            for k, v in templates.items():
                rows_t.append({
                    "Chave": k,
                    "Concessionária": v.get("concessionaria","—"),
                    "UC": v.get("uc","—"),
                    "Aprendido em": v.get("aprendido_em","—"),
                    "Manual": "✅" if v.get("manual") else "🤖",
                    "Tem bbox": "✅" if v.get("bbox_injetada") else "❌",
                })
            st.dataframe(pd.DataFrame(rows_t), use_container_width=True, hide_index=True)

            del_key = st.selectbox("Remover template:", ["— selecione —"] + list(templates.keys()),
                                   key="del_template_sel")
            if del_key != "— selecione —" and st.button("🗑 Remover", key="btn_del_tpl"):
                del templates[del_key]
                _robot.save_templates(templates)
                st.success(f"Template '{del_key}' removido.")
                st.rerun()


# ─── Helper interno: renderiza tabela de log ──────────────────────────────────
def _render_log_table(container, entries: list):
    if not entries:
        return
    df = pd.DataFrame(entries)
    cols_show = [c for c in ["ts","uc","status","injetada","saldo","obs"] if c in df.columns]
    df = df[cols_show].copy()
    df.columns = [{"ts":"Timestamp","uc":"UC","status":"Status",
                   "injetada":"Injetada kWh","saldo":"Saldo kWh","obs":"Obs"}.get(c,c)
                  for c in cols_show]

    STATUS_ICONS = {
        "baixado":          "✅ baixado",
        "não_disponível":   "⬜ não disponível",
        "precisa_template": "⚠️ precisa template",
        "erro":             "❌ erro",
        "erro_geral":       "❌ erro geral",
    }
    if "Status" in df.columns:
        df["Status"] = df["Status"].apply(lambda v: STATUS_ICONS.get(str(v), str(v)))

    container.dataframe(df, use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def page_integracoes():
    st.title("🔗 Conectar com o Google")

    goog = load_google_tokens()

    # ── Já conectado ──────────────────────────────────────────────────────────
    if goog.get("access_token"):
        nome_g  = goog.get("_google_name","")
        email_g = goog.get("_google_email","")
        pic_g   = goog.get("_google_pic","")

        st.markdown('<div class="alert alert-g">✅ Sua conta Google está conectada ao Hub.</div>',
                    unsafe_allow_html=True)
        st.write("")

        if pic_g:
            gc1, gc2 = st.columns([1, 6])
            gc1.image(pic_g, width=60)
            gc2.markdown(f"**{nome_g}**  \n{email_g}")
        else:
            st.markdown(f"**{nome_g}** · {email_g}")

        st.write("")
        st.markdown("**O que está disponível agora:**")
        st.markdown("- 📧 Enviar e-mails para geradores direto do card de atividade\n"
                    "- 📅 Ver suas reuniões do Google Calendar no Dashboard\n"
                    "- 📌 Criar eventos no Calendar a partir de atividades programadas")

        st.write("")
        if st.button("🔌 Desconectar conta Google", key="btn_google_disc"):
            save_google_tokens({})
            st.success("Conta desconectada.")
            st.rerun()
        return

    # ── Não conectado: UI simples para o analista ─────────────────────────────
    _, centro, _ = st.columns([1, 3, 1])
    with centro:
        st.write("")
        st.markdown(
            '<div style="text-align:center;padding:2rem 1rem">'
            '<div style="font-size:48px;margin-bottom:1rem">📅</div>'
            '<h2 style="color:#33001A;margin-bottom:.5rem">Conectar com o Google</h2>'
            '<p style="color:#5A3040;font-size:14px;margin-bottom:2rem">'
            'Clique no botão abaixo, faça login com sua conta Google<br>'
            'e autorize o acesso. É rápido e seguro.</p>'
            '</div>',
            unsafe_allow_html=True)

        # Verifica se o client_id foi configurado
        try:
            _ = st.secrets["google"]["client_id"]
            secrets_ok = True
        except Exception:
            secrets_ok = False

        if secrets_ok:
            auth_url = _google_auth_url()
            if auth_url:
                st.markdown(
                    f'<div style="text-align:center">'
                    f'<a href="{auth_url}" target="_self" style="'
                    f'display:inline-flex;align-items:center;gap:10px;'
                    f'background:#F36E21;color:white;padding:.8rem 2rem;'
                    f'border-radius:10px;font-size:16px;font-weight:600;'
                    f'text-decoration:none;box-shadow:0 2px 12px rgba(243,110,33,.35);'
                    f'transition:all .2s">'
                    f'<svg width="20" height="20" viewBox="0 0 48 48">'
                    f'<path fill="#fff" d="M44.5 20H24v8.5h11.8C34.7 33.9 30.1 37 24 37c-7.2 0-13-5.8-13-13s5.8-13 13-13c3.1 0 5.9 1.1 8.1 2.9l6.4-6.4C34.6 4.1 29.6 2 24 2 11.8 2 2 11.8 2 24s9.8 22 22 22c11 0 21-8 21-22 0-1.3-.2-2.7-.5-4z"/>'
                    f'</svg>'
                    f'Conectar com o Google</a></div>',
                    unsafe_allow_html=True)
                st.write("")
                st.caption("🔒 Suas credenciais são gerenciadas pelo Google. "
                           "O Hub Sunne não armazena sua senha.")
            else:
                st.error("Erro ao gerar link de autorização.")
        else:
            # Modo de demonstração / sem secrets configurados
            st.markdown(
                '<div class="alert alert-y">⚙️ <b>Configuração necessária (apenas uma vez, pela equipe técnica):</b><br><br>'
                'Adicione no Streamlit Cloud → Settings → Secrets:<br><br>'
                '<code>[google]<br>'
                'client_id     = "SEU_CLIENT_ID.apps.googleusercontent.com"<br>'
                'client_secret = "SEU_CLIENT_SECRET"</code><br><br>'
                'Após isso, qualquer analista poderá conectar com 1 clique.</div>',
                unsafe_allow_html=True)

        st.write("")
        st.markdown(
            '<div style="text-align:center;font-size:12px;color:#999">'
            'Após clicar, você será redirecionado para a página do Google.<br>'
            'Selecione sua conta de trabalho e clique em <b>Permitir</b>.'
            '</div>',
            unsafe_allow_html=True)


def main():
    st.markdown(SUNNE_CSS, unsafe_allow_html=True)

    # ── SESSÃO PERSISTENTE ────────────────────────────────────────────────────
    # Usa st.session_state que persiste enquanto a aba do browser estiver aberta.
    # Se o usuário estiver salvo no session_state, não pede login novamente.
    # O login só é limpo quando clica em "Sair".

    if "user" not in st.session_state:
        # ── Tela de login premium ─────────────────────────────────────────────
        st.markdown("""
        <style>
        [data-testid="stAppViewContainer"] {
          background: radial-gradient(ellipse 80% 60% at 50% -10%,
            rgba(243,110,33,.08) 0%, transparent 60%),
            var(--bg) !important;
          min-height: 100vh;
        }
        [data-testid="stSidebar"] { display: none !important; }
        .block-container { padding: 0 !important; max-width: 100% !important; }
        </style>
        """, unsafe_allow_html=True)

        _, centro, _ = st.columns([1, 1.1, 1])
        with centro:
            st.markdown("<div style='height:8vh'></div>", unsafe_allow_html=True)

            # Card de login
            st.markdown("""
            <div style="background:var(--surface);border:1px solid var(--border-m);
              border-radius:20px;padding:2.8rem 2.5rem 0;
              box-shadow:0 20px 60px rgba(0,0,0,.7),0 0 80px rgba(243,110,33,.05);">
              <div style="display:flex;align-items:center;gap:12px;margin-bottom:.5rem">
                <div style="width:38px;height:38px;
                  background:linear-gradient(135deg,#F36E21,#D45E14);
                  border-radius:9px;display:flex;align-items:center;
                  justify-content:center;font-size:18px;
                  box-shadow:0 0 20px rgba(243,110,33,.35)">☀️</div>
                <div>
                  <div style="font-family:'Syne',sans-serif;font-size:1.25rem;
                    color:#F0F0FF;font-weight:700;line-height:1;letter-spacing:-.02em">Sunne</div>
                  <div style="font-size:9px;color:#606080;letter-spacing:.14em;
                    text-transform:uppercase;margin-top:2px">Hub Operacional</div>
                </div>
              </div>
              <div style="height:1px;background:rgba(255,255,255,.06);margin:1.4rem 0"></div>
              <p style="font-size:12.5px;color:#606080;margin-bottom:1.6rem;line-height:1.6">
                Plataforma de gestão operacional de geração distribuída.
              </p>
            </div>
            """, unsafe_allow_html=True)

            with st.form("login_form"):
                st.markdown('<p style="font-size:10.5px;font-weight:600;color:#606080;'
                            'text-transform:uppercase;letter-spacing:.10em;margin-bottom:4px">'
                            'E-mail</p>', unsafe_allow_html=True)
                email_inp = st.text_input("", placeholder="seu@sunne.com.br",
                                          label_visibility="collapsed", key="login_email")
                st.markdown('<p style="font-size:10.5px;font-weight:600;color:#606080;'
                            'text-transform:uppercase;letter-spacing:.10em;margin-bottom:4px;'
                            'margin-top:.8rem">Senha</p>', unsafe_allow_html=True)
                senha_inp = st.text_input("", type="password",
                                          label_visibility="collapsed", key="login_senha")
                st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)
                btn_entrar = st.form_submit_button("Entrar", use_container_width=True)

            if btn_entrar:
                u = authenticate(email_inp, senha_inp)
                if u:
                    st.session_state["user"] = u
                    st.session_state.setdefault("page", "dash")
                    st.rerun()
                else:
                    st.markdown(
                        '<div class="alert alert-r" style="margin-top:.6rem">'
                        '⛔ E-mail ou senha incorretos.</div>',
                        unsafe_allow_html=True)

            st.markdown(
                '<p style="text-align:center;font-size:10.5px;color:#3A3A58;'
                'margin-top:1.4rem;font-family:Geist Mono,monospace">© 2026 Sunne · Hub Operacional v12</p>',
                unsafe_allow_html=True)
        return

    # ── OAuth Google: troca code na URL ──────────────────────────────────────
    params = st.query_params
    if "code" in params and not load_google_tokens().get("access_token"):
        with st.spinner("🔗 Conectando com o Google…"):
            tokens = _google_exchange_code(params["code"])
        if tokens.get("access_token"):
            save_google_tokens(tokens)
            st.query_params.clear()
            nome_g = tokens.get("_google_name","")
            st.toast(f"✅ Google conectado! Olá, {nome_g}! 👋", icon="🎉")
            st.rerun()
        else:
            st.query_params.clear()
            st.error("Não foi possível conectar. Tente novamente em Integrações.")

    # ── Auto-agendamentos ─────────────────────────────────────────────────────
    if ROBOT_DISPONIVEL and _robot.verificar_agendamento():
        if not st.session_state.get("robot_rodando", False):
            st.session_state["robot_rodando"]      = True
            st.session_state["robot_auto_trigger"] = True
            st.session_state["page"]               = "automacao"

    if RATEIO_BOT_OK and _rateio_bot.verificar_agendamento_rateio():
        if not st.session_state.get("rateio_bot_rodando", False):
            st.session_state["rateio_bot_rodando"] = True
            st.session_state["page"]               = "rateio"

    render_sidebar()

    page  = st.session_state.get("page","dash")

    # ── Protege páginas admin ─────────────────────────────────────────────────
    user_role = st.session_state["user"].get("role","user")
    if page == "gestor_dash" and user_role != "admin":
        page = "dash"

    pages = {
        "dash":         page_dashboard,
        "gestor_dash":  page_gestor_dashboard,
        "geradores":    page_geradores,
        "usinas":       page_usinas,
        "atividades":   page_atividades,
        "geracao":      page_geracao,
        "backoffice":   page_backoffice,
        "rateio":       page_rateio,
        "faturamento":  page_faturamento,
        "bi_analise":   page_bi_analise,
        "automacao":    page_automacao,
        "integracoes":  page_integracoes,
        "medicao":      page_medicao_cruzamento,
        "auditoria":    page_auditoria,
    }
    pages.get(page, page_dashboard)()


if __name__ == "__main__":
    main()
